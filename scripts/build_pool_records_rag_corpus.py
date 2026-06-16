#!/usr/bin/env python3
"""Build RAG-ready markdown from Memphis Pool Drive records."""
from __future__ import annotations

import argparse
import hashlib
import json
import re
import subprocess
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any


WORK_ORDER_MARKERS = {
    "job_name": ("JOB NAME/LOCATION:", "JOB NAME:", "LOCATION:"),
    "assigned_to": ("ASSIGNED TO:",),
    "measure_install": ("MEASURE/INSTALL:",),
    "work_order": ("W.O. #", "WO #", "WORK ORDER"),
    "customer_number": ("CUST #",),
}


def slugify(value: str, fallback: str = "record") -> str:
    clean = re.sub(r"[^A-Za-z0-9._-]+", "-", value).strip("-")
    return (clean[:120] or fallback).lower()


def sha16(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()[:16]


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def flatten_rows(rows: list[list[Any]]) -> list[str]:
    lines = []
    for row in rows:
        values = [cell_text(v) for v in row]
        if any(values):
            lines.append(" | ".join(v for v in values if v))
    return lines


def find_value_after_label(rows: list[list[Any]], labels: tuple[str, ...]) -> str:
    lowered = tuple(label.lower() for label in labels)
    for row in rows:
        values = [cell_text(v) for v in row]
        for idx, value in enumerate(values):
            if value.lower() in lowered:
                for candidate in values[idx + 1 :]:
                    if candidate:
                        return candidate
    return ""


def extract_location_block(rows: list[list[Any]]) -> tuple[str, str, str]:
    for row_index, row in enumerate(rows):
        values = [cell_text(v) for v in row]
        for idx, value in enumerate(values):
            if value.upper() == "JOB NAME/LOCATION:":
                name = next((v for v in values[idx + 1 :] if v), "")
                address = ""
                city_state = ""
                for lookahead in rows[row_index + 1 : row_index + 5]:
                    look_values = [cell_text(v) for v in lookahead]
                    tail = [v for v in look_values[idx + 1 :] if v]
                    if tail and not address:
                        address = tail[0]
                    elif tail and not city_state:
                        city_state = tail[0]
                        break
                return name, address, city_state
    return "", "", ""


def extract_excel(path: Path) -> dict[str, Any]:
    suffix = path.suffix.lower()
    sheets: dict[str, list[list[Any]]] = {}
    if suffix == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet in workbook.worksheets:
                rows = []
                for row in sheet.iter_rows(max_row=80, max_col=16, values_only=True):
                    values = [cell_text(v) for v in row]
                    if any(values):
                        rows.append(values)
                sheets[sheet.title] = rows
        finally:
            workbook.close()
    elif suffix == ".xls":
        import xlrd

        workbook = xlrd.open_workbook(str(path))
        for sheet in workbook.sheets():
            rows = []
            for row_idx in range(min(sheet.nrows, 80)):
                values = [cell_text(sheet.cell_value(row_idx, col_idx)) for col_idx in range(min(sheet.ncols, 16))]
                if any(values):
                    rows.append(values)
            sheets[sheet.name] = rows
    else:
        raise ValueError(f"Unsupported spreadsheet: {path}")

    first_rows = next(iter(sheets.values()), [])
    job_name, address, city_state = extract_location_block(first_rows)
    record = {
        "parser": suffix.lstrip("."),
        "sheets": sheets,
        "job_name": job_name,
        "job_address": address,
        "job_city_state": city_state,
        "assigned_to": find_value_after_label(first_rows, WORK_ORDER_MARKERS["assigned_to"]),
        "work_order_type": find_value_after_label(first_rows, WORK_ORDER_MARKERS["measure_install"]),
        "work_order_number": find_value_after_label(first_rows, WORK_ORDER_MARKERS["work_order"]),
        "customer_number": find_value_after_label(first_rows, WORK_ORDER_MARKERS["customer_number"]),
        "body_lines": flatten_rows(first_rows[:60]),
    }
    return record


def extract_pdf(path: Path) -> dict[str, Any]:
    text = ""
    parser = "pypdf"
    try:
        from pypdf import PdfReader

        reader = PdfReader(str(path))
        text = "\n\n".join((page.extract_text() or "").strip() for page in reader.pages).strip()
    except Exception as exc:
        parser = f"pypdf_failed:{type(exc).__name__}"
    if not text:
        try:
            with tempfile.TemporaryDirectory() as temp_dir:
                prefix = Path(temp_dir) / "page"
                subprocess.run(
                    ["pdftoppm", "-png", "-r", "300", str(path), str(prefix)],
                    check=True,
                    capture_output=True,
                    text=True,
                )
                image = Path(f"{prefix}-1.png")
                ocr_base = Path(temp_dir) / "ocr"
                subprocess.run(
                    ["tesseract", str(image), str(ocr_base), "--psm", "6"],
                    check=True,
                    capture_output=True,
                    text=True,
                )
                text = Path(f"{ocr_base}.txt").read_text(encoding="utf-8", errors="ignore").strip()
                parser = "tesseract"
        except Exception as exc:
            parser = f"{parser};tesseract_failed:{type(exc).__name__}"
    return {"parser": parser, "text": text}


def source_year(path: Path) -> str:
    for part in path.parts:
        match = re.search(r"20\d{2}", part)
        if match:
            return match.group(0)
    return "unknown-year"


def record_kind(path: Path) -> str:
    if path.suffix.lower() == ".pdf":
        return "bill"
    folder_parts = [part.lower() for part in path.parent.parts]
    if any("bill" in part for part in folder_parts):
        return "bill"
    return "work_order"


def markdown_for_record(source_root: Path, path: Path, extracted: dict[str, Any]) -> tuple[str, str, str]:
    rel = path.relative_to(source_root)
    kind = record_kind(path)
    year = source_year(path)
    title_base = extracted.get("job_name") or path.stem
    title = f"Pool {'Bill' if kind == 'bill' else 'Work Order'}: {title_base}"

    lines = [
        f"# {title}",
        "",
        f"Record kind: {kind}",
        f"Source year: {year}",
        f"Source file: {path}",
        f"Source relative path: {rel}",
        f"Source filename: {path.name}",
        f"Source hash: {sha16(path)}",
        f"Built at: {datetime.now(timezone.utc).isoformat()}",
        "",
    ]

    if kind == "work_order":
        lines.extend(
            [
                "## Extracted Work Order",
                "",
                f"Customer/job name: {extracted.get('job_name') or ''}",
                f"Job location: {extracted.get('job_address') or ''}",
                f"City/state: {extracted.get('job_city_state') or ''}",
                f"Work order number: {extracted.get('work_order_number') or ''}",
                f"Customer number: {extracted.get('customer_number') or ''}",
                f"Assigned to: {extracted.get('assigned_to') or ''}",
                f"Work order type: {extracted.get('work_order_type') or ''}",
                "",
                "## Search Text",
                "",
            ]
        )
        lines.extend(extracted.get("body_lines") or [])
    else:
        text = extracted.get("text") or ""
        lines.extend(
            [
                "## Extracted Bill Text",
                "",
                text[:12000] if text else "[No extractable PDF text; use source PDF for visual review.]",
            ]
        )

    out_rel = Path("drive-bills" if kind == "bill" else "drive-work-orders") / year / f"{slugify(path.stem)}.md"
    return kind, year, "\n".join(lines).rstrip() + "\n", str(out_rel)


def build(source_root: Path, output_root: Path) -> dict[str, Any]:
    output_root.mkdir(parents=True, exist_ok=True)
    records = []
    errors = []
    supported = {".xls", ".xlsx", ".pdf"}
    for path in sorted(source_root.rglob("*")):
        if not path.is_file() or path.suffix.lower() not in supported:
            continue
        try:
            if path.suffix.lower() in {".xls", ".xlsx"}:
                extracted = extract_excel(path)
            else:
                extracted = extract_pdf(path)
            kind, year, markdown, out_rel = markdown_for_record(source_root, path, extracted)
            out_path = output_root / out_rel
            out_path.parent.mkdir(parents=True, exist_ok=True)
            out_path.write_text(markdown, encoding="utf-8")
            records.append(
                {
                    "source_file": str(path),
                    "output_file": str(out_path),
                    "kind": kind,
                    "year": year,
                    "suffix": path.suffix.lower(),
                    "source_hash": sha16(path),
                }
            )
        except Exception as exc:
            errors.append({"source_file": str(path), "error": f"{type(exc).__name__}: {exc}"})
    return {"records": records, "errors": errors}


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-root", required=True)
    parser.add_argument("--output-root", default="/Users/stephengodman/godman-pool-data/jobs")
    parser.add_argument("--manifest", default="/Users/stephengodman/CodeX/work-artifacts/trello-mcp/audits/pool-records-rag-corpus-manifest-2026-06-15.json")
    args = parser.parse_args()

    payload = build(Path(args.source_root).expanduser(), Path(args.output_root).expanduser())
    manifest = Path(args.manifest).expanduser()
    manifest.parent.mkdir(parents=True, exist_ok=True)
    manifest.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(f"records: {len(payload['records'])}")
    print(f"errors: {len(payload['errors'])}")
    print(f"manifest: {manifest}")
    return 0 if not payload["errors"] else 2


if __name__ == "__main__":
    raise SystemExit(main())
