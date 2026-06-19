#!/usr/bin/env python3
"""Stephen-native Trello MCP server.

Exposes the Memphis Pool live board and Artesian backup board through MCP
without storing Trello secrets in MCP config. Credentials are read from
environment variables first, then Stephen's local credential profile.
"""
from __future__ import annotations

import argparse
import mimetypes
import os
import re
import hashlib
import json
import math
import shutil
import sqlite3
import subprocess
import sys
import time
import zipfile
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any
from urllib.parse import unquote_plus, urlparse
from xml.etree import ElementTree

import requests
from mcp.server.fastmcp import FastMCP


BASE_URL = "https://api.trello.com/1"
DEFAULT_TIMEOUT = 30
ARTIFACT_ROOT = Path("/Users/stephengodman/CodeX/work-artifacts/trello-mcp")
DEFAULT_PHOTOS_LIBRARY_PATH = Path("/Users/stephengodman/Pictures/Photos Library.photoslibrary")
PHOTO_MATCH_ROOT = ARTIFACT_ROOT / "photo-card-matches"
PHOTOS_INTAKE_ROOT = ARTIFACT_ROOT / "photos-intake"
PHOTO_GEOCODE_CACHE_PATH = PHOTO_MATCH_ROOT / "geocode-cache.json"
PHOTOKIT_EXPORTER_SOURCE = Path("/Users/stephengodman/CodeX/mcp_servers/photos_photokit_export.swift")
PHOTOKIT_EXPORTER_APP = Path("/Users/stephengodman/Applications/CodeXPhotoKitExport.app")
PHOTOKIT_EXPORTER_BIN = PHOTOKIT_EXPORTER_APP / "Contents" / "MacOS" / "photos_photokit_export"
OSXPHOTOS_BIN = Path(os.getenv("OSXPHOTOS_BIN", "/opt/homebrew/bin/osxphotos"))
OSXPHOTOS_AUTHORIZED_PYTHON = Path(
    os.getenv(
        "OSXPHOTOS_AUTHORIZED_PYTHON",
        "/opt/homebrew/Cellar/python@3.12/3.12.12_2/Frameworks/Python.framework/Versions/3.12/bin/python3.12",
    )
)
RAG_CLI = Path("/Users/stephengodman/000_AI/bin/rag")
POOL_JOBS_ROOT = Path("/Users/stephengodman/godman-pool-data/jobs")
LOCAL_TRELLO_CONFIG_PATH = Path("/Users/stephengodman/.trello-mcp/config.json")
TRELLO_MCP_LAUNCHER_PATH = Path("/Users/stephengodman/CodeX/mcp_servers/trello_mcp_launcher.sh")
CODEX_CONFIG_PATH = Path("/Users/stephengodman/.codex/config.toml")
CLAUDE_CONFIG_PATH = Path("/Users/stephengodman/.claude.json")
SUPPORTED_PHOTO_EXTENSIONS = {".heic", ".heif", ".jpg", ".jpeg", ".png", ".tif", ".tiff"}
SUPPORTED_SHEET_EXTENSIONS = {".xls", ".xlsx", ".csv"}
SUPPORTED_WORK_ORDER_EXTENSIONS = SUPPORTED_SHEET_EXTENSIONS | {".pdf", ".txt"}
MATCH_STOPWORDS = {
    "pool",
    "liner",
    "measure",
    "measures",
    "install",
    "reset",
    "memphis",
    "artesian",
    "covered",
    "need",
    "needs",
    "the",
    "and",
    "for",
    "http",
    "https",
    "www",
    "com",
    "trello",
    "cards",
    "card",
    "attachments",
    "attachment",
    "download",
    "xls",
    "xlsx",
    "csv",
    "pdf",
    "jpg",
    "jpeg",
    "drive",
    "road",
    "street",
    "lane",
    "court",
    "cove",
    "circle",
    "avenue",
    "trail",
    "terrace",
    "parkway",
    "boulevard",
}
KNOWN_BOARDS = {
    "memphis": {
        "id": "6a2f3f45d1e13218cfc620c3",
        "name": "Memphis Pool - New Admin 2026-06-14",
        "role": "new default live jobs and billing board",
    },
    "legacy_memphis": {
        "id": "60df29145c9a576f23056516",
        "name": "Memphis Pool",
        "role": "legacy source board; keep for verification/history until archived",
    },
    "artesian": {
        "id": "60431cb6af79f452db557abb",
        "name": "Artesian Pools",
        "role": "backup / synced PWA board",
    },
}
PROFILE_PATHS = [
    Path("/Users/stephengodman/.claude/0stephen-profile/credentials.md"),
    Path("/Users/stephengodman/.claude/stephen-profile/credentials.md"),
]
OP_VAULT = "3i56wtg5jxdvaiz7ksc6bmh65y"
OP_TRELLO_ITEM = "ifrjvey3q2pztbbyv5e7zhd6da"
OP_BIN = os.getenv("OP_BIN") or ("/opt/homebrew/bin/op" if Path("/opt/homebrew/bin/op").exists() else "op")
OP_TIMEOUT_SECONDS = int(os.getenv("TRELLO_OP_TIMEOUT_SECONDS", "8"))
OP_FAILURE_RETRY_SECONDS = int(os.getenv("TRELLO_OP_FAILURE_RETRY_SECONDS", "300"))
_CREDENTIAL_CACHE: tuple[str, str] | None = None
_OP_FAILURE_CACHE: dict[str, dict[str, Any]] = {}

mcp = FastMCP("trello")
mimetypes.add_type("image/heic", ".heic")
mimetypes.add_type("image/heif", ".heif")


class TrelloError(RuntimeError):
    pass


def _profile_text() -> str:
    chunks = []
    for path in PROFILE_PATHS:
        if path.exists():
            chunks.append(path.read_text(errors="ignore"))
    return "\n\n".join(chunks)


def _local_trello_config() -> dict[str, Any]:
    if not LOCAL_TRELLO_CONFIG_PATH.exists():
        return {}
    try:
        data = json.loads(LOCAL_TRELLO_CONFIG_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}
    return data if isinstance(data, dict) else {}


def _file_status(path: Path) -> dict[str, Any]:
    exists = path.exists()
    return {
        "path": str(path),
        "exists": exists,
        "mode": oct(path.stat().st_mode & 0o777) if exists else None,
        "size": path.stat().st_size if exists else None,
    }


def _text_file_contains(path: Path, needle: str) -> bool:
    if not path.exists():
        return False
    try:
        return needle in path.read_text(errors="ignore")
    except Exception:
        return False


def _runtime_environment_summary() -> dict[str, Any]:
    key = os.getenv("TRELLO_API_KEY") or os.getenv("TRELLO_KEY")
    token = os.getenv("TRELLO_API_TOKEN") or os.getenv("TRELLO_TOKEN")
    return {
        "key_present": bool(key),
        "token_present": bool(token),
        "complete": bool(key and token),
        "disable_op_fallback": _op_fallback_disabled(),
        "op_run_timeout_seconds": os.getenv("TRELLO_OP_RUN_TIMEOUT_SECONDS"),
    }


def _launcher_config_summary() -> dict[str, Any]:
    launcher = _file_status(TRELLO_MCP_LAUNCHER_PATH)
    launcher["executable"] = bool(TRELLO_MCP_LAUNCHER_PATH.exists() and os.access(TRELLO_MCP_LAUNCHER_PATH, os.X_OK))
    config_needle = str(TRELLO_MCP_LAUNCHER_PATH)
    return {
        "launcher": launcher,
        "codex_config": {
            **_file_status(CODEX_CONFIG_PATH),
            "references_launcher": _text_file_contains(CODEX_CONFIG_PATH, config_needle),
        },
        "claude_config": {
            **_file_status(CLAUDE_CONFIG_PATH),
            "references_launcher": _text_file_contains(CLAUDE_CONFIG_PATH, config_needle),
        },
    }


def _configured_boards() -> dict[str, dict[str, str]]:
    boards = json.loads(json.dumps(KNOWN_BOARDS))
    config = _local_trello_config()
    config_keys = {
        "memphis": "boardId",
        "legacy_memphis": "legacyMemphisBoardId",
        "artesian": "artesianBoardId",
    }
    for alias, key in config_keys.items():
        value = str(config.get(key) or "").strip()
        if value:
            boards.setdefault(alias, {})["id"] = value
            boards[alias]["id_source"] = str(LOCAL_TRELLO_CONFIG_PATH)
    return boards


def _trello_section(text: str) -> str:
    match = re.search(r"(?ims)^#{2,4}\s+Trello\b(.*?)(?=^#{2,4}\s+|\Z)", text)
    return match.group(1) if match else text


def _extract_labeled_value(text: str, labels: list[str]) -> str | None:
    for label in labels:
        pattern = rf"(?im)^\s*(?:[-*]\s*)?(?:\*\*)?{re.escape(label)}(?:\*\*)?\s*[:=](?:\*\*)?\s*(.+?)\s*$"
        match = re.search(pattern, text)
        if match:
            value = match.group(1).strip().strip("`").strip()
            if value and "[redacted]" not in value.lower():
                return value
    return None


def _op_fallback_disabled() -> bool:
    return os.getenv("TRELLO_DISABLE_OP_FALLBACK", "").strip().lower() in {"1", "true", "yes", "on"}


def _op_failure_cache_snapshot() -> dict[str, Any]:
    now = time.monotonic()
    snapshot = {}
    for field, cached in _OP_FAILURE_CACHE.items():
        age = round(now - float(cached.get("cached_at_monotonic") or now), 3)
        remaining = max(0.0, round(OP_FAILURE_RETRY_SECONDS - age, 3))
        snapshot[field] = {
            "status": cached.get("status"),
            "cached": True,
            "age_seconds": age,
            "retry_after_seconds": remaining,
        }
    return snapshot


def _cache_op_failure(result: dict[str, Any]) -> None:
    if result.get("ok"):
        _OP_FAILURE_CACHE.pop(str(result.get("field") or ""), None)
        return
    cacheable_statuses = {"timeout", "op_not_found", "error", "nonzero_exit", "empty"}
    if result.get("status") in cacheable_statuses:
        cached = {k: v for k, v in result.items() if k != "value"}
        cached["cached_at_monotonic"] = time.monotonic()
        _OP_FAILURE_CACHE[str(result.get("field") or "")] = cached


def _cached_op_failure(field: str) -> dict[str, Any] | None:
    cached = _OP_FAILURE_CACHE.get(field)
    if not cached:
        return None
    age = time.monotonic() - float(cached.get("cached_at_monotonic") or 0)
    if age >= OP_FAILURE_RETRY_SECONDS:
        _OP_FAILURE_CACHE.pop(field, None)
        return None
    result = {k: v for k, v in cached.items() if k != "cached_at_monotonic"}
    result.update(
        {
            "ok": False,
            "field": field,
            "status": f"cached_{cached.get('status')}",
            "cached": True,
            "age_seconds": round(age, 3),
            "retry_after_seconds": max(0.0, round(OP_FAILURE_RETRY_SECONDS - age, 3)),
        }
    )
    return result


def _op_field_result(field: str, *, use_failure_cache: bool = True) -> dict[str, Any]:
    if _op_fallback_disabled():
        return {"ok": False, "field": field, "status": "disabled_by_env", "elapsed_seconds": 0.0}
    if use_failure_cache:
        cached = _cached_op_failure(field)
        if cached:
            return cached
    started = time.monotonic()
    try:
        result = subprocess.run(
            [
                OP_BIN,
                "item",
                "get",
                OP_TRELLO_ITEM,
                "--vault",
                OP_VAULT,
                "--field",
                field,
                "--reveal",
            ],
            check=False,
            capture_output=True,
            text=True,
            timeout=OP_TIMEOUT_SECONDS,
        )
    except subprocess.TimeoutExpired:
        result = {"ok": False, "field": field, "status": "timeout", "elapsed_seconds": round(time.monotonic() - started, 3)}
        _cache_op_failure(result)
        return result
    except FileNotFoundError:
        result = {"ok": False, "field": field, "status": "op_not_found", "elapsed_seconds": round(time.monotonic() - started, 3)}
        _cache_op_failure(result)
        return result
    except Exception as exc:
        result = {
            "ok": False,
            "field": field,
            "status": "error",
            "error_type": type(exc).__name__,
            "elapsed_seconds": round(time.monotonic() - started, 3),
        }
        _cache_op_failure(result)
        return result
    elapsed = round(time.monotonic() - started, 3)
    if result.returncode != 0:
        detail = (result.stderr or result.stdout or "").strip()
        op_result = {
            "ok": False,
            "field": field,
            "status": "nonzero_exit",
            "returncode": result.returncode,
            "detail": detail[:180],
            "elapsed_seconds": elapsed,
        }
        _cache_op_failure(op_result)
        return op_result
    value = result.stdout.strip()
    op_result = {
        "ok": bool(value),
        "field": field,
        "status": "ok" if value else "empty",
        "value": value or None,
        "elapsed_seconds": elapsed,
    }
    _cache_op_failure(op_result)
    return op_result


def _op_field(field: str) -> str | None:
    result = _op_field_result(field)
    return result.get("value") if result.get("ok") else None


def _credential_source_summary(*, try_op: bool = False) -> dict[str, Any]:
    env_key = os.getenv("TRELLO_API_KEY") or os.getenv("TRELLO_KEY")
    env_token = os.getenv("TRELLO_API_TOKEN") or os.getenv("TRELLO_TOKEN")
    profile_path = next((path for path in PROFILE_PATHS if path.exists()), None)
    profile_text = _profile_text()
    section = _trello_section(profile_text)
    profile_key = _extract_labeled_value(section, ["API Key", "Trello API Key", "Key"])
    profile_token = _extract_labeled_value(section, ["API Token", "Trello API Token", "Token"])
    summary: dict[str, Any] = {
        "cache_loaded": _CREDENTIAL_CACHE is not None,
        "environment": {
            "key_present": bool(env_key),
            "token_present": bool(env_token),
            "complete": bool(env_key and env_token),
        },
        "profile": {
            "path": str(profile_path) if profile_path else None,
            "exists": profile_path is not None,
            "trello_section_found": bool(profile_text and section != profile_text),
            "key_present": bool(profile_key),
            "token_present": bool(profile_token),
            "complete": bool(profile_key and profile_token),
        },
        "onepassword": {
            "item_id": OP_TRELLO_ITEM,
            "vault_id": OP_VAULT,
            "timeout_seconds": OP_TIMEOUT_SECONDS,
            "failure_retry_seconds": OP_FAILURE_RETRY_SECONDS,
            "disabled_by_env": _op_fallback_disabled(),
            "failure_cache": _op_failure_cache_snapshot(),
            "checked": False,
        },
    }
    if try_op:
        op_checks = {}
        for field in ("API Key", "Token"):
            result = _op_field_result(field)
            op_checks[field] = {k: v for k, v in result.items() if k != "value"}
        summary["onepassword"].update(
            {
                "checked": True,
                "fields": op_checks,
                "complete": all(item.get("ok") for item in op_checks.values()),
            }
        )
    return summary


def _credential_recommendation(summary: dict[str, Any], *, try_op: bool) -> str:
    if summary["environment"]["complete"]:
        return "Live Trello operations can use TRELLO_API_KEY/TRELLO_API_TOKEN from the MCP runtime environment."
    if summary["profile"]["complete"]:
        return "Live Trello operations can use the local profile credentials."
    if summary["onepassword"].get("complete"):
        return "Live Trello operations can use the configured 1Password item."

    onepassword = summary.get("onepassword") or {}
    fields = onepassword.get("fields") or {}
    statuses = {str(item.get("status") or "") for item in fields.values()}
    cached = onepassword.get("failure_cache") or {}
    cached_statuses = {str(item.get("status") or "") for item in cached.values()}
    combined_statuses = statuses | cached_statuses

    if onepassword.get("disabled_by_env"):
        return (
            "No complete env/profile credentials found, and TRELLO_DISABLE_OP_FALLBACK is enabled. "
            "Set TRELLO_API_KEY and TRELLO_API_TOKEN in the MCP runtime for live Trello operations."
        )
    if any("timeout" in status for status in combined_statuses):
        return (
            "1Password Trello lookup timed out. Set TRELLO_API_KEY and TRELLO_API_TOKEN in the MCP runtime "
            "to make live Trello operations reliable."
        )
    if try_op:
        return (
            "No complete Trello credential source found. Add API Key and API Token to the profile, "
            "or set TRELLO_API_KEY and TRELLO_API_TOKEN in the MCP runtime."
        )
    return (
        "No complete env/profile credentials found in fast mode. Set TRELLO_API_KEY and TRELLO_API_TOKEN "
        "in the MCP runtime, or run diagnostics with try_onepassword=True to test 1Password."
    )


def _credentials() -> tuple[str, str]:
    global _CREDENTIAL_CACHE
    if _CREDENTIAL_CACHE:
        return _CREDENTIAL_CACHE

    key = os.getenv("TRELLO_API_KEY") or os.getenv("TRELLO_KEY")
    token = os.getenv("TRELLO_API_TOKEN") or os.getenv("TRELLO_TOKEN")

    if key and token:
        _CREDENTIAL_CACHE = (key, token)
        return key, token

    section = _trello_section(_profile_text())
    key = key or _extract_labeled_value(section, ["API Key", "Trello API Key", "Key"])
    token = token or _extract_labeled_value(section, ["API Token", "Trello API Token", "Token"])
    key = key or _op_field("API Key")
    token = token or _op_field("Token")

    if not key or not token:
        raise TrelloError(
            "Trello credentials not found. Set TRELLO_API_KEY/TRELLO_API_TOKEN "
            "or add API Key and API Token under the Trello section of the local profile. "
            "If 1Password is slow, set TRELLO_DISABLE_OP_FALLBACK=1 to fail fast."
        )

    _CREDENTIAL_CACHE = (key, token)
    return key, token


def _board_id(board: str | None = None) -> str:
    key = (board or "memphis").strip()
    boards = _configured_boards()
    if key in boards:
        return boards[key]["id"]
    return key


def _request(method: str, endpoint: str, *, params: dict[str, Any] | None = None) -> Any:
    key, token = _credentials()
    req_params = {"key": key, "token": token}
    if params:
        req_params.update({k: v for k, v in params.items() if v is not None})

    response = requests.request(
        method,
        f"{BASE_URL}/{endpoint.lstrip('/')}",
        params=req_params,
        timeout=DEFAULT_TIMEOUT,
    )
    if not response.ok:
        raise TrelloError(f"Trello {method} {endpoint} failed: {response.status_code} {response.text[:300]}")
    if response.status_code == 204 or not response.text:
        return {"ok": True}
    return response.json()


def _request_files(
    method: str,
    endpoint: str,
    *,
    params: dict[str, Any] | None = None,
    files: dict[str, Any] | None = None,
) -> Any:
    key, token = _credentials()
    req_params = {"key": key, "token": token}
    if params:
        req_params.update({k: v for k, v in params.items() if v is not None})

    response = requests.request(
        method,
        f"{BASE_URL}/{endpoint.lstrip('/')}",
        params=req_params,
        files=files,
        timeout=DEFAULT_TIMEOUT,
    )
    if not response.ok:
        raise TrelloError(f"Trello {method} {endpoint} failed: {response.status_code} {response.text[:300]}")
    if response.status_code == 204 or not response.text:
        return {"ok": True}
    return response.json()


def _download_request(url: str) -> bytes:
    key, token = _credentials()
    headers = {
        "Authorization": f'OAuth oauth_consumer_key="{key}", oauth_token="{token}"',
        "Accept": "*/*",
    }
    response = requests.get(url, headers=headers, timeout=DEFAULT_TIMEOUT)
    if not response.ok:
        response = requests.get(url, params={"key": key, "token": token}, timeout=DEFAULT_TIMEOUT)
    if not response.ok:
        raise TrelloError(f"Trello attachment download failed: {response.status_code} {response.text[:300]}")
    return response.content


def _card_summary(card: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": card.get("id"),
        "name": card.get("name"),
        "shortUrl": card.get("shortUrl"),
        "idList": card.get("idList"),
        "closed": card.get("closed"),
        "due": card.get("due"),
        "labels": [label.get("name") or label.get("color") for label in card.get("labels", [])],
    }


def _card_with_board_context(card: dict[str, Any], *, list_name: str | None = None) -> dict[str, Any]:
    return {
        "id": card.get("id"),
        "name": card.get("name"),
        "url": card.get("url") or card.get("shortUrl"),
        "shortUrl": card.get("shortUrl"),
        "idList": card.get("idList"),
        "list": list_name,
    }


def _find_card_by_query(query: str, board: str = "memphis") -> dict[str, Any]:
    q = (query or "").strip().lower()
    if not q:
        raise TrelloError("card query is required")
    lists = get_lists(board)
    list_names = {item["id"]: item["name"] for item in lists}
    cards = _request(
        "GET",
        f"boards/{_board_id(board)}/cards",
        params={"filter": "open", "fields": "id,name,url,shortUrl,idList,closed"},
    )
    matches = [card for card in cards if q in str(card.get("name") or "").lower()]
    exact = [card for card in matches if str(card.get("name") or "").strip().lower() == q]
    if len(exact) == 1:
        return _card_with_board_context(exact[0], list_name=list_names.get(exact[0].get("idList")))
    if len(matches) == 1:
        return _card_with_board_context(matches[0], list_name=list_names.get(matches[0].get("idList")))
    if not matches:
        raise TrelloError(f"No open card matching {query!r} on board {board!r}")
    preview = [
        {
            "id": card.get("id"),
            "name": card.get("name"),
            "list": list_names.get(card.get("idList"), "UNKNOWN"),
            "url": card.get("url") or card.get("shortUrl"),
        }
        for card in matches[:8]
    ]
    raise TrelloError(f"Multiple cards match {query!r}: {json.dumps(preview)}")


def _safe_name(name: str) -> str:
    clean = re.sub(r"[^A-Za-z0-9._-]+", "-", name or "artifact").strip("-")
    return clean[:120] or "artifact"


def _attachment_download_url(card_id: str, attachment: dict[str, Any]) -> str:
    url = attachment.get("url") or ""
    parsed = urlparse(url)
    path = parsed.path
    if "/download/" in path and parsed.netloc.endswith("trello.com"):
        return f"https://api.trello.com{path}"
    if parsed.netloc == "api.trello.com" and "/download/" in path:
        return url
    return f"{BASE_URL}/cards/{card_id}/attachments/{attachment['id']}/download/{_safe_name(attachment.get('name', 'attachment'))}"


def _sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def _artifact_dir(kind: str) -> Path:
    path = ARTIFACT_ROOT / kind
    path.mkdir(parents=True, exist_ok=True)
    return path


def _read_json_file(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return default


def _write_json_file(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, sort_keys=True), encoding="utf-8")


APPLE_PHOTOS_EPOCH = datetime(2001, 1, 1, tzinfo=timezone.utc)


def _apple_time_to_iso(value: Any) -> str | None:
    if value is None:
        return None
    try:
        return (APPLE_PHOTOS_EPOCH + timedelta(seconds=float(value))).astimezone().isoformat(timespec="seconds")
    except Exception:
        return None


def _datetime_to_apple_seconds(value: datetime) -> float:
    if value.tzinfo is None:
        value = value.replace(tzinfo=timezone.utc)
    return (value.astimezone(timezone.utc) - APPLE_PHOTOS_EPOCH).total_seconds()


def _looks_like_photos_library(path: Path) -> bool:
    return path.suffix == ".photoslibrary" or (path / "database" / "Photos.sqlite").exists()


def _photos_db_path(library_path: Path) -> Path:
    db = library_path / "database" / "Photos.sqlite"
    if not db.exists():
        raise TrelloError(f"Photos database not found: {db}")
    return db


def _photos_original_path(library_path: Path, directory: Any, filename: Any) -> Path:
    return library_path / "originals" / str(directory or "") / str(filename or "")


def _valid_photo_coordinate(lat: Any, lng: Any) -> bool:
    try:
        lat_f = float(lat)
        lng_f = float(lng)
    except Exception:
        return False
    if not (-90.0 <= lat_f <= 90.0 and -180.0 <= lng_f <= 180.0):
        return False
    # Apple Photos uses -180/-180 as a no-location sentinel in some libraries.
    if abs(lat_f + 180.0) < 0.0001 and abs(lng_f + 180.0) < 0.0001:
        return False
    if abs(lat_f) < 0.0001 and abs(lng_f) < 0.0001:
        return False
    return True


def _cover_candidate_score(photo: dict[str, Any]) -> int:
    width = int(photo.get("width") or 0)
    height = int(photo.get("height") or 0)
    size = int(photo.get("size") or 0)
    score = 0
    if width and height:
        score += 35 if width >= height else 15
        ratio = max(width, height) / max(1, min(width, height))
        if 1.1 <= ratio <= 2.4:
            score += 10
        score += min(30, int((width * height) / 1_000_000))
    score += min(25, int(size / 1_000_000))
    return score


def _photos_library_assets(
    library_path: Path,
    *,
    days_back: int = 60,
    max_files: int = 500,
    require_local: bool = False,
) -> list[dict[str, Any]]:
    library_path = library_path.expanduser()
    db = _photos_db_path(library_path)
    where = [
        "ZTRASHEDSTATE = 0",
        "ZHIDDEN = 0",
        "ZFILENAME IS NOT NULL",
        "ZKIND = 0",
        "ZLATITUDE IS NOT NULL",
        "ZLONGITUDE IS NOT NULL",
    ]
    params: list[Any] = []
    if days_back > 0:
        since = _datetime_to_apple_seconds(datetime.now(timezone.utc) - timedelta(days=days_back))
        where.append("ZDATECREATED >= ?")
        params.append(since)
    query = f"""
        SELECT ZUUID, ZDIRECTORY, ZFILENAME, ZDATECREATED, ZLATITUDE, ZLONGITUDE,
               ZWIDTH, ZHEIGHT, ZUNIFORMTYPEIDENTIFIER, ZCLOUDLOCALSTATE
        FROM ZASSET
        WHERE {' AND '.join(where)}
        ORDER BY ZDATECREATED DESC
        LIMIT ?
    """
    params.append(max(1, min(int(max_files), 5000)))
    rows: list[dict[str, Any]] = []
    conn = sqlite3.connect(f"file:{db}?mode=ro", uri=True, timeout=5)
    conn.row_factory = sqlite3.Row
    try:
        for row in conn.execute(query, params):
            if not _valid_photo_coordinate(row["ZLATITUDE"], row["ZLONGITUDE"]):
                continue
            original_path = _photos_original_path(library_path, row["ZDIRECTORY"], row["ZFILENAME"])
            suffix = original_path.suffix.lower()
            if suffix not in SUPPORTED_PHOTO_EXTENSIONS:
                continue
            local_exists = original_path.exists()
            if require_local and not local_exists:
                continue
            item = {
                "photo_id": row["ZUUID"],
                "source": "photos_library",
                "source_library": str(library_path),
                "path": str(original_path),
                "filename": row["ZFILENAME"],
                "local_exists": local_exists,
                "size": original_path.stat().st_size if local_exists else None,
                "sha256": _sha256(original_path) if local_exists else None,
                "taken_at": _apple_time_to_iso(row["ZDATECREATED"]),
                "latitude": float(row["ZLATITUDE"]),
                "longitude": float(row["ZLONGITUDE"]),
                "has_gps": True,
                "width": int(row["ZWIDTH"] or 0),
                "height": int(row["ZHEIGHT"] or 0),
                "uti": row["ZUNIFORMTYPEIDENTIFIER"],
                "cloud_local_state": row["ZCLOUDLOCALSTATE"],
            }
            item["cover_candidate_score"] = _cover_candidate_score(item)
            rows.append(item)
    finally:
        conn.close()
    return rows


def _photos_sample(
    row: sqlite3.Row,
    original_path: Path,
    *,
    local_exists: bool,
    has_gps: bool,
    include_paths: bool,
    include_coordinates: bool,
) -> dict[str, Any]:
    sample = {
        "photo_id": row["ZUUID"],
        "filename": row["ZFILENAME"],
        "taken_at": _apple_time_to_iso(row["ZDATECREATED"]),
        "extension": original_path.suffix.lower(),
        "local_exists": local_exists,
        "has_gps": has_gps,
        "width": int(row["ZWIDTH"] or 0),
        "height": int(row["ZHEIGHT"] or 0),
        "uti": row["ZUNIFORMTYPEIDENTIFIER"],
        "cloud_local_state": row["ZCLOUDLOCALSTATE"],
    }
    if local_exists:
        sample["size"] = original_path.stat().st_size
    if include_paths:
        sample["source_path"] = str(original_path)
    if include_coordinates and has_gps:
        sample["latitude"] = round(float(row["ZLATITUDE"]), 6)
        sample["longitude"] = round(float(row["ZLONGITUDE"]), 6)
    return sample


def _photos_library_intake_status(
    library_path: Path,
    *,
    days_back: int = 365,
    max_assets: int = 10000,
    sample_limit: int = 20,
    include_paths: bool = False,
    include_coordinates: bool = False,
) -> dict[str, Any]:
    library_path = library_path.expanduser()
    db = _photos_db_path(library_path)
    where = [
        "ZTRASHEDSTATE = 0",
        "ZHIDDEN = 0",
        "ZFILENAME IS NOT NULL",
        "ZKIND = 0",
    ]
    params: list[Any] = []
    if days_back > 0:
        since = _datetime_to_apple_seconds(datetime.now(timezone.utc) - timedelta(days=days_back))
        where.append("ZDATECREATED >= ?")
        params.append(since)
    query = f"""
        SELECT ZUUID, ZDIRECTORY, ZFILENAME, ZDATECREATED, ZLATITUDE, ZLONGITUDE,
               ZWIDTH, ZHEIGHT, ZUNIFORMTYPEIDENTIFIER, ZCLOUDLOCALSTATE
        FROM ZASSET
        WHERE {' AND '.join(where)}
        ORDER BY ZDATECREATED DESC
        LIMIT ?
    """
    limit = max(1, min(int(max_assets), 50000))
    params.append(limit)

    counts = {
        "assets_scanned": 0,
        "supported_image_assets": 0,
        "unsupported_extension_assets": 0,
        "local_originals": 0,
        "missing_originals": 0,
        "gps_assets": 0,
        "gps_local_originals": 0,
        "gps_missing_originals": 0,
        "no_gps_assets": 0,
        "local_original_bytes": 0,
    }
    extension_counts: dict[str, int] = {}
    cloud_local_state_counts: dict[str, int] = {}
    samples = {
        "gps_local_ready": [],
        "gps_missing_original": [],
        "local_without_gps": [],
    }
    newest_taken_at: str | None = None
    oldest_taken_at: str | None = None

    conn = sqlite3.connect(f"file:{db}?mode=ro", uri=True, timeout=5)
    conn.row_factory = sqlite3.Row
    try:
        for row in conn.execute(query, params):
            counts["assets_scanned"] += 1
            original_path = _photos_original_path(library_path, row["ZDIRECTORY"], row["ZFILENAME"])
            suffix = original_path.suffix.lower()
            extension_counts[suffix or "<none>"] = extension_counts.get(suffix or "<none>", 0) + 1
            cloud_key = str(row["ZCLOUDLOCALSTATE"])
            cloud_local_state_counts[cloud_key] = cloud_local_state_counts.get(cloud_key, 0) + 1
            if suffix not in SUPPORTED_PHOTO_EXTENSIONS:
                counts["unsupported_extension_assets"] += 1
                continue

            counts["supported_image_assets"] += 1
            local_exists = original_path.exists()
            has_gps = _valid_photo_coordinate(row["ZLATITUDE"], row["ZLONGITUDE"])
            taken_at = _apple_time_to_iso(row["ZDATECREATED"])
            if taken_at:
                newest_taken_at = newest_taken_at or taken_at
                oldest_taken_at = taken_at

            if local_exists:
                counts["local_originals"] += 1
                counts["local_original_bytes"] += original_path.stat().st_size
            else:
                counts["missing_originals"] += 1

            if has_gps:
                counts["gps_assets"] += 1
                if local_exists:
                    counts["gps_local_originals"] += 1
                else:
                    counts["gps_missing_originals"] += 1
            else:
                counts["no_gps_assets"] += 1

            if sample_limit <= 0:
                continue
            sample = _photos_sample(
                row,
                original_path,
                local_exists=local_exists,
                has_gps=has_gps,
                include_paths=include_paths,
                include_coordinates=include_coordinates,
            )
            if has_gps and local_exists and len(samples["gps_local_ready"]) < sample_limit:
                samples["gps_local_ready"].append(sample)
            elif has_gps and not local_exists and len(samples["gps_missing_original"]) < sample_limit:
                samples["gps_missing_original"].append(sample)
            elif local_exists and not has_gps and len(samples["local_without_gps"]) < sample_limit:
                samples["local_without_gps"].append(sample)
    finally:
        conn.close()

    recommendations = []
    if counts["gps_local_originals"]:
        recommendations.append("Run preview_photo_card_matches against local originals for immediate Trello-safe matches.")
    if counts["gps_missing_originals"]:
        recommendations.append("Use a native PhotoKit export/download worker or Photos 'Download Originals to this Mac' before relying on browser/iCloud.com automation.")
    if not counts["gps_local_originals"] and not counts["gps_missing_originals"]:
        recommendations.append("No GPS-ready pool-photo candidates found in this slice; widen days_back or import/download more phone originals.")

    return {
        "ok": True,
        "mode": "photos_intake_status",
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "source": {
            "photos_library_path": str(library_path),
            "database_path": str(db),
            "days_back": days_back,
            "max_assets": limit,
        },
        "counts": counts,
        "extension_counts": dict(sorted(extension_counts.items())),
        "cloud_local_state_counts": dict(sorted(cloud_local_state_counts.items())),
        "date_range": {"newest_taken_at": newest_taken_at, "oldest_taken_at": oldest_taken_at},
        "samples": samples,
        "recommendations": recommendations,
        "safety": {
            "read_only": True,
            "trello_writes": False,
            "photos_writes": False,
            "icloud_web_automation": False,
            "raw_gps_returned": include_coordinates,
            "source_paths_returned": include_paths,
            "secrets_returned": False,
        },
    }


def _compile_photokit_exporter() -> dict[str, Any]:
    if not PHOTOKIT_EXPORTER_SOURCE.exists():
        return {"ok": False, "error": "exporter_source_missing", "source": str(PHOTOKIT_EXPORTER_SOURCE)}
    macos_dir = PHOTOKIT_EXPORTER_APP / "Contents" / "MacOS"
    macos_dir.mkdir(parents=True, exist_ok=True)
    info_plist = PHOTOKIT_EXPORTER_APP / "Contents" / "Info.plist"
    info_plist.write_text(
        """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE plist PUBLIC "-//Apple//DTD PLIST 1.0//EN" "http://www.apple.com/DTDs/PropertyList-1.0.dtd">
<plist version="1.0">
<dict>
  <key>CFBundleDevelopmentRegion</key>
  <string>en</string>
  <key>CFBundleExecutable</key>
  <string>photos_photokit_export</string>
  <key>CFBundleIdentifier</key>
  <string>com.godmanautomations.codex.photokitexport</string>
  <key>CFBundleInfoDictionaryVersion</key>
  <string>6.0</string>
  <key>CFBundleName</key>
  <string>CodeX PhotoKit Export</string>
  <key>CFBundlePackageType</key>
  <string>APPL</string>
  <key>CFBundleShortVersionString</key>
  <string>1.0</string>
  <key>CFBundleVersion</key>
  <string>1</string>
  <key>NSPhotoLibraryUsageDescription</key>
  <string>CodeX uses Photos access to export Stephen's pool-job originals for Trello matching.</string>
  <key>NSPhotoLibraryAddUsageDescription</key>
  <string>CodeX does not add to Photos; this exists only to satisfy macOS Photos privacy metadata.</string>
</dict>
</plist>
""",
        encoding="utf-8",
    )
    started = time.monotonic()
    result = subprocess.run(
        [
            "swiftc",
            "-framework",
            "Photos",
            "-framework",
            "UniformTypeIdentifiers",
            str(PHOTOKIT_EXPORTER_SOURCE),
            "-o",
            str(PHOTOKIT_EXPORTER_BIN),
        ],
        capture_output=True,
        text=True,
        timeout=60,
    )
    sign_result = None
    if result.returncode == 0:
        sign_result = subprocess.run(
            ["codesign", "--force", "--deep", "--sign", "-", str(PHOTOKIT_EXPORTER_APP)],
            capture_output=True,
            text=True,
            timeout=30,
        )
    return {
        "ok": result.returncode == 0 and (sign_result is None or sign_result.returncode == 0),
        "source": str(PHOTOKIT_EXPORTER_SOURCE),
        "app": str(PHOTOKIT_EXPORTER_APP),
        "binary": str(PHOTOKIT_EXPORTER_BIN),
        "bundle_identifier": "com.godmanautomations.codex.photokitexport",
        "elapsed_seconds": round(time.monotonic() - started, 3),
        "stderr": (result.stderr or "").strip()[:1000] or None,
        "codesign_stderr": (sign_result.stderr or "").strip()[:1000] if sign_result else None,
    }


def _photos_missing_original_assets(
    library_path: Path,
    *,
    days_back: int = 365,
    max_assets: int = 10000,
    limit: int = 100,
    include_paths: bool = False,
    include_coordinates: bool = False,
) -> list[dict[str, Any]]:
    library_path = library_path.expanduser()
    db = _photos_db_path(library_path)
    where = [
        "ZTRASHEDSTATE = 0",
        "ZHIDDEN = 0",
        "ZFILENAME IS NOT NULL",
        "ZKIND = 0",
        "ZLATITUDE IS NOT NULL",
        "ZLONGITUDE IS NOT NULL",
    ]
    params: list[Any] = []
    if days_back > 0:
        since = _datetime_to_apple_seconds(datetime.now(timezone.utc) - timedelta(days=days_back))
        where.append("ZDATECREATED >= ?")
        params.append(since)
    query = f"""
        SELECT ZUUID, ZDIRECTORY, ZFILENAME, ZDATECREATED, ZLATITUDE, ZLONGITUDE,
               ZWIDTH, ZHEIGHT, ZUNIFORMTYPEIDENTIFIER, ZCLOUDLOCALSTATE
        FROM ZASSET
        WHERE {' AND '.join(where)}
        ORDER BY ZDATECREATED DESC
        LIMIT ?
    """
    params.append(max(1, min(int(max_assets), 50000)))
    export_limit = max(1, min(int(limit), 5000))
    assets: list[dict[str, Any]] = []
    conn = sqlite3.connect(f"file:{db}?mode=ro", uri=True, timeout=5)
    conn.row_factory = sqlite3.Row
    try:
        for row in conn.execute(query, params):
            if len(assets) >= export_limit:
                break
            if not _valid_photo_coordinate(row["ZLATITUDE"], row["ZLONGITUDE"]):
                continue
            original_path = _photos_original_path(library_path, row["ZDIRECTORY"], row["ZFILENAME"])
            if original_path.suffix.lower() not in SUPPORTED_PHOTO_EXTENSIONS:
                continue
            if original_path.exists():
                continue
            assets.append(
                _photos_sample(
                    row,
                    original_path,
                    local_exists=False,
                    has_gps=True,
                    include_paths=include_paths,
                    include_coordinates=include_coordinates,
                )
            )
    finally:
        conn.close()
    return assets


def _run_photokit_exporter(
    *,
    manifest_path: Path,
    output_dir: Path,
    limit: int,
    dry_run: bool,
    timeout_seconds: int,
) -> dict[str, Any]:
    if not PHOTOKIT_EXPORTER_BIN.exists():
        compile_result = _compile_photokit_exporter()
        if not compile_result.get("ok"):
            return {"ok": False, "compile": compile_result}
    run_limit = max(1, min(int(limit), 5000))
    requested_timeout = max(30, int(timeout_seconds))
    if dry_run:
        helper_asset_timeout = min(30, requested_timeout)
    else:
        helper_asset_timeout = max(30, min(180, requested_timeout // max(1, min(run_limit, 5))))
    outer_timeout = max(60, helper_asset_timeout * min(run_limit, 25) + 45)
    outer_timeout = min(1800, max(outer_timeout, min(requested_timeout + 30, 1800)))
    command = [
        "open",
        "-W",
        "-n",
        str(PHOTOKIT_EXPORTER_APP),
        "--args",
        "--manifest",
        str(manifest_path),
        "--output-dir",
        str(output_dir),
        "--result-json",
        str(manifest_path.parent / ("photokit-export-dry-run-result.json" if dry_run else "photokit-export-result.json")),
        "--limit",
        str(run_limit),
        "--timeout-seconds",
        str(helper_asset_timeout),
    ]
    command.append("--dry-run" if dry_run else "--apply")
    try:
        result = subprocess.run(command, capture_output=True, text=True, timeout=outer_timeout)
    except subprocess.TimeoutExpired:
        subprocess.run(["pkill", "-f", str(PHOTOKIT_EXPORTER_BIN)], capture_output=True, text=True, timeout=5)
        return {
            "ok": False,
            "error": "photokit_exporter_timeout",
            "message": "PhotoKit helper did not finish. macOS Photos privacy permission may be waiting or blocked.",
            "process": {
                "returncode": None,
                "stderr": None,
                "outer_timeout_seconds": outer_timeout,
                "helper_asset_timeout_seconds": helper_asset_timeout,
                "command_redacted": [
                    "open",
                    "-W",
                    "-n",
                    str(PHOTOKIT_EXPORTER_APP),
                    "--args",
                    "--manifest",
                    str(manifest_path),
                    "--output-dir",
                    str(output_dir),
                    "--limit",
                    str(limit),
                    "--dry-run" if dry_run else "--apply",
                ],
            },
        }
    result_json_path = manifest_path.parent / ("photokit-export-dry-run-result.json" if dry_run else "photokit-export-result.json")
    try:
        if result_json_path.exists():
            payload = json.loads(result_json_path.read_text(encoding="utf-8") or "{}")
        else:
            payload = json.loads(result.stdout or "{}")
    except Exception:
        payload = {"ok": False, "error": "invalid_exporter_json", "stdout": (result.stdout or "")[:1000]}
    payload.setdefault("process", {})
    payload["process"].update(
        {
            "returncode": result.returncode,
            "stderr": (result.stderr or "").strip()[:1000] or None,
            "outer_timeout_seconds": outer_timeout,
            "helper_asset_timeout_seconds": helper_asset_timeout,
            "command_redacted": [
                str(PHOTOKIT_EXPORTER_BIN),
                "--manifest",
                str(manifest_path),
                "--output-dir",
                str(output_dir),
                "--limit",
                str(limit),
                "--dry-run" if dry_run else "--apply",
            ],
        }
    )
    return payload


def _osxphotos_command_base() -> list[str]:
    if OSXPHOTOS_AUTHORIZED_PYTHON.exists():
        return [str(OSXPHOTOS_AUTHORIZED_PYTHON), "-m", "osxphotos"]
    return [str(OSXPHOTOS_BIN)]


def _run_osxphotos_exporter(
    *,
    manifest: dict[str, Any],
    manifest_path: Path,
    output_dir: Path,
    limit: int,
    dry_run: bool,
    timeout_seconds: int,
) -> dict[str, Any]:
    photo_ids = [str(item) for item in manifest.get("photo_ids") or [] if item]
    if not photo_ids:
        return {"ok": False, "error": "manifest_has_no_photo_ids"}
    output_dir.mkdir(parents=True, exist_ok=True)
    run_limit = max(1, min(int(limit), len(photo_ids), 5000))
    uuid_file = manifest_path.parent / "osxphotos-uuid-list.txt"
    uuid_file.write_text("\n".join(photo_ids[:run_limit]) + "\n", encoding="utf-8")

    command = [
        *_osxphotos_command_base(),
        "export",
        str(output_dir),
        "--library",
        str(Path(str(manifest.get("source", {}).get("photos_library_path") or DEFAULT_PHOTOS_LIBRARY_PATH)).expanduser()),
        "--uuid-from-file",
        str(uuid_file),
        "--download-missing",
        "--use-photokit",
        "--limit",
        str(run_limit),
        "--verbose",
    ]
    if dry_run:
        command.append("--dry-run")

    started = time.monotonic()
    try:
        result = subprocess.run(
            command,
            capture_output=True,
            text=True,
            timeout=max(60, int(timeout_seconds)),
        )
        timed_out = False
    except subprocess.TimeoutExpired as exc:
        result = None
        timed_out = True
        stdout = exc.stdout.decode(errors="replace") if isinstance(exc.stdout, bytes) else (exc.stdout or "")
        stderr = exc.stderr.decode(errors="replace") if isinstance(exc.stderr, bytes) else (exc.stderr or "")
    else:
        stdout = result.stdout or ""
        stderr = result.stderr or ""

    exported_files = []
    if output_dir.exists():
        for path in sorted(output_dir.rglob("*")):
            if path.is_file() and path.name != ".DS_Store":
                exported_files.append(
                    {
                        "path": str(path),
                        "filename": path.name,
                        "size": path.stat().st_size,
                    }
                )

    ok = bool(result and result.returncode == 0 and not timed_out)
    auth_denied = "could not get authorization to access Photos library" in stdout + stderr
    return {
        "ok": ok,
        "backend": "osxphotos",
        "dry_run": dry_run,
        "elapsed_seconds": round(time.monotonic() - started, 3),
        "requested_photo_ids": len(photo_ids),
        "limit": run_limit,
        "uuid_file": str(uuid_file),
        "output_dir": str(output_dir),
        "exported_files": exported_files[:200],
        "exported_file_count": len(exported_files),
        "process": {
            "returncode": None if result is None else result.returncode,
            "timed_out": timed_out,
            "stdout_tail": stdout[-4000:] if stdout else None,
            "stderr_tail": stderr[-4000:] if stderr else None,
            "authorization_denied": auth_denied,
            "command_redacted": [
                *_osxphotos_command_base(),
                "export",
                str(output_dir),
                "--library",
                "<photos-library>",
                "--uuid-from-file",
                str(uuid_file),
                "--download-missing",
                "--use-photokit",
                "--limit",
                str(run_limit),
                "--dry-run" if dry_run else "--apply",
            ],
        },
        "safety": {
            "photos_writes": False,
            "local_file_writes": not dry_run,
            "trello_writes": False,
            "icloud_network_access_allowed": not dry_run,
            "secrets_returned": False,
        },
    }


def _osxphotos_status() -> dict[str, Any]:
    command = [*_osxphotos_command_base(), "--version"]
    try:
        result = subprocess.run(command, capture_output=True, text=True, timeout=20)
    except Exception as exc:
        return {
            "ok": False,
            "command_redacted": command,
            "error_type": type(exc).__name__,
            "error": str(exc)[:500],
        }
    return {
        "ok": result.returncode == 0,
        "command_redacted": command,
        "returncode": result.returncode,
        "stdout": (result.stdout or "").strip()[:1000] or None,
        "stderr": (result.stderr or "").strip()[:1000] or None,
    }


def _known_city_context(text: str) -> str | None:
    value = text.lower()
    city_map = [
        ("collierville", "Collierville, TN"),
        ("germantown", "Germantown, TN"),
        ("cordova", "Cordova, TN"),
        ("bartlett", "Bartlett, TN"),
        ("lakeland", "Lakeland, TN"),
        ("arlington", "Arlington, TN"),
        ("eads", "Eads, TN"),
        ("millington", "Millington, TN"),
        ("memphis", "Memphis, TN"),
        ("olive branch", "Olive Branch, MS"),
        ("southaven", "Southaven, MS"),
        ("hernando", "Hernando, MS"),
        ("horn lake", "Horn Lake, MS"),
        ("west memphis", "West Memphis, AR"),
        ("holly springs", "Holly Springs, MS"),
    ]
    for needle, city in city_map:
        if needle in value:
            return city
    return None


def _address_queries(address: str, context: str = "") -> list[str]:
    base = re.sub(r"\s+", " ", address).strip(" ,")
    if not base:
        return []
    queries = []
    city = _known_city_context(f"{address}\n{context}")
    has_city_or_state = bool(re.search(r"\b(TN|MS|AR|Memphis|Collierville|Germantown|Cordova|Bartlett|Lakeland|Arlington|Eads|Millington|Southaven|Hernando|Olive Branch|West Memphis)\b", base, re.I))
    if has_city_or_state:
        queries.append(base)
    else:
        if city:
            queries.append(f"{base}, {city}")
        queries.append(f"{base}, Memphis, TN")
    queries.append(base)
    deduped = []
    seen = set()
    for query in queries:
        key = query.lower()
        if key not in seen:
            seen.add(key)
            deduped.append(query)
    return deduped


def _geocode_cache_key(query: str) -> str:
    return re.sub(r"\s+", " ", query).strip().lower()


def _census_geocode(query: str) -> dict[str, Any] | None:
    response = requests.get(
        "https://geocoding.geo.census.gov/geocoder/locations/onelineaddress",
        params={"address": query, "benchmark": "Public_AR_Current", "format": "json"},
        timeout=10,
    )
    if not response.ok:
        return None
    matches = ((response.json().get("result") or {}).get("addressMatches") or [])
    if not matches:
        return None
    first = matches[0]
    coords = first.get("coordinates") or {}
    if coords.get("x") is None or coords.get("y") is None:
        return None
    return {
        "query": query,
        "latitude": float(coords["y"]),
        "longitude": float(coords["x"]),
        "matched_address": first.get("matchedAddress"),
        "source": "us_census_geocoder",
    }


def _geocode_address(address: str, context: str = "", *, allow_network: bool = True) -> dict[str, Any] | None:
    cache = _read_json_file(PHOTO_GEOCODE_CACHE_PATH, {})
    for query in _address_queries(address, context=context):
        key = _geocode_cache_key(query)
        cached = cache.get(key)
        if isinstance(cached, dict):
            if cached.get("latitude") is not None and cached.get("longitude") is not None:
                return cached
            if cached.get("status") == "no_match":
                continue
        if not allow_network:
            continue
        try:
            result = _census_geocode(query)
        except Exception as exc:
            cache[key] = {"query": query, "status": "error", "error": str(exc)[:300], "updated_at": datetime.now().isoformat(timespec="seconds")}
            _write_json_file(PHOTO_GEOCODE_CACHE_PATH, cache)
            continue
        if result:
            result["updated_at"] = datetime.now().isoformat(timespec="seconds")
            cache[key] = result
            _write_json_file(PHOTO_GEOCODE_CACHE_PATH, cache)
            return result
        cache[key] = {"query": query, "status": "no_match", "updated_at": datetime.now().isoformat(timespec="seconds")}
        _write_json_file(PHOTO_GEOCODE_CACHE_PATH, cache)
    return None


def _card_locations_for_photo_match(
    board: str,
    *,
    include_complete: bool = False,
    limit_cards: int = 500,
    geocode_limit: int = 100,
    target_card_query: str = "",
    target_card_id: str = "",
) -> dict[str, Any]:
    trello_filter = "all" if include_complete else "open"
    all_cards = _request(
        "GET",
        f"boards/{_board_id(board)}/cards",
        params={"filter": trello_filter, "fields": "id,name,desc,shortUrl,idList,idBoard,closed,due,dateLastActivity"},
    )
    list_names = {item["id"]: item["name"] for item in get_lists(board=board, filter="all")}
    query = (target_card_query or "").strip()
    wanted_id = (target_card_id or "").strip()
    target_tokens = _match_tokens(query)
    target_filtered = 0
    cards = []
    for card in all_cards:
        list_name = list_names.get(card.get("idList"), "UNKNOWN")
        card_text = f"{card.get('name', '')}\n{card.get('desc', '')}\n{list_name}"
        short_url = str(card.get("shortUrl") or "")
        id_match = bool(wanted_id) and (
            wanted_id == str(card.get("id") or "")
            or wanted_id in short_url
        )
        query_match = (
            not query
            or query.lower() in card_text.lower()
            or (target_tokens and _score_text(target_tokens, card_text) == len(target_tokens))
        )
        if wanted_id and not id_match:
            target_filtered += 1
            continue
        if query and not query_match:
            target_filtered += 1
            continue
        cards.append(card)
    matched_card_count_before_limit = len(cards)
    cards = cards[: max(1, min(int(limit_cards), 1000))]
    locations = []
    skipped = {"no_address": 0, "not_geocoded": 0, "geocode_limit": 0, "target_filtered": target_filtered}
    geocode_attempts = 0
    cache = _read_json_file(PHOTO_GEOCODE_CACHE_PATH, {})
    for card in cards:
        list_name = list_names.get(card.get("idList"), "UNKNOWN")
        text = f"{card.get('name', '')}\n{card.get('desc', '')}\n{list_name}"
        addresses = _extract_address_candidates(text)
        if not addresses:
            skipped["no_address"] += 1
            continue
        geocoded = None
        for address in addresses:
            queries = _address_queries(address, context=text)
            cached_hit = any(_geocode_cache_key(query) in cache for query in queries)
            if not cached_hit and geocode_attempts >= max(0, int(geocode_limit)):
                skipped["geocode_limit"] += 1
                break
            if not cached_hit:
                geocode_attempts += 1
            geocoded = _geocode_address(address, context=text, allow_network=True)
            if geocoded:
                locations.append(
                    {
                        "card_id": card.get("id"),
                        "card_name": card.get("name"),
                        "card_url": card.get("shortUrl"),
                        "list": list_name,
                        "address": address,
                        "geocode": geocoded,
                        "latitude": geocoded.get("latitude"),
                        "longitude": geocoded.get("longitude"),
                        "closed": bool(card.get("closed")),
                    }
                )
                break
        if not geocoded:
            skipped["not_geocoded"] += 1
    return {
        "cards_seen": len(cards),
        "cards_available": len(all_cards),
        "locations": locations,
        "skipped": skipped,
        "geocode_attempts": geocode_attempts,
        "target": {
            "card_query": query or None,
            "card_id": wanted_id or None,
            "matched_card_count_before_limit": matched_card_count_before_limit,
        },
    }


def _photo_confidence(distance_ft: float, second_distance_ft: float | None = None) -> tuple[str, list[str]]:
    warnings = []
    if distance_ft <= 100:
        confidence = "very_strong"
    elif distance_ft <= 250:
        confidence = "likely"
    else:
        confidence = "review"
    if second_distance_ft is not None and second_distance_ft - distance_ft <= 75:
        confidence = "review"
        warnings.append("near_multiple_card_locations")
    return confidence, warnings


def _scan_photo_source(source_path: str, *, days_back: int = 60, max_files: int = 500, require_local: bool = False) -> list[dict[str, Any]]:
    root = Path(source_path or str(DEFAULT_PHOTOS_LIBRARY_PATH)).expanduser()
    if _looks_like_photos_library(root):
        return _photos_library_assets(root, days_back=days_back, max_files=max_files, require_local=require_local)
    photos = []
    for item in scan_pool_photos(str(root), max_files=max_files):
        item.setdefault("source", "local_folder")
        item.setdefault("local_exists", Path(str(item.get("path") or "")).exists())
        if item.get("has_gps") and (not require_local or item.get("local_exists")):
            item["cover_candidate_score"] = _cover_candidate_score(item)
            photos.append(item)
    return photos


def _stage_photo_for_plan(photo: dict[str, Any], card_name: str, plan_dir: Path) -> dict[str, Any]:
    source = Path(str(photo.get("path") or photo.get("source_path") or ""))
    if not source.is_file():
        return {"staged": False, "error": "source_file_missing", "source_path": str(source)}
    card_dir = plan_dir / "files" / _safe_name(card_name)
    card_dir.mkdir(parents=True, exist_ok=True)
    taken = str(photo.get("taken_at") or "unknown-date")[:10]
    uuid_part = str(photo.get("photo_id") or source.stem)[:8]
    target = card_dir / f"{taken}-{uuid_part}-{_safe_name(source.name)}"
    if not target.exists():
        shutil.copy2(source, target)
    return {
        "staged": True,
        "source_path": str(source),
        "staged_path": str(target),
        "sha256": _sha256(target),
        "size": target.stat().st_size,
    }


def _export_missing_originals_for_photo_plan(
    *,
    photos: list[dict[str, Any]],
    photos_library_path: Path,
    plan_dir: Path,
    limit: int,
    timeout_seconds: int,
) -> dict[str, Any]:
    selected = []
    seen: set[str] = set()
    for photo in photos:
        photo_id = str(photo.get("photo_id") or "").strip()
        if not photo_id or photo_id in seen:
            continue
        source = Path(str(photo.get("source_path") or photo.get("path") or ""))
        if source.is_file():
            continue
        selected.append(photo)
        seen.add(photo_id)
        if len(selected) >= max(1, min(int(limit), 5000)):
            break
    if not selected:
        return {
            "ok": True,
            "mode": "photo_plan_missing_original_export",
            "requested_photo_ids": 0,
            "exported_by_photo_id": {},
            "result": None,
        }

    export_dir = plan_dir / "missing-original-export"
    output_dir = export_dir / "exported-originals"
    manifest_path = export_dir / "photokit-export-manifest.json"
    manifest = {
        "ok": True,
        "mode": "photokit_export_manifest",
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "source": {
            "photos_library_path": str(photos_library_path),
            "source": "photo_card_match_plan",
        },
        "photo_ids": [str(item["photo_id"]) for item in selected],
        "assets": [
            {
                "photo_id": item.get("photo_id"),
                "filename": item.get("filename"),
                "taken_at": item.get("taken_at"),
                "width": item.get("width"),
                "height": item.get("height"),
                "cloud_local_state": item.get("cloud_local_state"),
            }
            for item in selected
        ],
        "output_dir": str(output_dir),
        "safety": {
            "photos_writes": False,
            "local_file_writes": True,
            "trello_writes": False,
            "secrets_returned": False,
        },
    }
    _write_json_file(manifest_path, manifest)
    result = _run_photokit_exporter(
        manifest_path=manifest_path,
        output_dir=output_dir,
        limit=len(selected),
        dry_run=False,
        timeout_seconds=timeout_seconds,
    )
    exported_by_photo_id = {}
    for row in result.get("results") or []:
        if row.get("status") != "exported" or not row.get("exported_path"):
            continue
        exported_by_photo_id[str(row.get("photo_id"))] = {
            "exported_path": row.get("exported_path"),
            "size": row.get("size"),
            "uti": row.get("uti"),
            "status": row.get("status"),
        }
    return {
        "ok": result.get("ok"),
        "mode": "photo_plan_missing_original_export",
        "manifest_json_path": str(manifest_path),
        "output_dir": str(output_dir),
        "requested_photo_ids": len(selected),
        "exported_by_photo_id": exported_by_photo_id,
        "exported_count": len(exported_by_photo_id),
        "result": result,
    }


def _load_photo_match_plan(plan_json_path: str) -> dict[str, Any]:
    path = Path(plan_json_path).expanduser()
    if not path.exists():
        raise TrelloError(f"Photo match plan JSON not found: {path}")
    payload = _read_json_file(path, None)
    if not isinstance(payload, dict) or payload.get("mode") != "photo_card_match_plan":
        raise TrelloError("File is not a photo_card_match_plan JSON")
    return payload


def _excel_preview(path: Path, max_rows: int = 40, max_cols: int = 20) -> dict[str, Any]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheets = {}
        for sheet in workbook.worksheets:
            rows = []
            for row in sheet.iter_rows(max_row=max_rows, max_col=max_cols, values_only=True):
                values = ["" if value is None else str(value) for value in row]
                if any(v.strip() for v in values):
                    rows.append(values)
            sheets[sheet.title] = rows
        workbook.close()
        return {"type": "xlsx", "sheets": sheets}

    if suffix == ".xls":
        import xlrd

        workbook = xlrd.open_workbook(path)
        sheets = {}
        for sheet in workbook.sheets():
            rows = []
            for r in range(min(sheet.nrows, max_rows)):
                values = ["" if sheet.cell_value(r, c) is None else str(sheet.cell_value(r, c)) for c in range(min(sheet.ncols, max_cols))]
                if any(v.strip() for v in values):
                    rows.append(values)
            sheets[sheet.name] = rows
        return {"type": "xls", "sheets": sheets}

    if suffix == ".csv":
        import csv

        rows = []
        with path.open(newline="", errors="ignore") as f:
            for i, row in enumerate(csv.reader(f)):
                if i >= max_rows:
                    break
                rows.append(row[:max_cols])
        return {"type": "csv", "sheets": {"csv": rows}}

    raise TrelloError(f"Unsupported sheet type: {path.suffix}")


def _flatten_preview_text(preview: dict[str, Any]) -> str:
    chunks = []
    for sheet, rows in preview.get("sheets", {}).items():
        chunks.append(str(sheet))
        for row in rows:
            chunks.append(" | ".join(str(v) for v in row if str(v).strip()))
    return "\n".join(chunks)


def _extract_address_candidates(text: str) -> list[str]:
    candidates = []
    street_pattern = re.compile(
        r"\b\d{1,6}\s+(?:[A-Za-z0-9'&.-]+\s+){0,8}"
        r"(?:st|street|rd|road|dr|drive|ave|avenue|ln|lane|ct|court|cv|cove|cir|circle|way|pkwy|parkway|blvd|trail|terrace)\b"
        r"(?:\s+(?:north|south|east|west|n|s|e|w|\d{1,5})\b)?",
        re.I,
    )
    for raw_url_address in re.findall(r"address=([^)\"]+)", text, flags=re.I):
        decoded = unquote_plus(raw_url_address).strip(" ,")
        if decoded:
            candidates.append(decoded[:220])
    for line in text.splitlines():
        s = _searchable_fragment(line)
        s = re.sub(r"\s+", " ", s).strip(" -:\t")
        if not s:
            continue
        for match in street_pattern.finditer(s):
            address = match.group(0).strip(" ,")
            tail = s[match.end() : match.end() + 80]
            city_match = re.match(
                r"\s*,?\s*((?:Memphis|Collierville|Germantown|Cordova|Bartlett|Lakeland|Arlington|Eads|Millington|Southaven|Hernando|Olive Branch|West Memphis|Corinth|Holly Springs)(?:,\s*(?:TN|MS|AR))?)\b",
                tail,
                flags=re.I,
            )
            if city_match and city_match.group(1).lower() not in address.lower():
                address = f"{address}, {city_match.group(1)}"
            candidates.append(address[:220])
    deduped = []
    for item in candidates:
        if item.lower() not in {x.lower() for x in deduped}:
            deduped.append(item)
    return deduped[:10]


def _searchable_fragment(value: Any) -> str:
    text = str(value or "")
    text = re.sub(r"\[([^\]]+)\]\([^)]+\)", r"\1", text)
    text = re.sub(r"https?://\S+", " ", text)
    text = text.replace("_", " ")
    return re.sub(r"\s+", " ", text).strip()


def _match_tokens(text: str) -> set[str]:
    tokens = set()
    for token in re.findall(r"[A-Za-z0-9]+", text.lower()):
        if len(token) >= 3 and token not in MATCH_STOPWORDS:
            tokens.add(token)
            if token == "bills":
                tokens.add("bill")
    return tokens


def _score_text(tokens: set[str], text: str) -> int:
    if not tokens:
        return 0
    haystack = _match_tokens(text)
    return len(tokens & haystack)


def _run_rag(args: list[str], timeout: int = 180) -> dict[str, Any]:
    if not RAG_CLI.exists():
        raise TrelloError(f"RAG CLI not found: {RAG_CLI}")
    env = os.environ.copy()
    env.setdefault("RAG_LOCAL_API_RETRY_ATTEMPTS", "1")
    try:
        result = subprocess.run(
            [str(RAG_CLI), *args, "--json"],
            check=False,
            capture_output=True,
            text=True,
            timeout=timeout,
            env=env,
        )
    except subprocess.TimeoutExpired as exc:
        raise TrelloError(f"RAG command timed out after {timeout}s: {' '.join(args)}") from exc
    if result.returncode != 0:
        detail = (result.stderr or result.stdout or "").strip()
        raise TrelloError(f"RAG command failed: {detail[:500]}")
    try:
        payload = json.loads(result.stdout)
    except json.JSONDecodeError as exc:
        raise TrelloError(f"RAG command returned non-JSON output: {result.stdout[:500]}") from exc
    if not payload.get("ok", False):
        raise TrelloError(f"RAG command returned errors: {payload.get('errors')}")
    return payload


def _pool_record_kind(result: dict[str, Any]) -> str:
    path = str(result.get("path") or "").lower()
    text = str(result.get("text") or "").lower()
    if "/drive-bills/" in path or "record kind: bill" in text:
        return "bill"
    if "/drive-work-orders/" in path or "record kind: work_order" in text:
        return "work_order"
    return "unknown"


def _pool_record_year(result: dict[str, Any]) -> str | None:
    path = str(result.get("path") or "")
    text = str(result.get("text") or "")
    for value in (path, text):
        match = re.search(r"\b(20\d{2})\b", value)
        if match:
            return match.group(1)
    return None


def _pool_record_summary(result: dict[str, Any], excerpt_chars: int = 1400) -> dict[str, Any]:
    text = str(result.get("text") or "")
    return {
        "kind": _pool_record_kind(result),
        "year": _pool_record_year(result),
        "score": result.get("score"),
        "distance": result.get("distance"),
        "file": result.get("file"),
        "path": result.get("path"),
        "chunk": result.get("chunk"),
        "text_excerpt": text[: max(200, min(excerpt_chars, 4000))],
    }


def _lexical_pool_record_search(query: str, record_kind: str = "any", max_results: int = 20) -> list[dict[str, Any]]:
    if not POOL_JOBS_ROOT.exists():
        return []
    kind = (record_kind or "any").strip().lower()
    tokens = _match_tokens(query)
    if not tokens:
        return []
    candidates = []
    for path in POOL_JOBS_ROOT.rglob("*.md"):
        result = {
            "text": path.read_text(encoding="utf-8", errors="ignore"),
            "file": path.name,
            "path": str(path),
            "chunk": "1/1",
        }
        result_kind = _pool_record_kind(result)
        if kind != "any" and result_kind != kind:
            continue
        haystack = _match_tokens(f"{path.name}\n{result['text']}")
        overlap = tokens & haystack
        if not overlap:
            continue
        score = min(0.91, 0.70 + len(overlap) * 0.06)
        result["score"] = score
        result["distance"] = round(max(0.0, 1.0 - score), 6)
        result["lexical_overlap"] = sorted(overlap)
        candidates.append(result)
    candidates.sort(key=lambda item: (len(item.get("lexical_overlap") or []), item.get("score") or 0.0), reverse=True)
    return [_pool_record_summary(item) for item in candidates[: max(1, min(max_results, 100))]]


def _label_names(card: dict[str, Any]) -> list[str]:
    names = []
    for label in card.get("labels") or []:
        value = label.get("name") or label.get("color")
        if value:
            names.append(str(value))
    return names


def _card_record_search_text(
    card: dict[str, Any],
    *,
    list_name: str | None = None,
    attachments: list[dict[str, Any]] | None = None,
) -> str:
    parts = [
        str(card.get("name") or ""),
        str(card.get("desc") or ""),
        " ".join(_label_names(card)),
        str(list_name or ""),
    ]
    for candidate in _extract_address_candidates("\n".join(parts)):
        parts.append(candidate)
    for item in attachments or []:
        name = str(item.get("name") or "")
        ext = str(item.get("extension") or "")
        if name:
            parts.append(name)
        if ext:
            parts.append(ext)

    seen = set()
    compact = []
    for part in parts:
        clean = _searchable_fragment(part)
        if not clean:
            continue
        key = clean.lower()
        if key in seen:
            continue
        seen.add(key)
        compact.append(clean)
    return " ".join(compact)[:1200]


def _card_name_anchor_tokens(card: dict[str, Any]) -> set[str]:
    name = _searchable_fragment(card.get("name") or "")
    head = re.split(r"\s+[-–]\s+", name, maxsplit=1)[0]
    ordered = [
        token.lower()
        for token in re.findall(r"[A-Za-z0-9]+", head)
        if len(token) >= 3 and token.lower() not in MATCH_STOPWORDS
    ]
    if not ordered:
        return set()
    if len(ordered) == 1:
        return {ordered[0]}
    return {ordered[-1]}


def _strong_card_tokens(card: dict[str, Any], query: str) -> set[str]:
    name = _searchable_fragment(card.get("name") or "")
    head = re.split(r"\s+[-–]\s+", name, maxsplit=1)[0]
    tokens = _match_tokens(head)
    for candidate in _extract_address_candidates(query):
        tokens.update(_match_tokens(candidate))
    return tokens


def _record_match_detail(
    card_tokens: set[str],
    record: dict[str, Any],
    *,
    strong_tokens: set[str] | None = None,
    name_anchor_tokens: set[str] | None = None,
) -> dict[str, Any]:
    searchable = " ".join(
        str(record.get(key) or "")
        for key in ("file", "path", "text_excerpt")
    )
    record_tokens = _match_tokens(searchable)
    overlap = sorted(card_tokens & record_tokens)
    strong_overlap = sorted((strong_tokens or set()) & record_tokens)
    name_anchor_overlap = sorted((name_anchor_tokens or set()) & record_tokens)
    score = float(record.get("score") or 0.0)
    overlap_count = len(overlap)
    strong_overlap_count = len(strong_overlap)
    has_name_anchor = not name_anchor_tokens or bool(name_anchor_overlap)

    if has_name_anchor and strong_overlap_count >= 3 and score >= 0.78:
        confidence = "exact"
        reason = "name_anchor_plus_multiple_specific_card_tokens"
    elif has_name_anchor and strong_overlap_count >= 2 and score >= 0.84:
        confidence = "exact"
        reason = "name_anchor_and_high_score"
    elif has_name_anchor and strong_overlap_count >= 1 and score >= 0.72:
        confidence = "likely"
        reason = "name_anchor_overlap"
    elif strong_overlap_count >= 2 and score >= 0.72:
        confidence = "review"
        reason = "specific_card_tokens_without_name_anchor"
    elif score >= 0.78 and overlap_count >= 2:
        confidence = "review"
        reason = "good_score_but_missing_name_anchor"
    else:
        confidence = "weak"
        reason = "low_score_or_low_overlap"

    return {
        "confidence": confidence,
        "reason": reason,
        "overlap_tokens": overlap[:20],
        "overlap_count": overlap_count,
        "strong_overlap_tokens": strong_overlap[:20],
        "strong_overlap_count": strong_overlap_count,
        "name_anchor_tokens": sorted(name_anchor_tokens or []),
        "name_anchor_overlap": name_anchor_overlap[:20],
        "score": score,
    }


def _confidence_rank(value: str) -> int:
    return {"exact": 4, "likely": 3, "review": 2, "weak": 1, "no_match": 0}.get(value, 0)


def _record_candidate_preview(candidate: dict[str, Any], excerpt_chars: int = 360) -> dict[str, Any]:
    match = candidate.get("match") or {}
    excerpt = str(candidate.get("text_excerpt") or "")
    return {
        "kind": candidate.get("kind"),
        "year": candidate.get("year"),
        "confidence": match.get("confidence"),
        "reason": match.get("reason"),
        "score": candidate.get("score"),
        "file": candidate.get("file"),
        "path": candidate.get("path"),
        "retrieval_sources": candidate.get("retrieval_sources") or [],
        "overlap_tokens": match.get("overlap_tokens") or [],
        "text_excerpt": excerpt[: max(120, min(excerpt_chars, 1200))],
    }


def _empty_match_buckets() -> dict[str, list[dict[str, Any]]]:
    return {"exact": [], "likely": [], "review": [], "weak": [], "no_match": []}


def _dms_to_decimal(values: Any, ref: str) -> float | None:
    try:
        deg, minute, sec = values
        decimal = float(deg) + float(minute) / 60 + float(sec) / 3600
        return -decimal if ref in {"S", "W"} else decimal
    except Exception:
        return None


def _photo_metadata(path: Path) -> dict[str, Any]:
    from PIL import ExifTags, Image

    try:
        with Image.open(path) as img:
            exif = img.getexif()
            if not exif:
                return {"path": str(path), "filename": path.name, "size": path.stat().st_size, "has_gps": False, "taken_at": None}
            named = {ExifTags.TAGS.get(k, k): v for k, v in exif.items()}
            gps_raw = named.get("GPSInfo")
            gps = {}
            if gps_raw:
                gps = {ExifTags.GPSTAGS.get(k, k): v for k, v in gps_raw.items()}
            lat = _dms_to_decimal(gps.get("GPSLatitude"), gps.get("GPSLatitudeRef")) if gps else None
            lng = _dms_to_decimal(gps.get("GPSLongitude"), gps.get("GPSLongitudeRef")) if gps else None
            return {
                "path": str(path),
                "filename": path.name,
                "size": path.stat().st_size,
                "sha256": _sha256(path),
                "taken_at": named.get("DateTimeOriginal") or named.get("DateTime"),
                "latitude": lat,
                "longitude": lng,
                "has_gps": lat is not None and lng is not None,
            }
    except Exception:
        if path.suffix.lower() not in {".heic", ".heif"}:
            raise
        return {
            "path": str(path),
            "filename": path.name,
            "size": path.stat().st_size,
            "sha256": _sha256(path),
            "taken_at": None,
            "latitude": None,
            "longitude": None,
            "has_gps": False,
            "metadata_note": "HEIC metadata should be read through Photos Library SQLite when possible.",
        }


def _distance_ft(lat1: float, lng1: float, lat2: float, lng2: float) -> float:
    radius_ft = 20902231.0
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lng2 - lng1)
    a = math.sin(dphi / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dlambda / 2) ** 2
    return radius_ft * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


@mcp.tool()
def trello_credential_diagnostics(try_onepassword: bool = False) -> dict[str, Any]:
    """Report non-secret Trello credential source availability.

    By default this does not call 1Password, so it cannot hang on an auth
    prompt. Set try_onepassword=True for a bounded, value-redacted field check.
    """
    summary = _credential_source_summary(try_op=try_onepassword)
    complete_sources = []
    if summary["environment"]["complete"]:
        complete_sources.append("environment")
    if summary["profile"]["complete"]:
        complete_sources.append("profile")
    if summary["onepassword"].get("complete"):
        complete_sources.append("onepassword")
    return {
        "ok": bool(complete_sources) or not try_onepassword,
        "complete_sources": complete_sources,
        "recommendation": _credential_recommendation(summary, try_op=try_onepassword),
        "diagnostics": summary,
        "secrets_returned": False,
    }


@mcp.tool()
def trello_config_diagnostics() -> dict[str, Any]:
    """Report non-secret local Trello MCP config status."""
    config = _local_trello_config()
    expected_keys = {
        "memphis": "boardId",
        "legacy_memphis": "legacyMemphisBoardId",
        "artesian": "artesianBoardId",
    }
    board_sources = {}
    for alias, key in expected_keys.items():
        value = str(config.get(key) or "").strip()
        board_sources[alias] = {
            "config_key": key,
            "present": bool(value),
            "chars": len(value),
            "active_board_id": _board_id(alias),
            "source": str(LOCAL_TRELLO_CONFIG_PATH) if value else "built_in_default",
        }
    return {
        "ok": True,
        "path": str(LOCAL_TRELLO_CONFIG_PATH),
        "exists": LOCAL_TRELLO_CONFIG_PATH.exists(),
        "mode": oct(LOCAL_TRELLO_CONFIG_PATH.stat().st_mode & 0o777) if LOCAL_TRELLO_CONFIG_PATH.exists() else None,
        "contains_credentials": any(
            key.lower() in {"key", "token", "apikey", "apitoken", "api_key", "api_token"}
            for key in config.keys()
        ),
        "board_sources": board_sources,
        "secrets_returned": False,
    }


@mcp.tool()
def trello_runtime_diagnostics(include_live_health: bool = False) -> dict[str, Any]:
    """Report non-secret Trello MCP runtime readiness.

    This is intended as the first operator check in a fresh MCP session. It
    does not call Trello unless include_live_health is true.
    """
    credential_status = trello_credential_diagnostics(False)
    config_status = trello_config_diagnostics()
    runtime_env = _runtime_environment_summary()
    launcher_status = _launcher_config_summary()
    complete_sources = credential_status.get("complete_sources") or []
    blockers = []

    if not runtime_env["complete"] and not complete_sources:
        blockers.append("no_complete_credential_source")
    if not launcher_status["launcher"]["exists"]:
        blockers.append("launcher_missing")
    elif not launcher_status["launcher"]["executable"]:
        blockers.append("launcher_not_executable")
    if not launcher_status["codex_config"]["references_launcher"]:
        blockers.append("codex_config_not_pointing_at_launcher")
    if not launcher_status["claude_config"]["references_launcher"]:
        blockers.append("claude_config_not_pointing_at_launcher")
    if config_status.get("contains_credentials"):
        blockers.append("local_board_config_contains_credential_like_keys")

    live_health: dict[str, Any] = {
        "checked": False,
        "ok": None,
        "error": None,
    }
    if include_live_health:
        try:
            health = trello_health()
            live_health = {
                "checked": True,
                "ok": bool(health.get("ok")),
                "member_present": bool(health.get("member")),
                "known_board_aliases": sorted((health.get("known_boards") or {}).keys()),
            }
        except Exception as exc:
            live_health = {
                "checked": True,
                "ok": False,
                "error": str(exc)[:500],
            }

    recommendation = credential_status.get("recommendation")
    if include_live_health and live_health.get("checked") and not live_health.get("ok"):
        error_text = str(live_health.get("error") or "").lower()
        if "invalid key" in error_text:
            recommendation = (
                "Trello credentials resolve into the runtime, but Trello rejects the API key. "
                "Update the Trello API key/token source before running live packet tools."
            )
        elif "invalid token" in error_text:
            recommendation = (
                "Trello credentials resolve into the runtime, but Trello rejects the token. "
                "Update the Trello API token source before running live packet tools."
            )

    return {
        "ok": not blockers and (not include_live_health or bool(live_health.get("ok"))),
        "blockers": blockers,
        "credential_sources": complete_sources,
        "runtime_environment": runtime_env,
        "launcher": launcher_status,
        "board_config": config_status,
        "live_health": live_health,
        "recommendation": recommendation,
        "safety": {
            "secrets_returned": False,
            "trello_writes": False,
            "google_drive_writes": False,
            "pi5_writes": False,
            "live_trello_read": bool(include_live_health),
        },
    }


@mcp.tool()
def trello_health() -> dict[str, Any]:
    """Verify Trello credentials by reading the authenticated member profile."""
    member = _request("GET", "members/me", params={"fields": "id,username,fullName"})
    return {
        "ok": True,
        "member": {k: member.get(k) for k in ("id", "username", "fullName")},
        "known_boards": _configured_boards(),
    }


@mcp.tool()
def list_known_boards() -> dict[str, Any]:
    """List Stephen's built-in board aliases and IDs."""
    return _configured_boards()


@mcp.tool()
def list_boards(filter: str = "open") -> list[dict[str, Any]]:
    """List boards visible to the authenticated Trello account."""
    boards = _request("GET", "members/me/boards", params={"filter": filter, "fields": "id,name,url,closed"})
    return [{k: b.get(k) for k in ("id", "name", "url", "closed")} for b in boards]


@mcp.tool()
def get_board(board: str = "memphis") -> dict[str, Any]:
    """Read a board by alias ('memphis' or 'artesian') or raw board ID."""
    return _request("GET", f"boards/{_board_id(board)}", params={"fields": "id,name,url,closed,dateLastActivity"})


@mcp.tool()
def get_lists(board: str = "memphis", filter: str = "open") -> list[dict[str, Any]]:
    """List Trello lists on a board."""
    lists = _request("GET", f"boards/{_board_id(board)}/lists", params={"filter": filter, "fields": "id,name,closed,pos"})
    return [{k: item.get(k) for k in ("id", "name", "closed", "pos")} for item in lists]


@mcp.tool()
def get_labels(board: str = "memphis") -> list[dict[str, Any]]:
    """List labels on a board."""
    labels = _request("GET", f"boards/{_board_id(board)}/labels", params={"fields": "id,name,color"})
    return [{k: label.get(k) for k in ("id", "name", "color")} for label in labels]


@mcp.tool()
def get_cards(board: str = "memphis", filter: str = "open", limit: int = 50) -> list[dict[str, Any]]:
    """List cards on a board, summarized for scanning."""
    cards = _request(
        "GET",
        f"boards/{_board_id(board)}/cards",
        params={"filter": filter, "fields": "id,name,shortUrl,idList,closed,due,labels"},
    )
    return [_card_summary(card) for card in cards[: max(1, min(limit, 500))]]


@mcp.tool()
def get_cards_in_list(list_id: str, limit: int = 50) -> list[dict[str, Any]]:
    """List cards in a Trello list."""
    cards = _request(
        "GET",
        f"lists/{list_id}/cards",
        params={"fields": "id,name,shortUrl,idList,closed,due,labels"},
    )
    return [_card_summary(card) for card in cards[: max(1, min(limit, 500))]]


@mcp.tool()
def get_card(card_id: str) -> dict[str, Any]:
    """Read one card with useful fields."""
    return _request(
        "GET",
        f"cards/{card_id}",
        params={"fields": "id,name,desc,shortUrl,idList,idBoard,closed,due,labels,dateLastActivity"},
    )


@mcp.tool()
def index_trello_job_cards(board: str = "memphis", limit: int = 500) -> list[dict[str, Any]]:
    """Index cards with lightweight address candidates for artifact matching."""
    cards = _request(
        "GET",
        f"boards/{_board_id(board)}/cards",
        params={"filter": "open", "fields": "id,name,desc,shortUrl,idList,idBoard,closed,due,dateLastActivity"},
    )
    indexed = []
    for card in cards[: max(1, min(limit, 1000))]:
        text = f"{card.get('name', '')}\n{card.get('desc', '')}"
        indexed.append(
            {
                "id": card.get("id"),
                "name": card.get("name"),
                "shortUrl": card.get("shortUrl"),
                "idList": card.get("idList"),
                "dateLastActivity": card.get("dateLastActivity"),
                "address_candidates": _extract_address_candidates(text),
            }
        )
    return indexed


@mcp.tool()
def search_cards(query: str, board: str = "memphis", limit: int = 25) -> list[dict[str, Any]]:
    """Search cards on one board by name or description text."""
    q = (query or "").strip().lower()
    if not q:
        return []
    cards = _request(
        "GET",
        f"boards/{_board_id(board)}/cards",
        params={"filter": "open", "fields": "id,name,desc,shortUrl,idList,closed,due,labels"},
    )
    hits = [card for card in cards if q in (card.get("name", "") + "\n" + card.get("desc", "")).lower()]
    return [_card_summary(card) for card in hits[: max(1, min(limit, 100))]]


@mcp.tool()
def create_card(list_id: str, name: str, desc: str = "", due: str | None = None, pos: str = "bottom") -> dict[str, Any]:
    """Create a card in a Trello list."""
    return _request("POST", "cards", params={"idList": list_id, "name": name, "desc": desc, "due": due, "pos": pos})


@mcp.tool()
def update_card(card_id: str, name: str | None = None, desc: str | None = None, due: str | None = None) -> dict[str, Any]:
    """Update basic card fields. Omitted fields are left unchanged."""
    return _request("PUT", f"cards/{card_id}", params={"name": name, "desc": desc, "due": due})


@mcp.tool()
def move_card(card_id: str, list_id: str, pos: str = "bottom") -> dict[str, Any]:
    """Move a card to another list."""
    return _request("PUT", f"cards/{card_id}", params={"idList": list_id, "pos": pos})


@mcp.tool()
def archive_card(card_id: str) -> dict[str, Any]:
    """Archive a card instead of deleting it."""
    return _request("PUT", f"cards/{card_id}", params={"closed": "true"})


@mcp.tool()
def add_comment(card_id: str, text: str) -> dict[str, Any]:
    """Add a comment to a Trello card."""
    return _request("POST", f"cards/{card_id}/actions/comments", params={"text": text})


@mcp.tool()
def add_checklist(card_id: str, name: str) -> dict[str, Any]:
    """Add a checklist to a Trello card."""
    return _request("POST", "checklists", params={"idCard": card_id, "name": name})


@mcp.tool()
def add_checklist_item(checklist_id: str, name: str, checked: bool = False) -> dict[str, Any]:
    """Add an item to a Trello checklist."""
    return _request(
        "POST",
        f"checklists/{checklist_id}/checkItems",
        params={"name": name, "checked": str(checked).lower()},
    )


@mcp.tool()
def add_attachment_url(card_id: str, url: str, name: str | None = None) -> dict[str, Any]:
    """Attach a URL to a Trello card."""
    return _request("POST", f"cards/{card_id}/attachments", params={"url": url, "name": name})


@mcp.tool()
def get_card_attachments(card_id: str, filter: str = "false") -> list[dict[str, Any]]:
    """List attachments on a Trello card."""
    attachments = _request(
        "GET",
        f"cards/{card_id}/attachments",
        params={"filter": filter, "fields": "id,name,url,bytes,date,mimeType,isUpload,previews"},
    )
    return [
        {
            "id": item.get("id"),
            "name": item.get("name"),
            "url": item.get("url"),
            "bytes": item.get("bytes"),
            "date": item.get("date"),
            "mimeType": item.get("mimeType"),
            "isUpload": item.get("isUpload"),
            "extension": Path(item.get("name") or "").suffix.lower(),
            "preview_count": len(item.get("previews") or []),
        }
        for item in attachments
    ]


@mcp.tool()
def index_trello_sheet_attachments(board: str = "memphis", limit_cards: int = 500) -> list[dict[str, Any]]:
    """Find XLS/XLSX/CSV attachments already present on board cards."""
    cards = _request(
        "GET",
        f"boards/{_board_id(board)}/cards",
        params={"filter": "open", "fields": "id,name,shortUrl,idList"},
    )
    hits = []
    for card in cards[: max(1, min(limit_cards, 1000))]:
        attachments = get_card_attachments(card["id"])
        for item in attachments:
            if item["extension"] in SUPPORTED_SHEET_EXTENSIONS:
                hits.append(
                    {
                        "card_id": card.get("id"),
                        "card_name": card.get("name"),
                        "card_url": card.get("shortUrl"),
                        "list_id": card.get("idList"),
                        "attachment": item,
                    }
                )
    return hits


@mcp.tool()
def audit_board_artifacts(board: str = "memphis", include_complete: bool = False, limit_cards: int = 500) -> dict[str, Any]:
    """Read-only board artifact audit for cards, addresses, and sheet attachments.

    Defaults to active/non-Complete lists so the live workflow is quick to audit.
    Set include_complete=True for a slower historical/backlog audit.
    """
    board_info = get_board(board)
    lists = get_lists(board)
    list_names = {item["id"]: item["name"] for item in lists}
    cards = _request(
        "GET",
        f"boards/{_board_id(board)}/cards",
        params={"filter": "open", "fields": "id,name,desc,shortUrl,idList,dateLastActivity"},
    )
    if not include_complete:
        cards = [card for card in cards if list_names.get(card.get("idList")) != "Complete"]
    cards = cards[: max(1, min(limit_cards, 1000))]

    list_summary: dict[str, dict[str, Any]] = {}
    extension_counts: dict[str, int] = {}
    cards_with_address = 0
    cards_with_sheets = 0
    cards_without_sheets = 0
    cards_without_address = 0
    sample_gaps = []

    for card in cards:
        list_name = list_names.get(card.get("idList"), "UNKNOWN")
        row = list_summary.setdefault(
            list_name,
            {
                "cards": 0,
                "cards_with_address_candidates": 0,
                "cards_with_sheet_attachments": 0,
                "sheet_attachments": 0,
            },
        )
        row["cards"] += 1

        address_candidates = _extract_address_candidates(f"{card.get('name', '')}\n{card.get('desc', '')}")
        if address_candidates:
            cards_with_address += 1
            row["cards_with_address_candidates"] += 1
        else:
            cards_without_address += 1

        attachments = get_card_attachments(card["id"])
        sheet_attachments = [item for item in attachments if item["extension"] in SUPPORTED_SHEET_EXTENSIONS]
        if sheet_attachments:
            cards_with_sheets += 1
            row["cards_with_sheet_attachments"] += 1
            row["sheet_attachments"] += len(sheet_attachments)
            for item in sheet_attachments:
                ext = item["extension"] or "unknown"
                extension_counts[ext] = extension_counts.get(ext, 0) + 1
        else:
            cards_without_sheets += 1
            if len(sample_gaps) < 12:
                sample_gaps.append(
                    {
                        "card_id": card.get("id"),
                        "card_url": card.get("shortUrl"),
                        "list": list_name,
                        "has_address_candidates": bool(address_candidates),
                    }
                )

    return {
        "board": {k: board_info.get(k) for k in ("id", "name", "url", "closed", "dateLastActivity")},
        "scope": {
            "include_complete": include_complete,
            "cards_audited": len(cards),
            "limit_cards": limit_cards,
        },
        "summary": {
            "cards_with_address_candidates": cards_with_address,
            "cards_without_address_candidates": cards_without_address,
            "cards_with_sheet_attachments": cards_with_sheets,
            "cards_without_sheet_attachments": cards_without_sheets,
            "sheet_attachments": sum(extension_counts.values()),
            "sheet_extensions": extension_counts,
        },
        "lists": list_summary,
        "sample_cards_without_sheet_attachments": sample_gaps,
    }


def _board_snapshot(
    board: str = "memphis",
    *,
    include_complete: bool = True,
    limit_cards: int = 1000,
    sample_cards_per_list: int = 8,
) -> dict[str, Any]:
    board_info = get_board(board)
    lists = get_lists(board, filter="all" if include_complete else "open")
    labels = get_labels(board)
    list_by_id = {item["id"]: item for item in lists}
    cards = _request(
        "GET",
        f"boards/{_board_id(board)}/cards",
        params={
            "filter": "all" if include_complete else "open",
            "fields": "id,name,shortUrl,idList,closed,due,dateLastActivity,labels,badges",
        },
    )
    if not include_complete:
        cards = [card for card in cards if not card.get("closed") and not list_by_id.get(card.get("idList"), {}).get("closed")]
    cards = cards[: max(1, min(int(limit_cards), 2000))]

    by_list: dict[str, dict[str, Any]] = {}
    label_counts: dict[str, int] = {}
    total_attachments = 0
    total_check_items = 0
    total_check_items_checked = 0
    due_cards = 0
    closed_cards = 0

    for item in lists:
        by_list[item["name"]] = {
            "list_id": item.get("id"),
            "closed": bool(item.get("closed")),
            "cards": 0,
            "open_cards": 0,
            "closed_cards": 0,
            "due_cards": 0,
            "attachment_count": 0,
            "check_items": 0,
            "check_items_checked": 0,
            "sample_cards": [],
        }

    for card in cards:
        list_info = list_by_id.get(card.get("idList")) or {}
        list_name = list_info.get("name") or "UNKNOWN"
        row = by_list.setdefault(
            list_name,
            {
                "list_id": card.get("idList"),
                "closed": bool(list_info.get("closed")),
                "cards": 0,
                "open_cards": 0,
                "closed_cards": 0,
                "due_cards": 0,
                "attachment_count": 0,
                "check_items": 0,
                "check_items_checked": 0,
                "sample_cards": [],
            },
        )
        badges = card.get("badges") or {}
        attachment_count = int(badges.get("attachments") or 0)
        check_items = int(badges.get("checkItems") or 0)
        check_items_checked = int(badges.get("checkItemsChecked") or 0)
        is_closed = bool(card.get("closed"))
        has_due = bool(card.get("due"))

        row["cards"] += 1
        row["closed_cards"] += 1 if is_closed else 0
        row["open_cards"] += 0 if is_closed else 1
        row["due_cards"] += 1 if has_due else 0
        row["attachment_count"] += attachment_count
        row["check_items"] += check_items
        row["check_items_checked"] += check_items_checked
        total_attachments += attachment_count
        total_check_items += check_items
        total_check_items_checked += check_items_checked
        due_cards += 1 if has_due else 0
        closed_cards += 1 if is_closed else 0

        labels_on_card = _label_names(card)
        for label in labels_on_card:
            label_counts[label] = label_counts.get(label, 0) + 1

        if len(row["sample_cards"]) < max(0, min(int(sample_cards_per_list), 25)):
            row["sample_cards"].append(
                {
                    "id": card.get("id"),
                    "name": card.get("name"),
                    "url": card.get("shortUrl"),
                    "closed": is_closed,
                    "due": card.get("due"),
                    "dateLastActivity": card.get("dateLastActivity"),
                    "labels": labels_on_card,
                    "attachment_count": attachment_count,
                    "check_items": check_items,
                    "check_items_checked": check_items_checked,
                }
            )

    return {
        "ok": True,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "board_alias": board,
        "board": {k: board_info.get(k) for k in ("id", "name", "url", "closed", "dateLastActivity")},
        "scope": {
            "include_complete": include_complete,
            "cards_seen": len(cards),
            "limit_cards": limit_cards,
            "sample_cards_per_list": sample_cards_per_list,
        },
        "summary": {
            "lists": len(lists),
            "labels": len(labels),
            "cards": len(cards),
            "open_cards": len(cards) - closed_cards,
            "closed_cards": closed_cards,
            "due_cards": due_cards,
            "attachment_count": total_attachments,
            "check_items": total_check_items,
            "check_items_checked": total_check_items_checked,
            "label_counts": dict(sorted(label_counts.items())),
        },
        "labels": labels,
        "lists": by_list,
        "safety": {
            "read_only": True,
            "trello_writes": False,
            "google_drive_writes": False,
            "auto_attach": False,
        },
    }


def _format_board_snapshot(snapshot: dict[str, Any]) -> str:
    board = snapshot.get("board") or {}
    summary = snapshot.get("summary") or {}
    scope = snapshot.get("scope") or {}
    lines = [
        f"# Board Snapshot - {board.get('name') or 'UNKNOWN'}",
        "",
        f"- Generated: {snapshot.get('generated_at') or datetime.now().isoformat(timespec='seconds')}",
        f"- Board URL: {board.get('url') or 'UNKNOWN'}",
        f"- Board ID: {board.get('id') or 'UNKNOWN'}",
        f"- Include complete/closed: {scope.get('include_complete')}",
        f"- Cards seen: {scope.get('cards_seen')} of limit {scope.get('limit_cards')}",
        "",
        "## Summary",
        "",
        f"- Lists: {summary.get('lists', 0)}",
        f"- Labels: {summary.get('labels', 0)}",
        f"- Cards: {summary.get('cards', 0)}",
        f"- Open cards: {summary.get('open_cards', 0)}",
        f"- Closed cards: {summary.get('closed_cards', 0)}",
        f"- Due cards: {summary.get('due_cards', 0)}",
        f"- Attachments: {summary.get('attachment_count', 0)}",
        f"- Checklist items: {summary.get('check_items_checked', 0)} checked of {summary.get('check_items', 0)}",
        "",
        "## Labels",
        "",
    ]

    label_counts = summary.get("label_counts") or {}
    if label_counts:
        for label, count in label_counts.items():
            lines.append(f"- {label}: {count}")
    else:
        lines.append("- None")

    lines.extend(["", "## Lists", ""])
    for list_name, row in (snapshot.get("lists") or {}).items():
        lines.append(f"### {list_name}")
        lines.append("")
        lines.append(f"- Cards: {row.get('cards', 0)}")
        lines.append(f"- Open: {row.get('open_cards', 0)}")
        lines.append(f"- Closed: {row.get('closed_cards', 0)}")
        lines.append(f"- Due: {row.get('due_cards', 0)}")
        lines.append(f"- Attachments: {row.get('attachment_count', 0)}")
        lines.append(f"- Checklist items: {row.get('check_items_checked', 0)} checked of {row.get('check_items', 0)}")
        samples = row.get("sample_cards") or []
        if samples:
            lines.append("")
            lines.append("Sample cards:")
            for card in samples:
                labels = ", ".join(card.get("labels") or [])
                detail = []
                if labels:
                    detail.append(f"labels: {labels}")
                if card.get("due"):
                    detail.append(f"due: {card.get('due')}")
                if card.get("attachment_count"):
                    detail.append(f"attachments: {card.get('attachment_count')}")
                suffix = f" ({'; '.join(detail)})" if detail else ""
                lines.append(f"- {card.get('name') or 'Untitled'} - {card.get('url') or card.get('id')}{suffix}")
        lines.append("")

    lines.extend(
        [
            "## Safety",
            "",
            "- Trello writes: false",
            "- Google Drive writes: false",
            "- Auto-attach: false",
        ]
    )
    return "\n".join(lines).rstrip() + "\n"


def _load_board_snapshot(path: str) -> dict[str, Any]:
    snapshot_path = Path(path).expanduser()
    if not snapshot_path.exists():
        raise TrelloError(f"Board snapshot JSON not found: {snapshot_path}")
    try:
        payload = json.loads(snapshot_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise TrelloError(f"Board snapshot JSON is invalid: {snapshot_path}") from exc
    if not isinstance(payload, dict) or not isinstance(payload.get("summary"), dict) or not isinstance(payload.get("lists"), dict):
        raise TrelloError("Board snapshot JSON must contain summary and lists objects")
    return payload


def _numeric_delta(primary: dict[str, Any], comparison: dict[str, Any], keys: list[str]) -> dict[str, dict[str, int]]:
    deltas = {}
    for key in keys:
        left = int(primary.get(key) or 0)
        right = int(comparison.get(key) or 0)
        deltas[key] = {"primary": left, "comparison": right, "delta": left - right}
    return deltas


def _sample_card_names(snapshot: dict[str, Any]) -> set[str]:
    names = set()
    for row in (snapshot.get("lists") or {}).values():
        for card in row.get("sample_cards") or []:
            name = str(card.get("name") or "").strip()
            if name:
                names.add(name)
    return names


def _compare_board_snapshots(primary: dict[str, Any], comparison: dict[str, Any]) -> dict[str, Any]:
    primary_summary = primary.get("summary") or {}
    comparison_summary = comparison.get("summary") or {}
    primary_lists = primary.get("lists") or {}
    comparison_lists = comparison.get("lists") or {}
    primary_labels = primary_summary.get("label_counts") or {}
    comparison_labels = comparison_summary.get("label_counts") or {}
    primary_sample_names = _sample_card_names(primary)
    comparison_sample_names = _sample_card_names(comparison)
    numeric_keys = [
        "lists",
        "labels",
        "cards",
        "open_cards",
        "closed_cards",
        "due_cards",
        "attachment_count",
        "check_items",
        "check_items_checked",
    ]

    list_deltas = {}
    for list_name in sorted(set(primary_lists) | set(comparison_lists)):
        left = primary_lists.get(list_name) or {}
        right = comparison_lists.get(list_name) or {}
        list_deltas[list_name] = _numeric_delta(
            left,
            right,
            ["cards", "open_cards", "closed_cards", "due_cards", "attachment_count", "check_items", "check_items_checked"],
        )

    label_deltas = {}
    for label in sorted(set(primary_labels) | set(comparison_labels)):
        left = int(primary_labels.get(label) or 0)
        right = int(comparison_labels.get(label) or 0)
        label_deltas[label] = {"primary": left, "comparison": right, "delta": left - right}

    warnings = []
    if set(primary_lists) != set(comparison_lists):
        warnings.append("list_names_differ")
    if primary_summary.get("cards") != comparison_summary.get("cards"):
        warnings.append("card_count_differs")
    if primary_summary.get("attachment_count") != comparison_summary.get("attachment_count"):
        warnings.append("attachment_count_differs")
    if primary_summary.get("check_items") != comparison_summary.get("check_items"):
        warnings.append("checklist_item_count_differs")

    return {
        "ok": True,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "primary_board": primary.get("board"),
        "comparison_board": comparison.get("board"),
        "summary_deltas": _numeric_delta(primary_summary, comparison_summary, numeric_keys),
        "lists": {
            "primary_only": sorted(set(primary_lists) - set(comparison_lists)),
            "comparison_only": sorted(set(comparison_lists) - set(primary_lists)),
            "deltas": list_deltas,
        },
        "labels": {
            "primary_only": sorted(set(primary_labels) - set(comparison_labels)),
            "comparison_only": sorted(set(comparison_labels) - set(primary_labels)),
            "deltas": label_deltas,
        },
        "sample_cards": {
            "primary_only": sorted(primary_sample_names - comparison_sample_names),
            "comparison_only": sorted(comparison_sample_names - primary_sample_names),
            "overlap": sorted(primary_sample_names & comparison_sample_names),
            "note": "Sample-card comparison is limited to cards captured in each snapshot sample.",
        },
        "warnings": warnings,
        "safety": {
            "local_file_read": True,
            "local_file_write": False,
            "trello_writes": False,
            "google_drive_writes": False,
            "live_trello_read": False,
        },
    }


def _format_board_snapshot_comparison(comparison: dict[str, Any]) -> str:
    primary = comparison.get("primary_board") or {}
    other = comparison.get("comparison_board") or {}
    lines = [
        f"# Board Snapshot Comparison - {primary.get('name') or 'Primary'} vs {other.get('name') or 'Comparison'}",
        "",
        f"- Generated: {comparison.get('generated_at') or datetime.now().isoformat(timespec='seconds')}",
        f"- Primary: {primary.get('name') or 'UNKNOWN'} ({primary.get('url') or primary.get('id') or 'UNKNOWN'})",
        f"- Comparison: {other.get('name') or 'UNKNOWN'} ({other.get('url') or other.get('id') or 'UNKNOWN'})",
        "",
        "## Summary Deltas",
        "",
    ]
    for key, row in (comparison.get("summary_deltas") or {}).items():
        lines.append(f"- {key}: primary {row.get('primary', 0)}, comparison {row.get('comparison', 0)}, delta {row.get('delta', 0)}")

    warnings = comparison.get("warnings") or []
    lines.extend(["", "## Warnings", ""])
    if warnings:
        for warning in warnings:
            lines.append(f"- {warning}")
    else:
        lines.append("- None")

    lists = comparison.get("lists") or {}
    lines.extend(["", "## Lists", ""])
    primary_only = lists.get("primary_only") or []
    comparison_only = lists.get("comparison_only") or []
    lines.append(f"- Primary-only lists: {', '.join(primary_only) if primary_only else 'none'}")
    lines.append(f"- Comparison-only lists: {', '.join(comparison_only) if comparison_only else 'none'}")
    lines.append("")
    for list_name, deltas in (lists.get("deltas") or {}).items():
        card_delta = (deltas.get("cards") or {}).get("delta", 0)
        attachment_delta = (deltas.get("attachment_count") or {}).get("delta", 0)
        checklist_delta = (deltas.get("check_items") or {}).get("delta", 0)
        lines.append(f"- {list_name}: cards delta {card_delta}, attachments delta {attachment_delta}, checklist items delta {checklist_delta}")

    labels = comparison.get("labels") or {}
    lines.extend(["", "## Labels", ""])
    lines.append(f"- Primary-only labels: {', '.join(labels.get('primary_only') or []) or 'none'}")
    lines.append(f"- Comparison-only labels: {', '.join(labels.get('comparison_only') or []) or 'none'}")

    samples = comparison.get("sample_cards") or {}
    lines.extend(["", "## Sample Cards", ""])
    lines.append(f"- Primary-only sample cards: {len(samples.get('primary_only') or [])}")
    lines.append(f"- Comparison-only sample cards: {len(samples.get('comparison_only') or [])}")
    lines.append(f"- Overlap sample cards: {len(samples.get('overlap') or [])}")
    lines.append(f"- Note: {samples.get('note')}")

    lines.extend(
        [
            "",
            "## Safety",
            "",
            "- Trello writes: false",
            "- Google Drive writes: false",
            "- Live Trello read: false",
        ]
    )
    return "\n".join(lines).rstrip() + "\n"


def _load_board_snapshot_comparison(path: str) -> dict[str, Any]:
    comparison_path = Path(path).expanduser()
    if not comparison_path.exists():
        raise TrelloError(f"Board snapshot comparison JSON not found: {comparison_path}")
    try:
        payload = json.loads(comparison_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise TrelloError(f"Board snapshot comparison JSON is invalid: {comparison_path}") from exc
    if not isinstance(payload, dict) or not isinstance(payload.get("summary_deltas"), dict):
        raise TrelloError("Board snapshot comparison JSON must contain summary_deltas")
    return payload


def _cutover_readiness(comparison: dict[str, Any], proposed_action: str) -> dict[str, Any]:
    summary_deltas = comparison.get("summary_deltas") or {}
    lists = comparison.get("lists") or {}
    labels = comparison.get("labels") or {}
    samples = comparison.get("sample_cards") or {}
    warnings = list(comparison.get("warnings") or [])
    blockers = []
    review_items = []

    critical_delta_keys = ["cards", "open_cards", "attachment_count", "check_items"]
    for key in critical_delta_keys:
        delta = int((summary_deltas.get(key) or {}).get("delta") or 0)
        if delta != 0:
            blockers.append(f"{key}_delta_{delta}")

    if lists.get("primary_only") or lists.get("comparison_only"):
        blockers.append("list_sets_differ")
    if warnings:
        review_items.extend(warnings)
    if labels.get("primary_only") or labels.get("comparison_only"):
        review_items.append("label_sets_differ")
    if samples.get("primary_only") or samples.get("comparison_only"):
        review_items.append("sample_cards_differ")

    ready = not blockers
    recommendation = (
        "Do not archive or delete the legacy board yet. Resolve blockers and regenerate snapshots."
        if blockers
        else "No blocking deltas found in this comparison. Review notes still require human confirmation before archive/cutover."
    )

    return {
        "ok": True,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "proposed_action": proposed_action,
        "ready_for_cutover": ready,
        "ready_for_archive": ready and proposed_action in {"archive_legacy_memphis", "archive_legacy_board"},
        "blockers": blockers,
        "review_items": sorted(set(review_items)),
        "recommendation": recommendation,
        "comparison_summary": {
            "primary_board": comparison.get("primary_board"),
            "comparison_board": comparison.get("comparison_board"),
            "summary_deltas": summary_deltas,
            "warnings": warnings,
            "primary_only_lists": lists.get("primary_only") or [],
            "comparison_only_lists": lists.get("comparison_only") or [],
            "primary_only_labels": labels.get("primary_only") or [],
            "comparison_only_labels": labels.get("comparison_only") or [],
            "primary_only_sample_cards": len(samples.get("primary_only") or []),
            "comparison_only_sample_cards": len(samples.get("comparison_only") or []),
            "overlap_sample_cards": len(samples.get("overlap") or []),
        },
        "required_human_checks": [
            "Confirm snapshots were generated from the intended new Memphis and legacy Memphis boards.",
            "Confirm important cards, attachments, and checklist history are accounted for.",
            "Confirm no active work remains only on the legacy board.",
            "Confirm archive/delete action is intentional before any Trello write tool is used.",
        ],
        "safety": {
            "local_file_read": True,
            "local_file_write": False,
            "trello_writes": False,
            "google_drive_writes": False,
            "live_trello_read": False,
            "archive_or_delete_performed": False,
        },
    }


def _format_cutover_readiness(readiness: dict[str, Any]) -> str:
    summary = readiness.get("comparison_summary") or {}
    primary = summary.get("primary_board") or {}
    comparison = summary.get("comparison_board") or {}
    lines = [
        "# Board Cutover Readiness",
        "",
        f"- Generated: {readiness.get('generated_at') or datetime.now().isoformat(timespec='seconds')}",
        f"- Proposed action: {readiness.get('proposed_action')}",
        f"- Primary board: {primary.get('name') or 'UNKNOWN'}",
        f"- Comparison board: {comparison.get('name') or 'UNKNOWN'}",
        f"- Ready for cutover: {readiness.get('ready_for_cutover')}",
        f"- Ready for archive: {readiness.get('ready_for_archive')}",
        "",
        "## Recommendation",
        "",
        readiness.get("recommendation") or "UNKNOWN",
        "",
        "## Blockers",
        "",
    ]
    blockers = readiness.get("blockers") or []
    if blockers:
        for blocker in blockers:
            lines.append(f"- {blocker}")
    else:
        lines.append("- None")

    lines.extend(["", "## Review Items", ""])
    review_items = readiness.get("review_items") or []
    if review_items:
        for item in review_items:
            lines.append(f"- {item}")
    else:
        lines.append("- None")

    lines.extend(["", "## Summary Deltas", ""])
    for key, row in (summary.get("summary_deltas") or {}).items():
        lines.append(f"- {key}: primary {row.get('primary', 0)}, comparison {row.get('comparison', 0)}, delta {row.get('delta', 0)}")

    lines.extend(["", "## Required Human Checks", ""])
    for item in readiness.get("required_human_checks") or []:
        lines.append(f"- {item}")

    lines.extend(
        [
            "",
            "## Safety",
            "",
            "- Trello writes: false",
            "- Google Drive writes: false",
            "- Live Trello read: false",
            "- Archive/delete performed: false",
        ]
    )
    return "\n".join(lines).rstrip() + "\n"


@mcp.tool()
def write_board_snapshot_report(
    board: str = "memphis",
    include_complete: bool = True,
    limit_cards: int = 1000,
    sample_cards_per_list: int = 8,
) -> dict[str, Any]:
    """Write a local read-only board snapshot report for cutover readiness."""
    snapshot = _board_snapshot(
        board=board,
        include_complete=include_complete,
        limit_cards=limit_cards,
        sample_cards_per_list=sample_cards_per_list,
    )
    audit_dir = _artifact_dir("audits")
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    board_name = _safe_name((snapshot.get("board") or {}).get("name") or board)
    stem = _safe_name(f"board-snapshot-{board_name}-{timestamp}")
    json_path = audit_dir / f"{stem}.json"
    md_path = audit_dir / f"{stem}.md"

    json_path.write_text(json.dumps(snapshot, indent=2, sort_keys=True), encoding="utf-8")
    md_path.write_text(_format_board_snapshot(snapshot), encoding="utf-8")

    return {
        "ok": True,
        "markdown_path": str(md_path),
        "json_path": str(json_path),
        "summary": snapshot.get("summary"),
        "scope": snapshot.get("scope"),
        "safety": {
            "local_file_write": True,
            "trello_writes": False,
            "google_drive_writes": False,
            "auto_attach": False,
        },
    }


@mcp.tool()
def write_board_snapshot_comparison_report(
    primary_snapshot_json: str,
    comparison_snapshot_json: str,
    label: str = "board-snapshot-comparison",
) -> dict[str, Any]:
    """Compare two local board snapshot JSON files and write Markdown/JSON."""
    primary = _load_board_snapshot(primary_snapshot_json)
    comparison_snapshot = _load_board_snapshot(comparison_snapshot_json)
    comparison = _compare_board_snapshots(primary, comparison_snapshot)
    comparison["source_paths"] = {
        "primary_snapshot_json": str(Path(primary_snapshot_json).expanduser()),
        "comparison_snapshot_json": str(Path(comparison_snapshot_json).expanduser()),
    }
    comparison["safety"]["local_file_write"] = True

    audit_dir = _artifact_dir("audits")
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    primary_name = _safe_name((primary.get("board") or {}).get("name") or "primary")
    comparison_name = _safe_name((comparison_snapshot.get("board") or {}).get("name") or "comparison")
    stem = _safe_name(f"{label}-{primary_name}-vs-{comparison_name}-{timestamp}")
    json_path = audit_dir / f"{stem}.json"
    md_path = audit_dir / f"{stem}.md"

    json_path.write_text(json.dumps(comparison, indent=2, sort_keys=True), encoding="utf-8")
    md_path.write_text(_format_board_snapshot_comparison(comparison), encoding="utf-8")

    return {
        "ok": True,
        "markdown_path": str(md_path),
        "json_path": str(json_path),
        "warnings": comparison.get("warnings"),
        "summary_deltas": comparison.get("summary_deltas"),
        "safety": {
            "local_file_read": True,
            "local_file_write": True,
            "trello_writes": False,
            "google_drive_writes": False,
            "live_trello_read": False,
        },
    }


@mcp.tool()
def write_cutover_readiness_report(
    comparison_json: str,
    proposed_action: str = "archive_legacy_memphis",
) -> dict[str, Any]:
    """Write a local cutover/archive readiness report from a snapshot comparison."""
    comparison = _load_board_snapshot_comparison(comparison_json)
    readiness = _cutover_readiness(comparison, proposed_action)
    readiness["source_paths"] = {
        "comparison_json": str(Path(comparison_json).expanduser()),
    }
    readiness["safety"]["local_file_write"] = True

    audit_dir = _artifact_dir("audits")
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    primary_name = _safe_name(((comparison.get("primary_board") or {}).get("name")) or "primary")
    comparison_name = _safe_name(((comparison.get("comparison_board") or {}).get("name")) or "comparison")
    stem = _safe_name(f"cutover-readiness-{primary_name}-vs-{comparison_name}-{timestamp}")
    json_path = audit_dir / f"{stem}.json"
    md_path = audit_dir / f"{stem}.md"

    json_path.write_text(json.dumps(readiness, indent=2, sort_keys=True), encoding="utf-8")
    md_path.write_text(_format_cutover_readiness(readiness), encoding="utf-8")

    return {
        "ok": True,
        "ready_for_cutover": readiness.get("ready_for_cutover"),
        "ready_for_archive": readiness.get("ready_for_archive"),
        "blockers": readiness.get("blockers"),
        "review_items": readiness.get("review_items"),
        "recommendation": readiness.get("recommendation"),
        "markdown_path": str(md_path),
        "json_path": str(json_path),
        "safety": {
            "local_file_read": True,
            "local_file_write": True,
            "trello_writes": False,
            "google_drive_writes": False,
            "live_trello_read": False,
            "archive_or_delete_performed": False,
        },
    }


@mcp.tool()
def write_cutover_packet(
    primary_board: str = "memphis",
    comparison_board: str = "legacy_memphis",
    proposed_action: str = "archive_legacy_memphis",
    include_complete: bool = True,
    limit_cards: int = 1000,
    sample_cards_per_list: int = 8,
) -> dict[str, Any]:
    """Write the full local cutover packet: snapshots, comparison, readiness.

    This performs live Trello reads through the snapshot tools when credentials
    are available. It never writes to Trello.
    """
    primary_snapshot = write_board_snapshot_report(
        board=primary_board,
        include_complete=include_complete,
        limit_cards=limit_cards,
        sample_cards_per_list=sample_cards_per_list,
    )
    comparison_snapshot = write_board_snapshot_report(
        board=comparison_board,
        include_complete=include_complete,
        limit_cards=limit_cards,
        sample_cards_per_list=sample_cards_per_list,
    )
    comparison_report = write_board_snapshot_comparison_report(
        primary_snapshot_json=str(primary_snapshot["json_path"]),
        comparison_snapshot_json=str(comparison_snapshot["json_path"]),
        label="cutover-comparison",
    )
    readiness_report = write_cutover_readiness_report(
        comparison_json=str(comparison_report["json_path"]),
        proposed_action=proposed_action,
    )

    return {
        "ok": True,
        "primary_board": primary_board,
        "comparison_board": comparison_board,
        "proposed_action": proposed_action,
        "artifacts": {
            "primary_snapshot_markdown": primary_snapshot.get("markdown_path"),
            "primary_snapshot_json": primary_snapshot.get("json_path"),
            "comparison_snapshot_markdown": comparison_snapshot.get("markdown_path"),
            "comparison_snapshot_json": comparison_snapshot.get("json_path"),
            "comparison_markdown": comparison_report.get("markdown_path"),
            "comparison_json": comparison_report.get("json_path"),
            "readiness_markdown": readiness_report.get("markdown_path"),
            "readiness_json": readiness_report.get("json_path"),
        },
        "readiness": {
            "ready_for_cutover": readiness_report.get("ready_for_cutover"),
            "ready_for_archive": readiness_report.get("ready_for_archive"),
            "blockers": readiness_report.get("blockers"),
            "review_items": readiness_report.get("review_items"),
            "recommendation": readiness_report.get("recommendation"),
        },
        "safety": {
            "local_file_write": True,
            "trello_reads": True,
            "trello_writes": False,
            "google_drive_writes": False,
            "archive_or_delete_performed": False,
        },
    }


@mcp.tool()
def inspect_cards_without_sheet_attachments(
    board: str = "memphis",
    include_complete: bool = False,
    limit_cards: int = 500,
) -> dict[str, Any]:
    """Read-only detail for cards that do not have XLS/XLSX/CSV attachments."""
    board_info = get_board(board)
    lists = get_lists(board)
    list_names = {item["id"]: item["name"] for item in lists}
    cards = _request(
        "GET",
        f"boards/{_board_id(board)}/cards",
        params={"filter": "open", "fields": "id,name,desc,shortUrl,idList,due,dateLastActivity,labels"},
    )
    if not include_complete:
        cards = [card for card in cards if list_names.get(card.get("idList")) != "Complete"]
    cards = cards[: max(1, min(limit_cards, 1000))]

    gaps = []
    for card in cards:
        attachments = get_card_attachments(card["id"])
        sheet_attachments = [item for item in attachments if item["extension"] in SUPPORTED_SHEET_EXTENSIONS]
        if sheet_attachments:
            continue

        list_name = list_names.get(card.get("idList"), "UNKNOWN")
        address_candidates = _extract_address_candidates(f"{card.get('name', '')}\n{card.get('desc', '')}")
        non_sheet_attachments = [
            {
                "id": item.get("id"),
                "name": item.get("name"),
                "extension": item.get("extension"),
                "mimeType": item.get("mimeType"),
                "bytes": item.get("bytes"),
                "date": item.get("date"),
            }
            for item in attachments
        ]
        labels = [label.get("name") or label.get("color") for label in card.get("labels", [])]

        if list_name in {"BILLS", "My Bill"}:
            reason = "billing_lane_sheet_may_not_be_expected"
        elif "SAFETY COVER" in list_name.upper():
            reason = "safety_cover_lane_may_use_non_tara_artifacts"
        elif list_name == "MEASURES":
            reason = "measure_lane_missing_expected_sheet_review"
        else:
            reason = "missing_sheet_review"

        gaps.append(
            {
                "card_id": card.get("id"),
                "card_name": card.get("name"),
                "card_url": card.get("shortUrl"),
                "list": list_name,
                "labels": labels,
                "due": card.get("due"),
                "dateLastActivity": card.get("dateLastActivity"),
                "address_candidates": address_candidates,
                "non_sheet_attachment_count": len(non_sheet_attachments),
                "non_sheet_attachments": non_sheet_attachments,
                "classification": reason,
            }
        )

    by_classification: dict[str, int] = {}
    by_list: dict[str, int] = {}
    for gap in gaps:
        by_classification[gap["classification"]] = by_classification.get(gap["classification"], 0) + 1
        by_list[gap["list"]] = by_list.get(gap["list"], 0) + 1

    return {
        "board": {k: board_info.get(k) for k in ("id", "name", "url", "closed", "dateLastActivity")},
        "scope": {
            "include_complete": include_complete,
            "cards_scanned": len(cards),
            "limit_cards": limit_cards,
        },
        "summary": {
            "cards_without_sheet_attachments": len(gaps),
            "by_list": by_list,
            "by_classification": by_classification,
        },
        "cards": gaps,
    }


@mcp.tool()
def find_sheet_candidates_for_card(
    card_id: str,
    search_boards_json: str = '["memphis", "legacy_memphis"]',
    include_complete: bool = True,
    limit_candidates: int = 20,
) -> dict[str, Any]:
    """Find likely XLS/XLSX/CSV attachment candidates for a card, read-only."""
    target = get_card(card_id)
    target_text = f"{target.get('name', '')}\n{target.get('desc', '')}"
    target_tokens = _match_tokens(target_text)
    search_boards = json.loads(search_boards_json)
    candidates = []

    for board in search_boards:
        board_info = get_board(board)
        lists = get_lists(board)
        list_names = {item["id"]: item["name"] for item in lists}
        cards = _request(
            "GET",
            f"boards/{_board_id(board)}/cards",
            params={"filter": "all" if include_complete else "open", "fields": "id,name,desc,shortUrl,idList,closed,dateLastActivity"},
        )

        for card in cards:
            if card.get("id") == card_id:
                continue
            name_desc_score = _score_text(target_tokens, f"{card.get('name', '')}\n{card.get('desc', '')}")
            if name_desc_score == 0:
                continue

            attachments = get_card_attachments(card["id"])
            sheet_attachments = [item for item in attachments if item["extension"] in SUPPORTED_SHEET_EXTENSIONS]
            if not sheet_attachments:
                continue

            attachment_score = 0
            for item in sheet_attachments:
                attachment_score = max(attachment_score, _score_text(target_tokens, item.get("name") or ""))

            score = name_desc_score * 2 + attachment_score
            candidates.append(
                {
                    "score": score,
                    "board_alias": board,
                    "board_id": board_info.get("id"),
                    "board_name": board_info.get("name"),
                    "card_id": card.get("id"),
                    "card_name": card.get("name"),
                    "card_url": card.get("shortUrl"),
                    "list": list_names.get(card.get("idList"), "UNKNOWN"),
                    "closed": card.get("closed"),
                    "dateLastActivity": card.get("dateLastActivity"),
                    "name_desc_score": name_desc_score,
                    "attachment_score": attachment_score,
                    "sheet_attachments": [
                        {
                            "id": item.get("id"),
                            "name": item.get("name"),
                            "extension": item.get("extension"),
                            "bytes": item.get("bytes"),
                            "date": item.get("date"),
                        }
                        for item in sheet_attachments
                    ],
                }
            )

    candidates.sort(key=lambda item: (item["score"], item["attachment_score"], item["dateLastActivity"] or ""), reverse=True)
    candidates = candidates[: max(1, min(limit_candidates, 100))]

    return {
        "target": {
            "card_id": target.get("id"),
            "card_name": target.get("name"),
            "card_url": target.get("shortUrl"),
            "tokens": sorted(target_tokens),
        },
        "scope": {
            "search_boards": search_boards,
            "include_complete": include_complete,
            "limit_candidates": limit_candidates,
        },
        "candidates": candidates,
    }


@mcp.tool()
def download_attachment(card_id: str, attachment_id: str, output_dir: str | None = None) -> dict[str, Any]:
    """Download one Trello attachment into the local artifact cache."""
    attachment = _request("GET", f"cards/{card_id}/attachments/{attachment_id}")
    target_dir = Path(output_dir).expanduser() if output_dir else _artifact_dir("attachments")
    target_dir.mkdir(parents=True, exist_ok=True)
    name = _safe_name(attachment.get("name") or attachment_id)
    path = target_dir / f"{card_id}-{attachment_id}-{name}"
    content = _download_request(_attachment_download_url(card_id, attachment))
    path.write_bytes(content)
    return {
        "path": str(path),
        "filename": attachment.get("name"),
        "bytes": len(content),
        "sha256": _sha256(path),
        "mimeType": attachment.get("mimeType"),
        "date": attachment.get("date"),
    }


@mcp.tool()
def read_xls_attachment(card_id: str, attachment_id: str, max_rows: int = 40, max_cols: int = 20) -> dict[str, Any]:
    """Download and preview an XLS/XLSX/CSV attachment from a Trello card."""
    downloaded = download_attachment(card_id, attachment_id)
    path = Path(downloaded["path"])
    if path.suffix.lower() not in SUPPORTED_SHEET_EXTENSIONS:
        raise TrelloError(f"Attachment is not a supported sheet file: {path.name}")
    preview = _excel_preview(path, max_rows=max_rows, max_cols=max_cols)
    text = _flatten_preview_text(preview)
    return {
        "download": downloaded,
        "preview": preview,
        "address_candidates": _extract_address_candidates(text),
    }


def _safe_attachment_name(name: str) -> str:
    clean = re.sub(r"[^A-Za-z0-9_.-]+", "_", name or "attachment").strip("._")
    return clean[:180] or "attachment"


def _attachment_summary(attachment: dict[str, Any]) -> dict[str, Any]:
    name = attachment.get("name") or ""
    return {
        "id": attachment.get("id"),
        "name": name,
        "url": attachment.get("url"),
        "bytes": attachment.get("bytes"),
        "date": attachment.get("date"),
        "mimeType": attachment.get("mimeType"),
        "isUpload": attachment.get("isUpload"),
        "extension": Path(name).suffix.lower(),
    }


def _work_order_attachment_score(attachment: dict[str, Any], name_filter: str = "") -> int:
    name = str(attachment.get("name") or "").lower()
    if name_filter and name_filter.lower() not in name:
        return -1000
    score = 0
    suffix = Path(name).suffix.lower()
    if suffix in {".xls", ".xlsx", ".csv", ".pdf"}:
        score += 40
    if "install" in name:
        score += 35
    if "work" in name and "order" in name:
        score += 30
    if "measure" in name:
        score += 20
    if "liner" in name:
        score += 15
    if "drawing" in name or "scan" in name:
        score += 5
    if suffix in {".jpg", ".jpeg", ".png", ".gif", ".webp", ".heic"}:
        score -= 80
    return score


def _normalize_extracted_text(text: str) -> str:
    lines = []
    seen = set()
    for line in (text or "").replace("\r", "\n").splitlines():
        clean = re.sub(r"\s+", " ", line).strip()
        if len(clean) < 2 or clean in seen:
            continue
        seen.add(clean)
        lines.append(clean)
    return "\n".join(lines)


def _extract_attachment_text(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        try:
            result = subprocess.run(
                ["pdftotext", "-layout", str(path), "-"],
                check=False,
                capture_output=True,
                text=True,
                timeout=30,
            )
            if result.returncode == 0 and result.stdout.strip():
                return _normalize_extracted_text(result.stdout)
        except Exception:
            pass
        try:
            from pypdf import PdfReader

            reader = PdfReader(str(path))
            text = "\n".join(page.extract_text() or "" for page in reader.pages)
            return _normalize_extracted_text(text)
        except Exception as exc:
            return f"[pdf text extraction failed: {exc!r}]"

    if suffix == ".xlsx":
        try:
            chunks: list[str] = []
            with zipfile.ZipFile(path) as zf:
                if "xl/sharedStrings.xml" in zf.namelist():
                    root = ElementTree.fromstring(zf.read("xl/sharedStrings.xml"))
                    for node in root.iter():
                        if (node.tag.endswith("}t") or node.tag == "t") and node.text:
                            chunks.append(node.text)
                for name in zf.namelist():
                    if not (name.startswith("xl/worksheets/") and name.endswith(".xml")):
                        continue
                    root = ElementTree.fromstring(zf.read(name))
                    for node in root.iter():
                        if node.text and node.text.strip():
                            chunks.append(node.text)
            return _normalize_extracted_text("\n".join(chunks))
        except Exception as exc:
            return f"[xlsx text extraction failed: {exc!r}]"

    if suffix == ".xls":
        chunks = []
        for cmd in (["strings", "-a", str(path)], ["strings", "-el", str(path)]):
            try:
                result = subprocess.run(cmd, check=False, capture_output=True, text=True, timeout=30)
                if result.stdout:
                    chunks.append(result.stdout)
            except Exception:
                continue
        return _normalize_extracted_text("\n".join(chunks))

    try:
        return _normalize_extracted_text(path.read_text(encoding="utf-8", errors="replace"))
    except Exception:
        return ""


def _work_order_hints(text: str) -> dict[str, Any]:
    lines = [line for line in (text or "").splitlines() if line.strip()]
    haul_lines = [
        line
        for line in lines
        if re.search(r"\bha(?:u)?l\s*-?\s*off\b|hauloff|trash|dispose|disposal", line, re.IGNORECASE)
    ]
    joined = "\n".join(haul_lines)
    haul_off = None
    if re.search(r"\bno\s+haul\s*-?\s*off\b|\bno\s+hauloff\b", joined, re.IGNORECASE):
        haul_off = "NO HAUL OFF"
    elif re.search(r"\bha(?:u)?l\s*-?\s*off\b|hauloff", joined, re.IGNORECASE):
        haul_off = "HAUL OFF MENTIONED"
    important = [
        line
        for line in lines
        if re.search(
            r"haul|liner|install|measure|steps|inlay|freeform|safety|cover|fold|address|phone|customer",
            line,
            re.IGNORECASE,
        )
    ]
    return {
        "haul_off": haul_off,
        "haul_lines": haul_lines[:12],
        "important_lines": important[:40],
    }


@mcp.tool()
def attach_file_to_card(
    file_path: str,
    card_id: str | None = None,
    card_query: str | None = None,
    board: str = "memphis",
    name: str | None = None,
    comment: str | None = None,
) -> dict[str, Any]:
    """Attach a local file to a Trello card.

    This is a real Trello write. Provide either card_id or a card_query that
    resolves to exactly one open card. Comments are optional and are never
    inferred from a spoken/user prompt.
    """
    if not card_id and not card_query:
        raise TrelloError("card_id or card_query is required")
    path = Path(file_path).expanduser()
    if not path.is_file():
        raise TrelloError(f"File does not exist: {path}")

    card = get_card(card_id) if card_id else _find_card_by_query(str(card_query), board=board)
    content_type = mimetypes.guess_type(str(path))[0] or "application/octet-stream"
    attachment_name = (name or path.name).strip()[:256]
    with path.open("rb") as fh:
        attachment = _request_files(
            "POST",
            f"cards/{card['id']}/attachments",
            params={"name": attachment_name},
            files={"file": (path.name, fh, content_type)},
        )
    comment_action = None
    if comment and comment.strip():
        comment_action = add_comment(str(card["id"]), comment.strip()[:1600])
    return {
        "ok": True,
        "card": _card_with_board_context(card),
        "attachment": _attachment_summary(attachment),
        "comment_action_id": comment_action.get("id") if isinstance(comment_action, dict) else None,
        "safety": {
            "trello_writes": True,
            "attachment_written": True,
            "comment_written": bool(comment_action),
            "user_prompt_used_as_comment": False,
        },
    }


@mcp.tool()
def set_card_cover(card_id: str, attachment_id: str) -> dict[str, Any]:
    """Set a Trello card cover to an existing attachment id."""
    if not card_id or not attachment_id:
        raise TrelloError("card_id and attachment_id are required")
    card = _request("PUT", f"cards/{card_id}", params={"idAttachmentCover": attachment_id})
    return {
        "ok": True,
        "card": _card_with_board_context(card),
        "attachment_id": attachment_id,
        "safety": {"trello_writes": True, "cover_changed": True},
    }


@mcp.tool()
def read_card_work_order(
    card_id: str | None = None,
    card_query: str | None = None,
    board: str = "memphis",
    attachment_filter: str = "",
    question: str = "",
    max_text_chars: int = 6000,
) -> dict[str, Any]:
    """Read the best work-order-like attachment on a Trello card.

    This is read-only. It prefers install/work-order spreadsheets and PDFs over
    photos, downloads the selected attachment to the local artifact cache, and
    extracts useful text plus haul-off hints.
    """
    if not card_id and not card_query:
        raise TrelloError("card_id or card_query is required")
    card = get_card(card_id) if card_id else _find_card_by_query(str(card_query), board=board)
    attachments = get_card_attachments(str(card["id"]))
    ranked = sorted(attachments, key=lambda item: _work_order_attachment_score(item, attachment_filter), reverse=True)
    candidates = [
        item
        for item in ranked
        if _work_order_attachment_score(item, attachment_filter) > 0
        and Path(item.get("name") or "").suffix.lower() in SUPPORTED_WORK_ORDER_EXTENSIONS
    ]
    if not candidates:
        return {
            "ok": False,
            "error": f"No readable work-order attachment found on {card.get('name')}",
            "card": _card_with_board_context(card),
            "attachment_filter": attachment_filter,
            "attachments": [_attachment_summary(item) for item in attachments[:30]],
            "safety": {"read_only": True, "trello_writes": False},
        }

    attachment = candidates[0]
    downloaded = download_attachment(str(card["id"]), str(attachment["id"]), output_dir=str(_artifact_dir("work-orders")))
    local_path = Path(downloaded["path"])
    text = _extract_attachment_text(local_path)
    limit = max(500, min(int(max_text_chars), 20000))
    return {
        "ok": True,
        "card": _card_with_board_context(card),
        "attachment": _attachment_summary(attachment),
        "download": downloaded,
        "question": question,
        "hints": _work_order_hints(text),
        "text": text[:limit],
        "text_truncated": len(text) > limit,
        "safety": {
            "read_only": True,
            "trello_writes": False,
            "local_file_write": True,
        },
    }


@mcp.tool()
def scan_tara_xls_forms(source_path: str, max_files: int = 200) -> list[dict[str, Any]]:
    """Read-only scan of local XLS/XLSX/CSV files for Tara/work-order matching."""
    root = Path(source_path).expanduser()
    if not root.exists():
        raise TrelloError(f"Source path does not exist: {root}")
    files = [root] if root.is_file() else [p for p in root.rglob("*") if p.suffix.lower() in SUPPORTED_SHEET_EXTENSIONS]
    results = []
    for path in files[: max(1, min(max_files, 1000))]:
        try:
            preview = _excel_preview(path, max_rows=25, max_cols=12)
            results.append(
                {
                    "path": str(path),
                    "filename": path.name,
                    "size": path.stat().st_size,
                    "modified_at": datetime.fromtimestamp(path.stat().st_mtime).isoformat(),
                    "sha256": _sha256(path),
                    "address_candidates": _extract_address_candidates(_flatten_preview_text(preview)),
                    "preview": preview,
                }
            )
        except Exception as exc:
            results.append({"path": str(path), "error": str(exc)})
    return results


@mcp.tool()
def photos_intake_status(
    photos_library_path: str = str(DEFAULT_PHOTOS_LIBRARY_PATH),
    days_back: int = 365,
    max_assets: int = 10000,
    sample_limit: int = 20,
    include_paths: bool = False,
    include_coordinates: bool = False,
) -> dict[str, Any]:
    """Read-only Mac Photos intake status for pool-photo matching.

    This inspects the local Photos Library database and reports how many image
    assets have GPS, how many originals are local, and how many GPS-bearing
    originals are still cloud-only from this Mac's point of view. It does not
    write to Photos, iCloud, or Trello.
    """
    if days_back < 0:
        raise TrelloError("days_back must be >= 0")
    if sample_limit < 0:
        raise TrelloError("sample_limit must be >= 0")
    root = Path(photos_library_path or str(DEFAULT_PHOTOS_LIBRARY_PATH)).expanduser()
    if not _looks_like_photos_library(root):
        raise TrelloError(f"Not a Photos Library path: {root}")

    report = _photos_library_intake_status(
        root,
        days_back=days_back,
        max_assets=max_assets,
        sample_limit=sample_limit,
        include_paths=include_paths,
        include_coordinates=include_coordinates,
    )
    batch_id = datetime.now().strftime("photos-intake-%Y%m%d-%H%M%S")
    report_dir = PHOTOS_INTAKE_ROOT / batch_id
    report_dir.mkdir(parents=True, exist_ok=True)
    json_path = report_dir / "mac-photos-intake-status.json"
    markdown_path = report_dir / "mac-photos-intake-status.md"
    report["reports"] = {"json": str(json_path), "markdown": str(markdown_path)}
    _write_json_file(json_path, report)

    counts = report["counts"]
    lines = [
        "# Mac Photos Intake Status",
        "",
        f"- Photos library: {report['source']['photos_library_path']}",
        f"- Days back: {days_back}",
        f"- Assets scanned: {counts['assets_scanned']}",
        f"- Supported image assets: {counts['supported_image_assets']}",
        f"- Local originals: {counts['local_originals']}",
        f"- Missing originals: {counts['missing_originals']}",
        f"- GPS assets: {counts['gps_assets']}",
        f"- GPS local originals: {counts['gps_local_originals']}",
        f"- GPS missing originals: {counts['gps_missing_originals']}",
        "",
        "## Recommendations",
        "",
    ]
    for item in report["recommendations"]:
        lines.append(f"- {item}")
    lines.extend(["", "## Sample Local GPS Originals", ""])
    for item in report["samples"]["gps_local_ready"][:sample_limit]:
        lines.append(f"- {item.get('taken_at')} | {item.get('filename')} | {item.get('extension')} | {item.get('width')}x{item.get('height')}")
    lines.extend(["", "## Sample GPS Assets Missing Originals", ""])
    for item in report["samples"]["gps_missing_original"][:sample_limit]:
        lines.append(f"- {item.get('taken_at')} | {item.get('filename')} | {item.get('extension')} | cloud_state={item.get('cloud_local_state')}")
    markdown_path.write_text("\n".join(lines), encoding="utf-8")
    return report


@mcp.tool()
def prepare_photokit_export_manifest(
    photos_library_path: str = str(DEFAULT_PHOTOS_LIBRARY_PATH),
    days_back: int = 365,
    max_assets: int = 10000,
    limit: int = 100,
    dry_run_helper: bool = True,
) -> dict[str, Any]:
    """Prepare an export manifest for cloud-only GPS photos.

    This reads the local Photos database, selects GPS-tagged assets whose
    original files are not local under the Photos Library, writes a local
    manifest, checks the PhotoKit exporter, and optionally dry-runs the
    exporter. It does not export/download originals unless the apply tool is
    run.
    """
    if days_back < 0:
        raise TrelloError("days_back must be >= 0")
    root = Path(photos_library_path or str(DEFAULT_PHOTOS_LIBRARY_PATH)).expanduser()
    if not _looks_like_photos_library(root):
        raise TrelloError(f"Not a Photos Library path: {root}")
    batch_id = datetime.now().strftime("photokit-export-%Y%m%d-%H%M%S")
    export_dir = PHOTOS_INTAKE_ROOT / batch_id
    output_dir = export_dir / "exported-originals"
    export_dir.mkdir(parents=True, exist_ok=True)

    assets = _photos_missing_original_assets(root, days_back=days_back, max_assets=max_assets, limit=limit)
    manifest = {
        "ok": True,
        "mode": "photokit_export_manifest",
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "source": {
            "photos_library_path": str(root),
            "days_back": days_back,
            "max_assets": max_assets,
        },
        "photo_ids": [str(item["photo_id"]) for item in assets],
        "assets": assets,
        "output_dir": str(output_dir),
        "summary": {
            "candidate_missing_gps_originals": len(assets),
            "limit": max(1, min(int(limit), 5000)),
        },
        "safety": {
            "read_only_manifest": True,
            "local_manifest_write": True,
            "photos_writes": False,
            "trello_writes": False,
            "icloud_network_access": False,
            "raw_gps_returned": False,
            "secrets_returned": False,
            "export_requires_token": "EXPORT_PHOTOS_FROM_ICLOUD",
        },
    }
    manifest_path = export_dir / "photokit-export-manifest.json"
    markdown_path = export_dir / "photokit-export-manifest.md"
    _write_json_file(manifest_path, manifest)
    photokit_result = _compile_photokit_exporter()
    osxphotos_result = _osxphotos_status()
    manifest["reports"] = {
        "manifest_json": str(manifest_path),
        "manifest_markdown": str(markdown_path),
        "export_output_dir": str(output_dir),
        "default_backend": "photokit",
        "osxphotos_backend": " ".join(_osxphotos_command_base()),
        "photokit_exporter": str(PHOTOKIT_EXPORTER_BIN),
    }
    manifest["photokit"] = photokit_result
    manifest["osxphotos"] = osxphotos_result
    if dry_run_helper and assets and photokit_result.get("ok"):
        manifest["dry_run"] = _run_photokit_exporter(
            manifest_path=manifest_path,
            output_dir=output_dir,
            limit=min(len(assets), max(1, min(int(limit), 25))),
            dry_run=True,
            timeout_seconds=120,
        )

    lines = [
        "# PhotoKit Export Manifest",
        "",
        f"- Photos library: {root}",
        f"- Candidate missing GPS originals: {len(assets)}",
        f"- Manifest: {manifest_path}",
        f"- Export output dir: {output_dir}",
        f"- PhotoKit exporter ready: {photokit_result.get('ok')}",
        f"- osxphotos ready: {osxphotos_result.get('ok')}",
        "",
        "## Apply Command",
        "",
        "Use the MCP apply tool `export_photokit_photo_originals` with backend `photokit` and apply token `EXPORT_PHOTOS_FROM_ICLOUD`.",
        "",
        "## Sample Assets",
        "",
    ]
    for item in assets[:25]:
        lines.append(f"- {item.get('taken_at')} | {item.get('filename')} | {item.get('extension')} | cloud_state={item.get('cloud_local_state')}")
    markdown_path.write_text("\n".join(lines), encoding="utf-8")
    _write_json_file(manifest_path, manifest)
    return manifest


@mcp.tool()
def export_photokit_photo_originals(
    manifest_json_path: str,
    dry_run: bool = True,
    apply_token: str = "",
    limit: int = 25,
    timeout_seconds: int = 300,
    backend: str = "photokit",
) -> dict[str, Any]:
    """Export/download Photos originals through the configured Photos backend.

    Defaults to dry-run. A real export writes files only under the manifest's
    local output directory and requires apply_token=EXPORT_PHOTOS_FROM_ICLOUD.
    It does not write to Trello or mutate the Photos library. The default
    backend is the signed PhotoKit helper because it is authorized by macOS
    Photos privacy controls in this room. The osxphotos backend remains
    available as a diagnostic fallback.
    """
    manifest_path = Path(manifest_json_path).expanduser()
    if not manifest_path.exists():
        raise TrelloError(f"Manifest JSON not found: {manifest_path}")
    manifest = _read_json_file(manifest_path, None)
    if not isinstance(manifest, dict) or manifest.get("mode") != "photokit_export_manifest":
        raise TrelloError("File is not a photokit_export_manifest JSON")
    required_token = "EXPORT_PHOTOS_FROM_ICLOUD"
    apply_enabled = not dry_run
    if apply_enabled and apply_token != required_token:
        raise TrelloError(f"Refusing Photos export. Re-run with apply_token={required_token!r} to export originals.")
    output_dir = Path(str(manifest.get("output_dir") or manifest_path.parent / "exported-originals")).expanduser()
    backend_key = (backend or "photokit").strip().lower()
    if backend_key == "osxphotos":
        result = _run_osxphotos_exporter(
            manifest=manifest,
            manifest_path=manifest_path,
            output_dir=output_dir,
            limit=limit,
            dry_run=dry_run,
            timeout_seconds=timeout_seconds,
        )
    elif backend_key in {"photokit", "swift", "helper"}:
        result = _run_photokit_exporter(
            manifest_path=manifest_path,
            output_dir=output_dir,
            limit=limit,
            dry_run=dry_run,
            timeout_seconds=timeout_seconds,
        )
    else:
        raise TrelloError("backend must be 'osxphotos' or 'photokit'")
    result["mode"] = "photos_original_export"
    result["manifest_json_path"] = str(manifest_path)
    result["backend_requested"] = backend_key
    result["required_apply_token"] = required_token
    result.setdefault("safety", {})
    result["safety"].update(
        {
            "photos_writes": False,
            "local_file_writes": not dry_run,
            "trello_writes": False,
            "icloud_network_access_allowed": not dry_run,
            "secrets_returned": False,
        }
    )
    return result


@mcp.tool()
def scan_pool_photos(source_path: str, max_files: int = 500) -> list[dict[str, Any]]:
    """Read-only scan of local photo files for GPS/time metadata."""
    root = Path(source_path).expanduser()
    if not root.exists():
        raise TrelloError(f"Source path does not exist: {root}")
    if _looks_like_photos_library(root):
        return _photos_library_assets(root, days_back=3650, max_files=max_files, require_local=False)
    files = [root] if root.is_file() else [p for p in root.rglob("*") if p.suffix.lower() in SUPPORTED_PHOTO_EXTENSIONS]
    results = []
    for path in files[: max(1, min(max_files, 5000))]:
        try:
            results.append(_photo_metadata(path))
        except Exception as exc:
            results.append({"path": str(path), "filename": path.name, "error": str(exc)})
    return results


@mcp.tool()
def match_photos_to_locations(
    source_path: str,
    locations_json: str,
    radius_ft: int = 300,
    max_files: int = 500,
) -> list[dict[str, Any]]:
    """Match GPS-tagged photos to supplied card/location records.

    locations_json must be a JSON array with objects containing card_id, name,
    latitude, and longitude. This keeps geocoding outside the matcher until the
    address source is verified.
    """
    locations = json.loads(locations_json)
    photos = [p for p in scan_pool_photos(source_path, max_files=max_files) if p.get("has_gps")]
    matches = []
    for photo in photos:
        ranked = []
        for loc in locations:
            if loc.get("latitude") is None or loc.get("longitude") is None:
                continue
            distance = _distance_ft(float(photo["latitude"]), float(photo["longitude"]), float(loc["latitude"]), float(loc["longitude"]))
            if distance <= radius_ft:
                ranked.append(
                    {
                        "card_id": loc.get("card_id"),
                        "card_name": loc.get("name"),
                        "card_url": loc.get("url"),
                        "address": loc.get("address"),
                        "distance_ft": round(distance, 1),
                    }
                )
        if ranked:
            ranked.sort(key=lambda item: item["distance_ft"])
            best = ranked[0]
            confidence = "very_strong" if best["distance_ft"] <= 100 else "likely" if best["distance_ft"] <= 300 else "review"
            matches.append({"photo": photo, "best_match": best, "confidence": confidence, "all_matches": ranked[:5]})
    return matches


@mcp.tool()
def preview_photo_card_matches(
    source_path: str = str(DEFAULT_PHOTOS_LIBRARY_PATH),
    board: str = "memphis",
    days_back: int = 60,
    radius_ft: int = 350,
    max_photos: int = 500,
    limit_cards: int = 500,
    include_complete: bool = False,
    geocode_limit: int = 100,
    stage_files: bool = True,
    max_stage_files: int = 120,
    export_missing_originals: bool = False,
    missing_export_limit: int = 25,
    missing_export_timeout_seconds: int = 180,
    missing_export_apply_token: str = "",
    target_card_query: str = "",
    target_card_id: str = "",
) -> dict[str, Any]:
    """Preview phone/Photos-library photo matches to Trello cards.

    This is a guarded dry-run. It reads local Photos metadata, geocodes card
    address candidates into a local cache, stages matched local originals into a
    local plan folder, and writes a JSON plan. When explicitly requested, it can
    export matched cloud-backed originals into that local plan folder before
    staging. It does not write to Trello.
    """
    if days_back < 0:
        raise TrelloError("days_back must be >= 0")
    if radius_ft < 25 or radius_ft > 2000:
        raise TrelloError("radius_ft must be between 25 and 2000")
    required_export_token = "EXPORT_PHOTOS_FROM_ICLOUD"
    if export_missing_originals and missing_export_apply_token != required_export_token:
        raise TrelloError(f"Refusing Photos original export. Re-run with missing_export_apply_token={required_export_token!r}.")
    batch_id = datetime.now().strftime("photo-card-match-%Y%m%d-%H%M%S")
    plan_dir = PHOTO_MATCH_ROOT / batch_id
    plan_dir.mkdir(parents=True, exist_ok=True)

    source_root = Path(source_path or str(DEFAULT_PHOTOS_LIBRARY_PATH)).expanduser()
    photos = _scan_photo_source(str(source_root), days_back=days_back, max_files=max_photos, require_local=False)
    local_photos = [photo for photo in photos if photo.get("local_exists", True)]
    gps_photos = [photo for photo in photos if photo.get("has_gps")]
    gps_missing_originals = [photo for photo in gps_photos if not photo.get("local_exists", True)]
    card_location_payload = _card_locations_for_photo_match(
        board,
        include_complete=include_complete,
        limit_cards=limit_cards,
        geocode_limit=geocode_limit,
        target_card_query=target_card_query,
        target_card_id=target_card_id,
    )
    locations = card_location_payload["locations"]

    staged_count = 0
    matches_by_card: dict[str, dict[str, Any]] = {}
    unmatched_photo_count = 0
    matched_photo_items: list[dict[str, Any]] = []
    for photo in gps_photos:
        ranked = []
        for loc in locations:
            if loc.get("latitude") is None or loc.get("longitude") is None:
                continue
            distance = _distance_ft(float(photo["latitude"]), float(photo["longitude"]), float(loc["latitude"]), float(loc["longitude"]))
            if distance <= radius_ft:
                ranked.append({**loc, "distance_ft": round(distance, 1)})
        if not ranked:
            unmatched_photo_count += 1
            continue
        ranked.sort(key=lambda item: item["distance_ft"])
        best = ranked[0]
        second_distance = ranked[1]["distance_ft"] if len(ranked) > 1 else None
        confidence, warnings = _photo_confidence(float(best["distance_ft"]), second_distance)
        card_id = str(best["card_id"])
        group = matches_by_card.setdefault(
            card_id,
            {
                "card_id": card_id,
                "card_name": best.get("card_name"),
                "card_url": best.get("card_url"),
                "list": best.get("list"),
                "address": best.get("address"),
                "geocode": best.get("geocode"),
                "photos": [],
            },
        )
        photo_id = str(photo.get("photo_id") or photo.get("sha256") or photo.get("path"))
        photo_item = {
            "photo_id": photo_id,
            "filename": photo.get("filename"),
            "source_path": photo.get("path"),
            "local_exists": photo.get("local_exists", True),
            "cloud_local_state": photo.get("cloud_local_state"),
            "taken_at": photo.get("taken_at"),
            "latitude": photo.get("latitude"),
            "longitude": photo.get("longitude"),
            "size": photo.get("size"),
            "sha256": photo.get("sha256"),
            "width": photo.get("width"),
            "height": photo.get("height"),
            "distance_ft": best.get("distance_ft"),
            "confidence": confidence,
            "warnings": warnings,
            "cover_candidate_score": photo.get("cover_candidate_score") or 0,
            "all_matches": [
                {
                    "card_id": item.get("card_id"),
                    "card_name": item.get("card_name"),
                    "distance_ft": item.get("distance_ft"),
                    "address": item.get("address"),
                }
                for item in ranked[:5]
            ],
            "staged": False,
        }
        group["photos"].append(photo_item)
        matched_photo_items.append(photo_item)

    missing_export_result = None
    if stage_files and export_missing_originals and _looks_like_photos_library(source_root):
        missing_export_result = _export_missing_originals_for_photo_plan(
            photos=matched_photo_items,
            photos_library_path=source_root,
            plan_dir=plan_dir,
            limit=min(max(1, int(missing_export_limit)), max(1, int(max_stage_files))),
            timeout_seconds=missing_export_timeout_seconds,
        )
        exported_by_photo_id = missing_export_result.get("exported_by_photo_id") or {}
        for photo_item in matched_photo_items:
            exported = exported_by_photo_id.get(str(photo_item.get("photo_id")))
            if not exported:
                continue
            photo_item["source_path"] = exported.get("exported_path")
            photo_item["exported_original"] = True
            photo_item["exported_original_size"] = exported.get("size")
            photo_item["exported_original_uti"] = exported.get("uti")

    if stage_files:
        for group in matches_by_card.values():
            for photo_item in group["photos"]:
                if staged_count >= max(0, int(max_stage_files)):
                    break
                stage_result = _stage_photo_for_plan(photo_item, str(group.get("card_name") or group.get("card_id")), plan_dir)
                photo_item.update(stage_result)
                if stage_result.get("staged"):
                    staged_count += 1

    cards = list(matches_by_card.values())
    for group in cards:
        group["photos"].sort(key=lambda item: (item.get("confidence") != "very_strong", -(item.get("cover_candidate_score") or 0), item.get("taken_at") or ""))
        staged_photos = [item for item in group["photos"] if item.get("staged")]
        cover = max(staged_photos or group["photos"], key=lambda item: item.get("cover_candidate_score") or 0, default=None)
        group["selected_cover_photo_id"] = cover.get("photo_id") if cover else None
        group["photo_count"] = len(group["photos"])
        group["staged_photo_count"] = len(staged_photos)

    cards.sort(key=lambda item: (item.get("list") or "", item.get("card_name") or ""))
    plan = {
        "ok": True,
        "mode": "photo_card_match_plan",
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "plan_dir": str(plan_dir),
        "source": {
            "source_path": str(source_root),
            "days_back": days_back,
            "max_photos": max_photos,
            "photos_scanned": len(photos),
            "local_photos_with_originals": len(local_photos),
            "gps_photos": len(gps_photos),
            "gps_local_photos": len([photo for photo in gps_photos if photo.get("local_exists", True)]),
            "gps_missing_originals": len(gps_missing_originals),
        },
        "board": {"alias": board, "id": _board_id(board), "include_complete": include_complete},
        "settings": {
            "radius_ft": radius_ft,
            "limit_cards": limit_cards,
            "geocode_limit": geocode_limit,
            "stage_files": stage_files,
            "max_stage_files": max_stage_files,
            "export_missing_originals": export_missing_originals,
            "missing_export_limit": missing_export_limit,
            "missing_export_timeout_seconds": missing_export_timeout_seconds,
            "target_card_query": (target_card_query or "").strip() or None,
            "target_card_id": (target_card_id or "").strip() or None,
        },
        "card_location_summary": {
            "cards_seen": card_location_payload.get("cards_seen"),
            "cards_available": card_location_payload.get("cards_available"),
            "geocoded_card_locations": len(locations),
            "skipped": card_location_payload.get("skipped"),
            "geocode_attempts": card_location_payload.get("geocode_attempts"),
            "target": card_location_payload.get("target"),
        },
        "summary": {
            "matched_cards": len(cards),
            "matched_photos": sum(len(group["photos"]) for group in cards),
            "staged_photos": staged_count,
            "missing_original_matched_photos": sum(
                1 for group in cards for photo in group["photos"] if not photo.get("local_exists", True)
            ),
            "exported_missing_originals": (missing_export_result or {}).get("exported_count", 0),
            "unmatched_gps_photos": unmatched_photo_count,
        },
        "missing_original_export": missing_export_result,
        "cards": cards,
        "safety": {
            "read_only_preview": True,
            "local_file_write": bool(stage_files),
            "icloud_network_access": bool(export_missing_originals),
            "trello_writes": False,
            "photos_writes": False,
            "google_drive_writes": False,
            "secrets_returned": False,
            "apply_requires_token": "APPLY_PHOTO_CARD_MATCH_PLAN",
        },
    }
    json_path = plan_dir / "photo-card-match-plan.json"
    _write_json_file(json_path, plan)
    markdown_path = plan_dir / "photo-card-match-plan.md"
    lines = [
        "# Photo Card Match Plan",
        "",
        f"- Source: {plan['source']['source_path']}",
        f"- Board: {board}",
        f"- Target card query: {(target_card_query or '').strip() or 'none'}",
        f"- Target card id: {(target_card_id or '').strip() or 'none'}",
        f"- Matched cards: {plan['summary']['matched_cards']}",
        f"- Matched photos: {plan['summary']['matched_photos']}",
        f"- Staged photos: {plan['summary']['staged_photos']}",
        f"- Missing-original matched photos: {plan['summary']['missing_original_matched_photos']}",
        f"- Exported missing originals: {plan['summary']['exported_missing_originals']}",
        "",
    ]
    for group in cards[:80]:
        lines.append(f"## {group.get('card_name')}")
        lines.append(f"- Card: {group.get('card_url')}")
        lines.append(f"- Address: {group.get('address')}")
        lines.append(f"- Photos: {group.get('photo_count')} staged {group.get('staged_photo_count')}")
        if group.get("selected_cover_photo_id"):
            lines.append(f"- Cover candidate: {group.get('selected_cover_photo_id')}")
        lines.append("")
        for photo in group["photos"][:12]:
            lines.append(f"  - {photo.get('filename')} | {photo.get('confidence')} | {photo.get('distance_ft')} ft | staged={photo.get('staged')}")
        lines.append("")
    markdown_path.write_text("\n".join(lines), encoding="utf-8")
    return {
        "ok": True,
        "plan_json_path": str(json_path),
        "plan_markdown_path": str(markdown_path),
        "summary": plan["summary"],
        "source": plan["source"],
        "card_location_summary": plan["card_location_summary"],
        "sample_cards": cards[:10],
        "safety": plan["safety"],
    }


@mcp.tool()
def apply_photo_card_match_plan(
    plan_json_path: str,
    dry_run: bool = True,
    apply_token: str = "",
    allowed_confidences_json: str = '["very_strong", "likely"]',
    limit_cards: int = 25,
    limit_photos_per_card: int = 10,
    set_covers: bool = True,
    skip_existing_names: bool = True,
) -> dict[str, Any]:
    """Apply a previewed photo-card match plan to Trello.

    Defaults to dry-run. Real Trello attachment/cover writes only happen when
    dry_run is false and apply_token is APPLY_PHOTO_CARD_MATCH_PLAN.
    """
    plan = _load_photo_match_plan(plan_json_path)
    try:
        allowed_raw = json.loads(allowed_confidences_json)
    except json.JSONDecodeError as exc:
        raise TrelloError("allowed_confidences_json must be a JSON array") from exc
    allowed = {str(item).strip() for item in allowed_raw if str(item).strip()}
    valid = {"very_strong", "likely", "review"}
    invalid = allowed - valid
    if invalid:
        raise TrelloError(f"Unsupported confidence values: {sorted(invalid)}")
    if not allowed:
        raise TrelloError("At least one confidence must be allowed")

    apply_enabled = not dry_run
    required_token = "APPLY_PHOTO_CARD_MATCH_PLAN"
    if apply_enabled and apply_token != required_token:
        raise TrelloError(f"Refusing Trello writes. Re-run with apply_token={required_token!r} to apply photo attachments.")

    results = []
    errors = []
    cards = (plan.get("cards") or [])[: max(1, min(int(limit_cards), 100))]
    for group in cards:
        card_id = str(group.get("card_id") or "")
        if not card_id:
            continue
        existing_by_name: dict[str, dict[str, Any]] = {}
        if apply_enabled and skip_existing_names:
            try:
                for item in get_card_attachments(card_id):
                    if item.get("name"):
                        existing_by_name[str(item["name"])] = item
            except Exception as exc:
                errors.append({"card_id": card_id, "stage": "list_existing_attachments", "error": str(exc)[:500]})
                continue
        card_result = {
            "card_id": card_id,
            "card_name": group.get("card_name"),
            "card_url": group.get("card_url"),
            "uploads": [],
            "cover": None,
        }
        selected_cover_id = str(group.get("selected_cover_photo_id") or "")
        selected_cover_attachment_id = None
        uploadable_seen = 0
        upload_limit = max(1, min(int(limit_photos_per_card), 50))
        for photo in group.get("photos") or []:
            if str(photo.get("confidence")) not in allowed:
                continue
            staged_path = Path(str(photo.get("staged_path") or "")).expanduser()
            if not staged_path.is_file():
                card_result["uploads"].append({"photo_id": photo.get("photo_id"), "status": "skipped", "reason": "staged_file_missing"})
                continue
            if uploadable_seen >= upload_limit:
                card_result["uploads"].append({"photo_id": photo.get("photo_id"), "status": "skipped", "reason": "limit_photos_per_card"})
                continue
            uploadable_seen += 1
            taken = str(photo.get("taken_at") or "unknown-date")[:10]
            attachment_name = f"Pool photo {taken} {str(photo.get('photo_id') or staged_path.stem)[:8]}{staged_path.suffix.lower()}"[:256]
            if dry_run:
                row = {
                    "photo_id": photo.get("photo_id"),
                    "status": "would_upload",
                    "attachment_name": attachment_name,
                    "staged_path": str(staged_path),
                    "confidence": photo.get("confidence"),
                    "distance_ft": photo.get("distance_ft"),
                }
            elif skip_existing_names and attachment_name in existing_by_name:
                existing = existing_by_name[attachment_name]
                row = {
                    "photo_id": photo.get("photo_id"),
                    "status": "skipped_existing_name",
                    "attachment_name": attachment_name,
                    "attachment_id": existing.get("id"),
                }
                if str(photo.get("photo_id")) == selected_cover_id:
                    selected_cover_attachment_id = existing.get("id")
            else:
                try:
                    uploaded = attach_file_to_card(
                        file_path=str(staged_path),
                        card_id=card_id,
                        name=attachment_name,
                        comment=None,
                    )
                    attachment = uploaded.get("attachment") or {}
                    row = {
                        "photo_id": photo.get("photo_id"),
                        "status": "uploaded",
                        "attachment_name": attachment_name,
                        "attachment_id": attachment.get("id"),
                    }
                    if str(photo.get("photo_id")) == selected_cover_id:
                        selected_cover_attachment_id = attachment.get("id")
                except Exception as exc:
                    row = {"photo_id": photo.get("photo_id"), "status": "error", "error": str(exc)[:500]}
                    errors.append({"card_id": card_id, "photo_id": photo.get("photo_id"), "stage": "upload", "error": str(exc)[:500]})
            card_result["uploads"].append(row)
        if dry_run:
            card_result["cover"] = {"status": "would_set" if set_covers and selected_cover_id else "not_requested", "photo_id": selected_cover_id or None}
        elif set_covers and selected_cover_attachment_id:
            try:
                cover = set_card_cover(card_id=card_id, attachment_id=str(selected_cover_attachment_id))
                card_result["cover"] = {"status": "set", "attachment_id": cover.get("attachment_id")}
            except Exception as exc:
                card_result["cover"] = {"status": "error", "error": str(exc)[:500]}
                errors.append({"card_id": card_id, "stage": "set_cover", "error": str(exc)[:500]})
        results.append(card_result)

    upload_count = sum(1 for card in results for item in card.get("uploads", []) if item.get("status") == "uploaded")
    would_upload_count = sum(1 for card in results for item in card.get("uploads", []) if item.get("status") == "would_upload")
    return {
        "ok": not errors,
        "mode": "dry_run" if dry_run else "applied",
        "plan_json_path": str(Path(plan_json_path).expanduser()),
        "allowed_confidences": sorted(allowed),
        "summary": {
            "cards_considered": len(results),
            "uploaded": upload_count,
            "would_upload": would_upload_count,
            "errors": len(errors),
        },
        "results": results,
        "errors": errors,
        "safety": {
            "trello_writes": bool(apply_enabled),
            "attachments_written": upload_count if apply_enabled else 0,
            "cover_writes": bool(apply_enabled and set_covers),
            "photos_writes": False,
            "google_drive_writes": False,
            "secrets_returned": False,
            "dry_run_default": True,
            "required_apply_token": required_token,
        },
    }


@mcp.tool()
def search_pool_records(
    query: str,
    record_kind: str = "any",
    year: str | None = None,
    limit: int = 8,
    min_relevance: float = 0.0,
) -> dict[str, Any]:
    """Search indexed Memphis/Artesian work orders and bills from the local RAG corpus.

    record_kind can be any, work_order, or bill. year should be a four-digit
    year when supplied. This is read-only and restricted to the pool_jobs source.
    """
    q = (query or "").strip()
    if not q:
        raise TrelloError("query is required")
    kind = (record_kind or "any").strip().lower()
    if kind not in {"any", "work_order", "bill"}:
        raise TrelloError("record_kind must be one of: any, work_order, bill")
    if year is not None and not re.fullmatch(r"20\d{2}", str(year)):
        raise TrelloError("year must be a four-digit 20xx year")

    requested_limit = max(1, min(int(limit), 25))
    search_limit = requested_limit if kind == "any" and year is None else min(max(requested_limit * 4, 12), 50)
    args = ["search", q, "--source", "pool_jobs", "--top-k", str(search_limit)]
    if min_relevance:
        args.extend(["--min-relevance", str(min_relevance)])
    payload = _run_rag(args, timeout=180)

    filtered = []
    for result in payload.get("results") or []:
        result_kind = _pool_record_kind(result)
        result_year = _pool_record_year(result)
        if kind != "any" and result_kind != kind:
            continue
        if year is not None and result_year != str(year):
            continue
        filtered.append(_pool_record_summary(result))
        if len(filtered) >= requested_limit:
            break

    return {
        "ok": True,
        "query": q,
        "filters": {"record_kind": kind, "year": str(year) if year is not None else None},
        "source": "pool_jobs",
        "returned": len(filtered),
        "searched_results": len(payload.get("results") or []),
        "results": filtered,
    }


@mcp.tool()
def ask_pool_records(question: str, top_k: int = 8, min_relevance: float = 0.0) -> dict[str, Any]:
    """Ask a natural-language question against indexed pool work orders and bills.

    This is read-only and restricted to the pool_jobs source. Use search_pool_records
    first when you need exact files or tight bill/work-order/year filters.
    """
    q = (question or "").strip()
    if not q:
        raise TrelloError("question is required")
    args = ["ask", q, "--source", "pool_jobs", "--top-k", str(max(1, min(int(top_k), 12)))]
    if min_relevance:
        args.extend(["--min-relevance", str(min_relevance)])
    payload = _run_rag(args, timeout=240)
    return {
        "ok": True,
        "question": q,
        "source": "pool_jobs",
        "answer": payload.get("answer"),
        "sources": [_pool_record_summary(result, excerpt_chars=700) for result in payload.get("results") or []],
        "errors": payload.get("errors") or [],
    }


@mcp.tool()
def find_records_for_card(
    card_id: str,
    record_kind: str = "any",
    include_attachment_names: bool = True,
    limit: int = 8,
    min_relevance: float = 0.0,
) -> dict[str, Any]:
    """Find likely indexed work-order/bill records for one Trello card.

    This is read-only. It reads card fields and optional attachment names, searches
    the local pool_jobs RAG source, and returns review candidates with confidence.
    It never writes to Trello and never attaches records automatically.
    """
    card = get_card(card_id)
    list_name = None
    if card.get("idList"):
        try:
            list_info = _request("GET", f"lists/{card['idList']}", params={"fields": "id,name"})
            list_name = list_info.get("name")
        except Exception:
            list_name = None

    attachments = get_card_attachments(card["id"]) if include_attachment_names else []
    query = _card_record_search_text(card, list_name=list_name, attachments=attachments)
    kind = (record_kind or "any").strip().lower()
    if kind not in {"any", "work_order", "bill"}:
        raise TrelloError("record_kind must be one of: any, work_order, bill")
    if kind == "any" and list_name and "BILL" in list_name.upper():
        kind = "bill"
    if not query:
        return {
            "ok": True,
            "target": {
                "card_id": card.get("id"),
                "card_name": card.get("name"),
                "card_url": card.get("shortUrl"),
                "list": list_name,
            },
            "query": "",
            "summary": {"candidates": 0, "best_confidence": "no_match"},
            "candidates": [],
            "note": "Card had no useful searchable text.",
        }

    search = search_pool_records(
        query,
        record_kind=kind,
        year=None,
        limit=max(1, min(int(limit), 25)),
        min_relevance=min_relevance,
    )
    lexical = _lexical_pool_record_search(
        query,
        record_kind=kind,
        max_results=max(5, min(int(limit) * 3, 50)),
    )
    card_tokens = _match_tokens(query)
    strong_tokens = _strong_card_tokens(card, query)
    name_anchor_tokens = _card_name_anchor_tokens(card)
    if kind == "bill" or (list_name and "BILL" in list_name.upper()):
        name_anchor_tokens = set()
    merged_records: dict[str, dict[str, Any]] = {}
    for record in search.get("results") or []:
        path = str(record.get("path") or record.get("file") or "")
        if path:
            merged_records[path] = {**record, "retrieval_sources": ["rag"]}
    for record in lexical:
        path = str(record.get("path") or record.get("file") or "")
        if not path:
            continue
        if path in merged_records:
            existing = merged_records[path]
            existing["score"] = max(float(existing.get("score") or 0.0), float(record.get("score") or 0.0))
            existing.setdefault("retrieval_sources", ["rag"]).append("lexical")
            if len(str(record.get("text_excerpt") or "")) > len(str(existing.get("text_excerpt") or "")):
                existing["text_excerpt"] = record.get("text_excerpt")
        else:
            merged_records[path] = {**record, "retrieval_sources": ["lexical"]}

    candidates = []
    for record in merged_records.values():
        detail = _record_match_detail(
            card_tokens,
            record,
            strong_tokens=strong_tokens,
            name_anchor_tokens=name_anchor_tokens,
        )
        candidates.append({**record, "match": detail})

    candidates.sort(
        key=lambda item: (
            _confidence_rank(item.get("match", {}).get("confidence", "no_match")),
            float(item.get("score") or 0.0),
            item.get("match", {}).get("overlap_count", 0),
        ),
        reverse=True,
    )
    candidates = candidates[: max(1, min(int(limit), 25))]

    best_confidence = candidates[0]["match"]["confidence"] if candidates else "no_match"
    by_kind: dict[str, int] = {}
    by_confidence: dict[str, int] = {}
    for item in candidates:
        kind = item.get("kind") or "unknown"
        confidence = item.get("match", {}).get("confidence", "unknown")
        by_kind[kind] = by_kind.get(kind, 0) + 1
        by_confidence[confidence] = by_confidence.get(confidence, 0) + 1

    return {
        "ok": True,
        "target": {
            "card_id": card.get("id"),
            "card_name": card.get("name"),
            "card_url": card.get("shortUrl"),
            "list": list_name,
            "labels": _label_names(card),
            "due": card.get("due"),
            "dateLastActivity": card.get("dateLastActivity"),
            "attachment_names_used": [item.get("name") for item in attachments if item.get("name")],
        },
        "query": query,
        "source": "pool_jobs",
        "filters": {"record_kind": kind},
        "summary": {
            "candidates": len(candidates),
            "best_confidence": best_confidence,
            "by_kind": by_kind,
            "by_confidence": by_confidence,
            "rag_candidates_seen": search.get("searched_results"),
            "lexical_candidates_seen": len(lexical),
        },
        "candidates": candidates,
        "safety": {
            "read_only": True,
            "trello_writes": False,
            "auto_attach": False,
        },
    }


@mcp.tool()
def review_board_record_matches(
    board: str = "memphis",
    record_kind: str = "any",
    include_complete: bool = False,
    list_names_json: str = "[]",
    limit_cards: int = 20,
    max_candidates_per_card: int = 2,
    min_relevance: float = 0.0,
) -> dict[str, Any]:
    """Review likely pool record matches across a board, read-only.

    Scans a capped set of cards, runs find_records_for_card for each, and
    groups results by best confidence. list_names_json may be a JSON array of
    Trello list names to restrict the scan, for example ["MEASURES", "BILLS"].
    """
    kind = (record_kind or "any").strip().lower()
    if kind not in {"any", "work_order", "bill"}:
        raise TrelloError("record_kind must be one of: any, work_order, bill")

    try:
        selected_lists = json.loads(list_names_json or "[]")
    except json.JSONDecodeError as exc:
        raise TrelloError("list_names_json must be a JSON array of list names") from exc
    if not isinstance(selected_lists, list):
        raise TrelloError("list_names_json must be a JSON array of list names")
    selected_list_names = {str(item).strip() for item in selected_lists if str(item).strip()}

    board_info = get_board(board)
    lists = get_lists(board, filter="all" if include_complete else "open")
    list_names = {item["id"]: item["name"] for item in lists}
    cards = _request(
        "GET",
        f"boards/{_board_id(board)}/cards",
        params={
            "filter": "all" if include_complete else "open",
            "fields": "id,name,desc,shortUrl,idList,closed,due,labels,dateLastActivity",
        },
    )

    scoped_cards = []
    for card in cards:
        list_name = list_names.get(card.get("idList"), "UNKNOWN")
        if card.get("closed") and not include_complete:
            continue
        if not include_complete and list_name == "Complete":
            continue
        if selected_list_names and list_name not in selected_list_names:
            continue
        scoped_cards.append(card)

    scoped_card_count = len(scoped_cards)
    scoped_cards.sort(key=lambda item: item.get("dateLastActivity") or "", reverse=True)
    scoped_cards = scoped_cards[: max(1, min(int(limit_cards), 100))]

    buckets = _empty_match_buckets()
    by_list: dict[str, dict[str, int]] = {}
    errors = []
    cards_reviewed = 0
    card_limit = max(1, min(int(max_candidates_per_card), 8))

    for card in scoped_cards:
        list_name = list_names.get(card.get("idList"), "UNKNOWN")
        list_row = by_list.setdefault(
            list_name,
            {"cards": 0, "exact": 0, "likely": 0, "review": 0, "weak": 0, "no_match": 0},
        )
        list_row["cards"] += 1
        cards_reviewed += 1

        try:
            effective_kind = "bill" if kind == "any" and "BILL" in list_name.upper() else kind
            match = find_records_for_card(
                card["id"],
                record_kind=effective_kind,
                include_attachment_names=True,
                limit=max(3, card_limit),
                min_relevance=min_relevance,
            )
            confidence = match.get("summary", {}).get("best_confidence") or "no_match"
            if confidence not in buckets:
                confidence = "weak"
            list_row[confidence] += 1
            buckets[confidence].append(
                {
                    "card_id": card.get("id"),
                    "card_name": card.get("name"),
                    "card_url": card.get("shortUrl"),
                    "list": list_name,
                    "labels": _label_names(card),
                    "due": card.get("due"),
                    "dateLastActivity": card.get("dateLastActivity"),
                    "query": match.get("query"),
                    "candidate_count": match.get("summary", {}).get("candidates", 0),
                    "by_kind": match.get("summary", {}).get("by_kind", {}),
                    "top_candidates": [
                        _record_candidate_preview(item)
                        for item in (match.get("candidates") or [])[:card_limit]
                    ],
                }
            )
        except Exception as exc:
            list_row["no_match"] += 1
            errors.append(
                {
                    "card_id": card.get("id"),
                    "card_name": card.get("name"),
                    "card_url": card.get("shortUrl"),
                    "list": list_name,
                    "error": str(exc)[:500],
                }
            )
            buckets["no_match"].append(
                {
                    "card_id": card.get("id"),
                    "card_name": card.get("name"),
                    "card_url": card.get("shortUrl"),
                    "list": list_name,
                    "candidate_count": 0,
                    "top_candidates": [],
                    "error": str(exc)[:500],
                }
            )

    bucket_counts = {name: len(items) for name, items in buckets.items()}
    needs_human_review = bucket_counts["review"] + bucket_counts["weak"] + bucket_counts["no_match"]

    return {
        "ok": True,
        "board": {k: board_info.get(k) for k in ("id", "name", "url", "closed", "dateLastActivity")},
        "scope": {
            "record_kind": kind,
            "include_complete": include_complete,
            "selected_lists": sorted(selected_list_names),
            "cards_available_after_scope": scoped_card_count,
            "cards_reviewed": cards_reviewed,
            "limit_cards": limit_cards,
            "max_candidates_per_card": card_limit,
            "source": "pool_jobs",
        },
        "summary": {
            "bucket_counts": bucket_counts,
            "actionable_without_record_writes": bucket_counts["exact"] + bucket_counts["likely"],
            "needs_human_review": needs_human_review,
            "errors": len(errors),
        },
        "by_list": by_list,
        "buckets": buckets,
        "errors": errors,
        "safety": {
            "read_only": True,
            "trello_writes": False,
            "google_drive_writes": False,
            "auto_attach": False,
        },
    }


def _format_board_record_match_report(review: dict[str, Any]) -> str:
    board = review.get("board") or {}
    scope = review.get("scope") or {}
    summary = review.get("summary") or {}
    bucket_counts = summary.get("bucket_counts") or {}
    lines = [
        f"# Board Record Match Report - {board.get('name') or 'UNKNOWN'}",
        "",
        f"- Generated: {datetime.now().isoformat(timespec='seconds')}",
        f"- Board URL: {board.get('url') or 'UNKNOWN'}",
        f"- Source: {scope.get('source') or 'UNKNOWN'}",
        f"- Record kind: {scope.get('record_kind') or 'any'}",
        f"- Include complete: {scope.get('include_complete')}",
        f"- Selected lists: {', '.join(scope.get('selected_lists') or []) or 'all active lists'}",
        f"- Cards reviewed: {scope.get('cards_reviewed')} of {scope.get('cards_available_after_scope')}",
        "",
        "## Summary",
        "",
        f"- Exact: {bucket_counts.get('exact', 0)}",
        f"- Likely: {bucket_counts.get('likely', 0)}",
        f"- Review: {bucket_counts.get('review', 0)}",
        f"- Weak: {bucket_counts.get('weak', 0)}",
        f"- No match: {bucket_counts.get('no_match', 0)}",
        f"- Actionable without Trello writes: {summary.get('actionable_without_record_writes', 0)}",
        f"- Needs human review: {summary.get('needs_human_review', 0)}",
        f"- Errors: {summary.get('errors', 0)}",
        "",
        "## By List",
        "",
    ]

    for list_name, row in sorted((review.get("by_list") or {}).items()):
        lines.append(
            f"- {list_name}: {row.get('cards', 0)} cards; "
            f"exact {row.get('exact', 0)}, likely {row.get('likely', 0)}, "
            f"review {row.get('review', 0)}, weak {row.get('weak', 0)}, no_match {row.get('no_match', 0)}"
        )

    for bucket in ("exact", "likely", "review", "weak", "no_match"):
        lines.extend(["", f"## {bucket.replace('_', ' ').title()}", ""])
        items = (review.get("buckets") or {}).get(bucket) or []
        if not items:
            lines.append("- None")
            continue
        for item in items:
            lines.append(f"### {item.get('card_name') or 'Untitled card'}")
            lines.append("")
            lines.append(f"- Card: {item.get('card_url') or item.get('card_id')}")
            lines.append(f"- List: {item.get('list') or 'UNKNOWN'}")
            lines.append(f"- Candidate count: {item.get('candidate_count', 0)}")
            if item.get("error"):
                lines.append(f"- Error: {item.get('error')}")
            for index, candidate in enumerate(item.get("top_candidates") or [], start=1):
                lines.append(f"- Candidate {index}: {candidate.get('file') or 'UNKNOWN'}")
                lines.append(f"  - Kind: {candidate.get('kind') or 'unknown'}")
                lines.append(f"  - Year: {candidate.get('year') or 'unknown'}")
                lines.append(f"  - Confidence: {candidate.get('confidence') or 'unknown'}")
                lines.append(f"  - Reason: {candidate.get('reason') or 'unknown'}")
                lines.append(f"  - Path: {candidate.get('path') or 'UNKNOWN'}")
                overlap = ", ".join(candidate.get("overlap_tokens") or [])
                lines.append(f"  - Overlap: {overlap or 'none'}")
            lines.append("")

    if review.get("errors"):
        lines.extend(["", "## Errors", ""])
        for error in review.get("errors") or []:
            lines.append(f"- {error.get('card_name') or error.get('card_id')}: {error.get('error')}")

    lines.extend(
        [
            "",
            "## Safety",
            "",
            "- Trello writes: false",
            "- Google Drive writes: false",
            "- Auto-attach: false",
        ]
    )
    return "\n".join(lines).rstrip() + "\n"


def _record_match_marker(card_id: str, record_path: str) -> str:
    seed = f"{card_id}|{record_path}".encode("utf-8", errors="ignore")
    digest = hashlib.sha256(seed).hexdigest()[:16]
    return f"codex-record-match:{digest}"


def _card_comment_actions(card_id: str, limit: int = 100) -> list[dict[str, Any]]:
    actions = _request(
        "GET",
        f"cards/{card_id}/actions",
        params={"filter": "commentCard", "limit": max(1, min(int(limit), 1000))},
    )
    return actions if isinstance(actions, list) else []


def _card_has_comment_marker(card_id: str, marker: str) -> bool:
    for action in _card_comment_actions(card_id):
        text = str((action.get("data") or {}).get("text") or "")
        if marker in text:
            return True
    return False


def _proposed_match_actions(review: dict[str, Any], allowed_confidences: set[str]) -> list[dict[str, Any]]:
    actions = []
    for bucket in ("exact", "likely", "review", "weak", "no_match"):
        if bucket not in allowed_confidences:
            continue
        for item in (review.get("buckets") or {}).get(bucket) or []:
            candidates = item.get("top_candidates") or []
            if not candidates:
                continue
            candidate = candidates[0]
            confidence = candidate.get("confidence") or bucket
            if confidence not in allowed_confidences:
                continue
            record_path = candidate.get("path") or ""
            record_name = candidate.get("file") or "matched pool record"
            marker = _record_match_marker(str(item.get("card_id") or ""), record_path)
            comment = (
                "Historical pool record candidate found.\n\n"
                f"- Confidence: {confidence}\n"
                f"- Record: {record_name}\n"
                f"- Kind: {candidate.get('kind') or 'unknown'}\n"
                f"- Year: {candidate.get('year') or 'unknown'}\n"
                f"- Source path: {record_path}\n"
                f"- Marker: {marker}\n\n"
                "Review before attaching or treating this as final."
            )
            actions.append(
                {
                    "card_id": item.get("card_id"),
                    "card_name": item.get("card_name"),
                    "card_url": item.get("card_url"),
                    "list": item.get("list"),
                    "confidence": confidence,
                    "reason": candidate.get("reason"),
                    "record_kind": candidate.get("kind"),
                    "record_year": candidate.get("year"),
                    "record_file": record_name,
                    "record_path": record_path,
                    "idempotency_marker": marker,
                    "proposed_actions": [
                        {
                            "type": "comment",
                            "status": "proposed_only",
                            "text": comment,
                        },
                        {
                            "type": "attachment_review",
                            "status": "proposed_only",
                            "note": "Local record path identified; no Trello attachment was created.",
                            "path": record_path,
                        },
                    ],
                }
            )
    actions.sort(
        key=lambda item: (
            _confidence_rank(str(item.get("confidence") or "")),
            str(item.get("card_name") or ""),
        ),
        reverse=True,
    )
    return actions


def _format_record_match_action_plan(review: dict[str, Any], actions: list[dict[str, Any]]) -> str:
    board = review.get("board") or {}
    scope = review.get("scope") or {}
    lines = [
        f"# Proposed Record Match Actions - {board.get('name') or 'UNKNOWN'}",
        "",
        f"- Generated: {datetime.now().isoformat(timespec='seconds')}",
        f"- Board URL: {board.get('url') or 'UNKNOWN'}",
        f"- Selected lists: {', '.join(scope.get('selected_lists') or []) or 'all active lists'}",
        f"- Cards reviewed: {scope.get('cards_reviewed')} of {scope.get('cards_available_after_scope')}",
        f"- Proposed actions: {len(actions)}",
        "",
        "## Important",
        "",
        "- This is a proposal only.",
        "- No Trello comments were added.",
        "- No Trello attachments were created.",
        "- No Google Drive writes were performed.",
        "",
        "## Actions",
        "",
    ]

    if not actions:
        lines.append("- None")
    for index, action in enumerate(actions, start=1):
        comment = next(
            (item for item in action.get("proposed_actions") or [] if item.get("type") == "comment"),
            {},
        )
        lines.append(f"### {index}. {action.get('card_name') or 'Untitled card'}")
        lines.append("")
        lines.append(f"- Card: {action.get('card_url') or action.get('card_id')}")
        lines.append(f"- List: {action.get('list') or 'UNKNOWN'}")
        lines.append(f"- Confidence: {action.get('confidence') or 'unknown'}")
        lines.append(f"- Record: {action.get('record_file') or 'UNKNOWN'}")
        lines.append(f"- Record kind: {action.get('record_kind') or 'unknown'}")
        lines.append(f"- Record year: {action.get('record_year') or 'unknown'}")
        lines.append(f"- Record path: {action.get('record_path') or 'UNKNOWN'}")
        lines.append(f"- Marker: {action.get('idempotency_marker') or 'missing'}")
        lines.append("")
        lines.append("Proposed comment:")
        lines.append("")
        lines.append("```text")
        lines.append(str(comment.get("text") or "").rstrip())
        lines.append("```")
        lines.append("")

    lines.extend(
        [
            "## Safety",
            "",
            "- Trello writes: false",
            "- Google Drive writes: false",
            "- Auto-attach: false",
        ]
    )
    return "\n".join(lines).rstrip() + "\n"


@mcp.tool()
def write_board_record_match_report(
    board: str = "memphis",
    record_kind: str = "any",
    include_complete: bool = False,
    list_names_json: str = "[]",
    limit_cards: int = 25,
    max_candidates_per_card: int = 2,
    min_relevance: float = 0.0,
) -> dict[str, Any]:
    """Write a local Markdown/JSON audit report for board record matches.

    This wraps review_board_record_matches and writes local files under the MCP
    audit artifact folder. It does not write to Trello or Google Drive.
    """
    review = review_board_record_matches(
        board=board,
        record_kind=record_kind,
        include_complete=include_complete,
        list_names_json=list_names_json,
        limit_cards=limit_cards,
        max_candidates_per_card=max_candidates_per_card,
        min_relevance=min_relevance,
    )
    audit_dir = _artifact_dir("audits")
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    board_name = _safe_name((review.get("board") or {}).get("name") or board)
    selected = "-".join((review.get("scope") or {}).get("selected_lists") or ["all-active"])
    stem = _safe_name(f"board-record-matches-{board_name}-{selected}-{timestamp}")
    json_path = audit_dir / f"{stem}.json"
    md_path = audit_dir / f"{stem}.md"

    json_path.write_text(json.dumps(review, indent=2, sort_keys=True), encoding="utf-8")
    md_path.write_text(_format_board_record_match_report(review), encoding="utf-8")

    return {
        "ok": True,
        "markdown_path": str(md_path),
        "json_path": str(json_path),
        "summary": review.get("summary"),
        "scope": review.get("scope"),
        "safety": {
            "local_file_write": True,
            "trello_writes": False,
            "google_drive_writes": False,
            "auto_attach": False,
        },
    }


@mcp.tool()
def write_record_match_action_plan(
    board: str = "memphis",
    record_kind: str = "any",
    include_complete: bool = False,
    list_names_json: str = "[]",
    allowed_confidences_json: str = '["exact", "likely"]',
    limit_cards: int = 25,
    max_candidates_per_card: int = 2,
    min_relevance: float = 0.0,
) -> dict[str, Any]:
    """Write a local proposed-action plan for exact/likely record matches.

    This produces proposed Trello comment/attachment-review actions as local
    Markdown and JSON only. It does not write to Trello or Google Drive.
    """
    try:
        allowed_raw = json.loads(allowed_confidences_json or '["exact", "likely"]')
    except json.JSONDecodeError as exc:
        raise TrelloError("allowed_confidences_json must be a JSON array") from exc
    if not isinstance(allowed_raw, list):
        raise TrelloError("allowed_confidences_json must be a JSON array")
    allowed = {str(item).strip().lower() for item in allowed_raw if str(item).strip()}
    invalid = allowed - {"exact", "likely", "review", "weak", "no_match"}
    if invalid:
        raise TrelloError(f"Unsupported confidence values: {sorted(invalid)}")
    if not allowed:
        raise TrelloError("At least one allowed confidence is required")

    review = review_board_record_matches(
        board=board,
        record_kind=record_kind,
        include_complete=include_complete,
        list_names_json=list_names_json,
        limit_cards=limit_cards,
        max_candidates_per_card=max_candidates_per_card,
        min_relevance=min_relevance,
    )
    actions = _proposed_match_actions(review, allowed)
    payload = {
        "ok": True,
        "board": review.get("board"),
        "scope": {
            **(review.get("scope") or {}),
            "allowed_confidences": sorted(allowed),
        },
        "review_summary": review.get("summary"),
        "proposed_action_count": len(actions),
        "actions": actions,
        "safety": {
            "proposal_only": True,
            "local_file_write": True,
            "trello_writes": False,
            "google_drive_writes": False,
            "auto_attach": False,
        },
    }

    audit_dir = _artifact_dir("audits")
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    board_name = _safe_name((review.get("board") or {}).get("name") or board)
    selected = "-".join((review.get("scope") or {}).get("selected_lists") or ["all-active"])
    confidence_label = "-".join(sorted(allowed))
    stem = _safe_name(f"record-match-action-plan-{board_name}-{selected}-{confidence_label}-{timestamp}")
    json_path = audit_dir / f"{stem}.json"
    md_path = audit_dir / f"{stem}.md"

    json_path.write_text(json.dumps(payload, indent=2, sort_keys=True), encoding="utf-8")
    md_path.write_text(_format_record_match_action_plan(review, actions), encoding="utf-8")

    return {
        "ok": True,
        "markdown_path": str(md_path),
        "json_path": str(json_path),
        "summary": {
            "proposed_action_count": len(actions),
            "allowed_confidences": sorted(allowed),
            "review_summary": review.get("summary"),
        },
        "safety": payload["safety"],
    }


def _load_action_plan(plan_json_path: str) -> dict[str, Any]:
    path = Path(plan_json_path).expanduser()
    if not path.exists():
        raise TrelloError(f"Action plan JSON not found: {path}")
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise TrelloError(f"Action plan JSON is invalid: {path}") from exc
    if not isinstance(payload, dict) or not isinstance(payload.get("actions"), list):
        raise TrelloError("Action plan JSON must contain an actions array")
    return payload


def _comment_text_for_action(action: dict[str, Any]) -> str:
    for proposed in action.get("proposed_actions") or []:
        if proposed.get("type") == "comment" and proposed.get("text"):
            return str(proposed["text"])
    raise TrelloError(f"Action is missing proposed comment text: {action.get('card_name') or action.get('card_id')}")


def _marker_for_action(action: dict[str, Any]) -> str:
    marker = str(action.get("idempotency_marker") or "").strip()
    if marker:
        return marker
    card_id = str(action.get("card_id") or "").strip()
    record_path = str(action.get("record_path") or "").strip()
    if not card_id or not record_path:
        return ""
    return _record_match_marker(card_id, record_path)


def _comment_text_with_marker(action: dict[str, Any], marker: str) -> str:
    comment_text = _comment_text_for_action(action)
    if marker and marker not in comment_text:
        comment_text = f"{comment_text.rstrip()}\n- Marker: {marker}\n"
    return comment_text


@mcp.tool()
def apply_record_match_action_plan(
    plan_json_path: str,
    dry_run: bool = True,
    apply_token: str = "",
    limit_actions: int = 50,
    validate_trello_cards: bool = False,
) -> dict[str, Any]:
    """Validate or apply a record-match action plan.

    Defaults to dry-run. Real Trello comments are only created when dry_run is
    false and apply_token is exactly APPLY_RECORD_MATCH_ACTION_PLAN. This tool
    never creates Trello attachments; attachment entries remain review notes.
    Dry-runs validate local record paths by default and only read Trello cards
    when validate_trello_cards is true.
    """
    payload = _load_action_plan(plan_json_path)
    actions = payload.get("actions") or []
    actions = actions[: max(1, min(int(limit_actions), 100))]
    apply_enabled = not dry_run
    required_token = "APPLY_RECORD_MATCH_ACTION_PLAN"
    if apply_enabled and apply_token != required_token:
        raise TrelloError(f"Refusing Trello writes. Re-run with apply_token={required_token!r} to apply comments.")
    if apply_enabled:
        validate_trello_cards = True

    results = []
    errors = []
    for index, action in enumerate(actions, start=1):
        row = {
            "index": index,
            "card_id": action.get("card_id"),
            "card_name": action.get("card_name"),
            "card_url": action.get("card_url"),
            "confidence": action.get("confidence"),
            "record_path": action.get("record_path"),
            "idempotency_marker": _marker_for_action(action),
            "idempotency_marker_source": "plan" if action.get("idempotency_marker") else "computed",
            "validated": False,
            "applied": False,
            "planned_writes": [],
        }
        try:
            if not action.get("card_id"):
                raise TrelloError("Action missing card_id")
            marker = row["idempotency_marker"]
            if not marker:
                raise TrelloError("Action missing idempotency_marker")
            if validate_trello_cards:
                card = get_card(str(action["card_id"]))
                row["current_card_name"] = card.get("name")
                row["trello_card_validated"] = True
                row["duplicate_comment_found"] = _card_has_comment_marker(str(action["card_id"]), marker)
            else:
                row["current_card_name"] = None
                row["trello_card_validated"] = False
                row["trello_card_validation_skipped"] = True
                row["duplicate_comment_found"] = None
            record_path = Path(str(action.get("record_path") or "")).expanduser()
            row["record_exists"] = record_path.exists()
            if action.get("record_path") and not record_path.exists():
                raise TrelloError(f"Record path no longer exists: {record_path}")
            comment_text = _comment_text_with_marker(action, str(marker))
            row["planned_writes"].append({"type": "comment", "chars": len(comment_text)})
            row["planned_writes"].append(
                {
                    "type": "attachment_review",
                    "applied": False,
                    "note": "Attachments are not auto-created by this tool.",
                }
            )
            row["validated"] = True
            if apply_enabled:
                if row.get("duplicate_comment_found"):
                    row["applied"] = False
                    row["skipped"] = True
                    row["skip_reason"] = "duplicate_comment_marker_found"
                else:
                    comment = add_comment(str(action["card_id"]), comment_text)
                    row["applied"] = True
                    row["comment_action_id"] = comment.get("id")
        except Exception as exc:
            row["error"] = str(exc)[:500]
            errors.append(row)
        results.append(row)

    applied_count = sum(1 for item in results if item.get("applied"))
    validated_count = sum(1 for item in results if item.get("validated"))
    return {
        "ok": not errors,
        "mode": "dry_run" if dry_run else "apply",
        "plan_json_path": str(Path(plan_json_path).expanduser()),
        "summary": {
            "actions_seen": len(actions),
            "validated": validated_count,
            "applied_comments": applied_count,
            "errors": len(errors),
        },
        "results": results,
        "safety": {
            "dry_run": dry_run,
            "trello_comments_written": applied_count,
            "trello_attachments_written": 0,
            "requires_apply_token_for_writes": True,
            "required_apply_token": required_token,
            "validate_trello_cards": validate_trello_cards,
        },
    }


@mcp.tool()
def preview_stephen_bill(
    board: str = "memphis",
    list_name: str = "Jobs that I need to bill for",
    limit: int = 100,
    read_work_orders: bool = True,
    extra_card_queries: list[str] | None = None,
) -> dict[str, Any]:
    """Build Stephen's read-only billing preview from Trello jobs ready to bill."""
    command = [
        "/Users/stephengodman/CodeX/bin/trello-preview-bill",
        "--json",
        "--board",
        board,
        "--list",
        list_name,
        "--limit",
        str(max(1, min(int(limit), 500))),
    ]
    for query in extra_card_queries or []:
        command.extend(["--extra-card-query", query])
    if not read_work_orders:
        command.append("--no-work-orders")
    result = subprocess.run(command, check=False, capture_output=True, text=True, timeout=180)
    output = result.stdout.strip() or result.stderr.strip()
    try:
        payload = json.loads(output)
    except json.JSONDecodeError as exc:
        raise TrelloError(f"billing preview returned invalid JSON: {output[:600]}") from exc
    payload.setdefault("ok", result.returncode == 0)
    payload.setdefault("safety", {})
    payload["safety"].update(
        {
            "read_only": True,
            "trello_writes": False,
            "google_drive_writes": False,
            "pi5_writes": False,
            "local_report_files_written": not bool(payload.get("error")),
            "secrets_returned": False,
        }
    )
    return payload


@mcp.tool()
def draft_stephen_bill(
    board: str = "memphis",
    list_name: str = "Jobs that I need to bill for",
    limit: int = 100,
    include_review: bool = False,
    read_work_orders: bool = False,
    extra_card_queries: list[str] | None = None,
) -> dict[str, Any]:
    """Create a local PDF/CSV/Markdown draft bill from Trello jobs ready to bill."""
    command = [
        "/Users/stephengodman/CodeX/bin/trello-draft-bill",
        "--json",
        "--board",
        board,
        "--list",
        list_name,
        "--limit",
        str(max(1, min(int(limit), 500))),
    ]
    for query in extra_card_queries or []:
        command.extend(["--extra-card-query", query])
    if include_review:
        command.append("--include-review")
    if read_work_orders:
        command.append("--read-work-orders")
    result = subprocess.run(command, check=False, capture_output=True, text=True, timeout=240)
    output = result.stdout.strip() or result.stderr.strip()
    try:
        payload = json.loads(output)
    except json.JSONDecodeError as exc:
        raise TrelloError(f"bill draft returned invalid JSON: {output[:600]}") from exc
    payload.setdefault("ok", result.returncode == 0)
    payload.setdefault("safety", {})
    payload["safety"].update(
        {
            "trello_writes": False,
            "google_drive_writes": False,
            "pi5_writes": False,
            "local_bill_files_written": not bool(payload.get("error")),
            "secrets_returned": False,
        }
    )
    return payload


@mcp.tool()
def publish_stephen_bill(
    board: str = "memphis",
    bills_list: str = "BILLS",
    bill_json: str = "",
    apply_token: str = "",
    comment_job_cards: bool = False,
    move_to_list: str = "",
) -> dict[str, Any]:
    """Plan or publish a local Stephen bill draft to the Trello BILLS list.

    Default mode is read-only planning. Trello writes require
    apply_token="PUBLISH_STEPHEN_BILL".
    """
    command = [
        "/Users/stephengodman/CodeX/bin/trello-publish-bill",
        "--json",
        "--board",
        board,
        "--bills-list",
        bills_list,
    ]
    if bill_json:
        command.extend(["--bill-json", bill_json])
    if apply_token:
        command.extend(["--apply-token", apply_token])
    if comment_job_cards:
        command.append("--comment-job-cards")
    if move_to_list:
        command.extend(["--move-to-list", move_to_list])
    result = subprocess.run(command, check=False, capture_output=True, text=True, timeout=300)
    output = result.stdout.strip() or result.stderr.strip()
    try:
        payload = json.loads(output)
    except json.JSONDecodeError as exc:
        raise TrelloError(f"bill publish returned invalid JSON: {output[:600]}") from exc
    payload.setdefault("ok", result.returncode == 0)
    payload.setdefault("safety", {})
    payload["safety"].setdefault("secrets_returned", False)
    return payload


@mcp.tool()
def route_pool_stops(
    stop_queries: list[str],
    board: str = "memphis",
    warehouse_first: bool = True,
    manual_stops: list[str] | None = None,
    origin: str = "Current Location",
) -> dict[str, Any]:
    """Build a route from Trello pool job names plus optional manual stops.

    Manual stops use NAME=ADDRESS, for example:
    "Keathley pump drop=8925 Cedar Mills Cove, Cordova, TN".
    """
    if not stop_queries and not manual_stops:
        raise TrelloError("stop_queries or manual_stops is required")
    command = [
        "/Users/stephengodman/CodeX/bin/trello-route-stops",
        "--json",
        "--board",
        board,
        "--origin",
        origin,
    ]
    if warehouse_first:
        command.append("--warehouse-first")
    for item in manual_stops or []:
        command.extend(["--manual-stop", item])
    command.extend(stop_queries)
    result = subprocess.run(command, check=False, capture_output=True, text=True, timeout=180)
    output = result.stdout.strip() or result.stderr.strip()
    try:
        payload = json.loads(output)
    except json.JSONDecodeError as exc:
        raise TrelloError(f"route command returned invalid JSON: {output[:600]}") from exc
    payload.setdefault("ok", result.returncode == 0)
    payload.setdefault("safety", {})
    payload["safety"].update(
        {
            "trello_writes": False,
            "google_drive_writes": False,
            "pi5_writes": False,
            "secrets_returned": False,
        }
    )
    return payload


@mcp.tool()
def copy_board(source_board: str = "memphis", name: str | None = None, keep_from_source: str = "cards") -> dict[str, Any]:
    """Copy a Trello board. Defaults to copying Memphis Pool with cards."""
    source_id = _board_id(source_board)
    if not name:
        today = datetime.now().strftime("%Y-%m-%d")
        source_name = KNOWN_BOARDS.get(source_board, {}).get("name", source_board)
        name = f"{source_name} - Stephen Admin Copy {today}"
    return _request(
        "POST",
        "boards",
        params={
            "name": name,
            "idBoardSource": source_id,
            "keepFromSource": keep_from_source,
            "defaultLists": "false",
        },
    )


def _cli_bool(value: str) -> bool:
    return value.strip().lower() in {"1", "true", "yes", "on"}


def _run_cli(argv: list[str]) -> int:
    parser = argparse.ArgumentParser(description="Trello MCP local operator commands")
    parser.add_argument("--runtime-diagnostics", action="store_true", help="Print redacted runtime diagnostics and exit.")
    parser.add_argument("--include-live-health", action="store_true", help="Include a read-only Trello health call with diagnostics.")
    parser.add_argument("--write-cutover-packet", action="store_true", help="Write the local Memphis cutover packet and exit.")
    parser.add_argument("--search-cards", action="store_true", help="Search Trello cards by name/description and exit.")
    parser.add_argument("--read-work-order", action="store_true", help="Read a work-order-like attachment from a Trello card and exit.")
    parser.add_argument("--attach-file", action="store_true", help="Attach a local file to a Trello card and exit. This writes to Trello.")
    parser.add_argument("--set-cover", action="store_true", help="Set a Trello card cover attachment and exit. This writes to Trello.")
    parser.add_argument("--primary-board", default="memphis")
    parser.add_argument("--comparison-board", default="legacy_memphis")
    parser.add_argument("--proposed-action", default="archive_legacy_memphis")
    parser.add_argument("--board", default="memphis")
    parser.add_argument("--query", default=None)
    parser.add_argument("--card-id", default=None)
    parser.add_argument("--card-query", default=None)
    parser.add_argument("--file-path", default=None)
    parser.add_argument("--name", default=None)
    parser.add_argument("--comment", default=None)
    parser.add_argument("--attachment-id", default=None)
    parser.add_argument("--attachment-filter", default="")
    parser.add_argument("--question", default="")
    parser.add_argument("--max-text-chars", type=int, default=6000)
    parser.add_argument("--include-complete", default="true")
    parser.add_argument("--limit-cards", type=int, default=1000)
    parser.add_argument("--sample-cards-per-list", type=int, default=8)
    args = parser.parse_args(argv)

    try:
        if args.runtime_diagnostics:
            payload = trello_runtime_diagnostics(include_live_health=args.include_live_health)
        elif args.search_cards:
            if not args.query:
                raise TrelloError("--query is required with --search-cards")
            payload = {
                "ok": True,
                "board": args.board,
                "query": args.query,
                "cards": search_cards(query=args.query, board=args.board, limit=args.limit_cards),
                "safety": {
                    "read_only": True,
                    "trello_writes": False,
                    "google_drive_writes": False,
                    "pi5_writes": False,
                },
            }
        elif args.read_work_order:
            payload = read_card_work_order(
                card_id=args.card_id,
                card_query=args.card_query,
                board=args.board,
                attachment_filter=args.attachment_filter,
                question=args.question,
                max_text_chars=args.max_text_chars,
            )
        elif args.attach_file:
            if not args.file_path:
                raise TrelloError("--file-path is required with --attach-file")
            payload = attach_file_to_card(
                file_path=args.file_path,
                card_id=args.card_id,
                card_query=args.card_query,
                board=args.board,
                name=args.name,
                comment=args.comment,
            )
        elif args.set_cover:
            if not args.card_id or not args.attachment_id:
                raise TrelloError("--card-id and --attachment-id are required with --set-cover")
            payload = set_card_cover(card_id=args.card_id, attachment_id=args.attachment_id)
        elif args.write_cutover_packet:
            payload = write_cutover_packet(
                primary_board=args.primary_board,
                comparison_board=args.comparison_board,
                proposed_action=args.proposed_action,
                include_complete=_cli_bool(args.include_complete),
                limit_cards=args.limit_cards,
                sample_cards_per_list=args.sample_cards_per_list,
            )
        else:
            parser.error("choose --runtime-diagnostics, --search-cards, --read-work-order, --attach-file, --set-cover, or --write-cutover-packet")
            return 2
    except Exception as exc:
        payload = {
            "ok": False,
            "error_type": type(exc).__name__,
            "error": str(exc)[:800],
            "safety": {
                "secrets_returned": False,
                "trello_writes": False,
                "google_drive_writes": False,
                "pi5_writes": False,
            },
        }
        print(json.dumps(payload, indent=2, sort_keys=True))
        return 1

    print(json.dumps(payload, indent=2, sort_keys=True))
    return 0 if payload.get("ok") else 1


if __name__ == "__main__":
    if len(sys.argv) > 1:
        raise SystemExit(_run_cli(sys.argv[1:]))
    mcp.run()
